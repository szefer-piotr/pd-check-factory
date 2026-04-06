"""DM review Excel export and apply-import for pd_draft_specs."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pdcheck_factory.json_util import load_schema, read_json, validate, write_json


def _anchor_text(anchor: Dict[str, Any]) -> str:
    return f"{anchor.get('anchor_type', '')}: {anchor.get('anchor_description', '')}"


def _window_text(w: Dict[str, Any]) -> str:
    parts = [w.get("window_text", ""), w.get("window_type", "")]
    lb = w.get("lower_bound")
    ub = w.get("upper_bound")
    if lb is not None or ub is not None:
        parts.append(f"bounds=({lb},{ub})")
    u = w.get("unit")
    if u:
        parts.append(f"unit={u}")
    return " | ".join(str(p) for p in parts if p)


HEADERS = [
    "spec_id",
    "status",
    "deviation_title",
    "deviation_category",
    "protocol_rule_description",
    "candidate_trigger_condition",
    "timing_anchor",
    "allowed_window",
    "confidence",
    "ambiguity_flag",
    "reviewer_notes",
    "dm_decision",
    "dm_comments",
    "proposed_text",
]


def export_dm_workbook(
    *,
    pd_specs_path: Path,
    output_path: Path,
) -> Path:
    data = read_json(pd_specs_path)
    specs = data.get("pd_draft_specs", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "DM Review"
    ws.append(HEADERS)

    for spec in specs:
        ws.append(
            [
                spec.get("spec_id", ""),
                spec.get("status", "draft"),
                spec.get("deviation_title", ""),
                spec.get("deviation_category", ""),
                spec.get("protocol_rule_description", ""),
                spec.get("candidate_trigger_condition", ""),
                _anchor_text(spec.get("timing_anchor", {})),
                _window_text(spec.get("allowed_window", {})),
                spec.get("confidence", ""),
                spec.get("ambiguity_flag", ""),
                spec.get("reviewer_notes", ""),
                "",  # dm_decision
                "",  # dm_comments
                "",  # proposed_text
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _read_header_map(ws: Worksheet) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    m: Dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        key = _norm_header(str(cell))
        if key:
            m[key] = idx
    return m


def _get_cell(row: Tuple[Any, ...], colmap: Dict[str, int], *names: str) -> str:
    for n in names:
        idx = colmap.get(n)
        if idx is None:
            continue
        if idx >= len(row):
            continue
        v = row[idx]
        if v is None:
            return ""
        return str(v).strip()
    return ""


def _map_decision(raw: str) -> Optional[str]:
    r = (raw or "").strip().lower()
    if not r:
        return None
    if r in ("approve", "approved", "accept", "yes"):
        return "approved"
    if r in ("reject", "rejected", "no"):
        return "rejected"
    if r in ("revise", "reviewed", "needs_revision", "pending"):
        return "reviewed"
    return None


def apply_dm_workbook(
    *,
    pd_specs_path: Path,
    workbook_path: Path,
    output_specs_path: Path,
) -> Path:
    payload = read_json(pd_specs_path)
    schema = load_schema("pd_draft_spec.schema.json")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active
    colmap = _read_header_map(ws)

    updates_by_id: Dict[str, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        sid = _get_cell(row, colmap, "spec_id")
        if not sid:
            continue
        updates_by_id[sid] = {
            "dm_decision": _get_cell(row, colmap, "dm_decision"),
            "dm_comments": _get_cell(row, colmap, "dm_comments"),
            "proposed_text": _get_cell(row, colmap, "proposed_text"),
        }

    specs: List[Dict[str, Any]] = list(payload.get("pd_draft_specs", []))
    for spec in specs:
        sid = spec.get("spec_id")
        if sid not in updates_by_id:
            continue
        u = updates_by_id[sid]
        decision = _map_decision(u["dm_decision"])
        if decision:
            spec["status"] = decision
        notes = (spec.get("reviewer_notes") or "").strip()
        if u["dm_comments"]:
            notes = (notes + "\n" + u["dm_comments"]).strip()
            spec["reviewer_notes"] = notes
        if u["proposed_text"]:
            spec["candidate_trigger_condition"] = u["proposed_text"]

    payload["pd_draft_specs"] = specs
    payload["generated_at"] = datetime.now(timezone.utc).isoformat()

    errs = validate(payload, schema)
    if errs:
        raise ValueError("Updated specs failed validation: " + "; ".join(errs[:15]))

    write_json(output_specs_path, payload)
    return output_specs_path
