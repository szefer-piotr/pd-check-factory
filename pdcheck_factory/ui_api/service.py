from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import uuid
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

from pdcheck_factory import (
    blob_io,
    coding_workbook_export,
    extraction_resolve,
    paths,
    pipeline_v2,
    review_sources,
    study_artifact_sync,
)
from pdcheck_factory.json_util import read_json, write_json
from pdcheck_factory.deviation_contract import (
    has_flat_pd_spec_fields,
    lift_pd_spec_row,
    pd_spec_field,
    split_pd_spec_row,
)
from pdcheck_factory.pd_spec_import import parse_pd_spec_xlsx, programmable_from_manual_or_programmable


class UiApiError(Exception):
    def __init__(self, code: str, message: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code


PROCESSING_CORE_STEP_IDS: List[str] = [
    "extract-inputs",
    "index-protocol",
    "acrf-split-toc",
    "acrf-summary-text",
]

PROCESSING_BACKEND_STEP_IDS: List[str] = [
    "extract-inputs",
    "index-protocol",
    "acrf-split-toc",
    "acrf-summary-text",
    "extract-rules",
    "extract-deviations",
]

STEP_ORDER: List[str] = [
    "extract-inputs",
    "index-protocol",
    "acrf-split-toc",
    "acrf-summary-text",
    "extract-rules",
    "extract-deviations",
    "import-pd-spec-ground",
    "import-pd-spec-map",
    "import-pd-spec-enrich",
    "merge-pd-spec-imports",
    "review-and-finalize",
]

STEP_DEPENDENCIES: Dict[str, List[str]] = {
    "extract-inputs": [],
    "index-protocol": ["extract-inputs"],
    "acrf-split-toc": ["extract-inputs"],
    "acrf-summary-text": ["acrf-split-toc"],
    "extract-rules": ["index-protocol"],
    "extract-deviations": ["extract-rules", "acrf-summary-text"],
    "import-pd-spec-ground": ["index-protocol", "acrf-summary-text"],
    "import-pd-spec-map": [],
    "import-pd-spec-enrich": ["index-protocol", "acrf-summary-text"],
    "merge-pd-spec-imports": ["import-pd-spec-ground"],
    "review-and-finalize": [],
}

ENTRY_MODE_EXTRACTED = "extracted"
ENTRY_MODE_IMPORTED_PD_SPEC = "imported_pd_spec"

WORKFLOW_CHOICE_EXTRACT = "extract"
WORKFLOW_CHOICE_MAP = "map"
WORKFLOW_CHOICE_ENRICH = "enrich"
VALID_WORKFLOW_CHOICES = {WORKFLOW_CHOICE_EXTRACT, WORKFLOW_CHOICE_MAP, WORKFLOW_CHOICE_ENRICH}

UI_STAGE_PROJECT = "project"
UI_STAGE_SETUP = "setup"
UI_STAGE_SUMMARY = "summary"
UI_STAGE_PROCESSING = "processing"
UI_STAGE_REVIEW = "review"
VALID_UI_STAGES = {
    UI_STAGE_PROJECT,
    UI_STAGE_SETUP,
    UI_STAGE_SUMMARY,
    UI_STAGE_PROCESSING,
    UI_STAGE_REVIEW,
}

WORKFLOW_STEPS: Dict[str, List[str]] = {
    WORKFLOW_CHOICE_EXTRACT: [
        "extract-inputs",
        "index-protocol",
        "acrf-split-toc",
        "acrf-summary-text",
        "extract-rules",
        "extract-deviations",
        "review-and-finalize",
    ],
    WORKFLOW_CHOICE_MAP: [
        "import-pd-spec-map",
        "merge-pd-spec-imports",
        "review-and-finalize",
    ],
    WORKFLOW_CHOICE_ENRICH: [
        "extract-inputs",
        "index-protocol",
        "acrf-split-toc",
        "acrf-summary-text",
        "import-pd-spec-enrich",
        "merge-pd-spec-imports",
        "review-and-finalize",
    ],
}

IMPORT_STEP_ORDER: List[str] = [
    "extract-inputs",
    "index-protocol",
    "acrf-split-toc",
    "acrf-summary-text",
    "import-pd-spec-ground",
    "import-pd-spec-map",
    "import-pd-spec-enrich",
    "merge-pd-spec-imports",
    "review-and-finalize",
]

IMPORT_STEP_DEPENDENCIES: Dict[str, List[str]] = {
    "extract-inputs": [],
    "index-protocol": ["extract-inputs"],
    "acrf-split-toc": ["extract-inputs"],
    "acrf-summary-text": ["acrf-split-toc"],
    "import-pd-spec-ground": ["index-protocol", "acrf-summary-text"],
    "import-pd-spec-map": [],
    "import-pd-spec-enrich": ["index-protocol", "acrf-summary-text"],
    "merge-pd-spec-imports": ["import-pd-spec-ground"],
    "review-and-finalize": ["import-pd-spec-ground"],
}

STEP7_EXPORT_COLUMNS: List[str] = [
    "study_id",
    "exported_at",
    "rule_id",
    "rule_title",
    "rule_text",
    "rule_paragraph_refs",
    "deviation_id",
    "deviation_text",
    "paragraph_refs",
    "supporting_sentences",
    "data_support_note",
    "status",
    "dm_comment",
    "entry_source",
    "programmable",
    "programmability_note",
    "pseudo_logic",
]


@dataclass(frozen=True)
class StudyPaths:
    protocol_source: Path
    acrf_source: Path
    paragraph_index: Path
    acrf_sections_toc_dir: Path
    acrf_summary_text_merged: Path
    rules_parsed: Path
    deviations_parsed: Path
    deviations_review_state: Path
    deviations_validated: Path
    pseudo_logic_validated: Path
    final_json: Path
    final_xlsx: Path


@dataclass
class UiStepService:
    output_dir: Path

    def _mirror_upload(self, study_id: str, *local_paths: Path) -> None:
        for p in local_paths:
            study_artifact_sync.mirror_upload_path(study_id, self.output_dir, p)

    def _study_paths(self, study_id: str) -> StudyPaths:
        proto = extraction_resolve.resolve_protocol_rendered_source_md(study_id, self.output_dir)
        acrf = extraction_resolve.resolve_acrf_rendered_source_md(study_id, self.output_dir)
        sections_toc = extraction_resolve.resolve_acrf_sections_toc_dir(study_id, self.output_dir)
        return StudyPaths(
            protocol_source=proto,
            acrf_source=acrf,
            paragraph_index=paths.local_protocol_paragraph_index_json(study_id, self.output_dir),
            acrf_sections_toc_dir=sections_toc,
            acrf_summary_text_merged=paths.local_acrf_summary_text_merged(study_id, self.output_dir),
            rules_parsed=paths.local_rules_parsed_json(study_id, self.output_dir),
            deviations_parsed=paths.local_deviations_parsed_json(study_id, self.output_dir),
            deviations_review_state=paths.local_deviations_review_state(study_id, self.output_dir),
            deviations_validated=paths.local_deviations_validated_json(study_id, self.output_dir),
            pseudo_logic_validated=paths.local_pseudo_logic_validated_json(study_id, self.output_dir),
            final_json=paths.local_final_deviations_json(study_id, self.output_dir),
            final_xlsx=paths.local_final_deviations_xlsx(study_id, self.output_dir),
        )

    def _require_study_id(self, study_id: str) -> str:
        normalized = (study_id or "").strip()
        if not normalized:
            raise UiApiError("VALIDATION_ERROR", "studyId is required", 400)
        return normalized

    def _assert_safe_study_id(self, study_id: str) -> None:
        if "/" in study_id or "\\" in study_id or ".." in study_id:
            raise UiApiError(
                "VALIDATION_ERROR",
                "studyId must not contain path separators or '..'",
                400,
            )

    def _get_entry_mode(self, study_id: str) -> str:
        manifest = self._read_upload_manifest_obj(study_id)
        mode = str(manifest.get("entryMode") or ENTRY_MODE_EXTRACTED).strip()
        if mode not in {ENTRY_MODE_EXTRACTED, ENTRY_MODE_IMPORTED_PD_SPEC}:
            return ENTRY_MODE_EXTRACTED
        return mode

    def _effective_step_order(self, study_id: str) -> List[str]:
        return list(STEP_ORDER)

    def _effective_step_dependencies(self, study_id: str) -> Dict[str, List[str]]:
        return dict(STEP_DEPENDENCIES)

    def _assert_step_dependencies(self, statuses: Dict[str, str], step_id: str, study_id: str) -> None:
        if step_id == "review-and-finalize":
            extract_done = statuses.get("extract-deviations") in {"done", "skipped"}
            import_done = any(
                statuses.get(step) in {"done", "skipped"}
                for step in (
                    "import-pd-spec-ground",
                    "import-pd-spec-map",
                    "import-pd-spec-enrich",
                )
            )
            if not (extract_done or import_done):
                raise UiApiError(
                    "STEP_BLOCKED",
                    "Step 'review-and-finalize' is blocked. Complete deviation extraction or PD spec import first.",
                    409,
                )
            return
        for dependency in self._effective_step_dependencies(study_id).get(step_id, []):
            dep_status = statuses.get(dependency)
            if dep_status not in {"done", "skipped"}:
                raise UiApiError(
                    "STEP_BLOCKED",
                    f"Step '{step_id}' is blocked. Complete '{dependency}' first.",
                    409,
                )

    def _has_import_snapshot(self, study_id: str) -> bool:
        review_dir = paths.local_review_dir(study_id, self.output_dir)
        return review_dir.exists() and any(review_dir.glob("deviations_import_*.json"))

    def _has_merged_snapshot(self, study_id: str) -> bool:
        review_dir = paths.local_review_dir(study_id, self.output_dir)
        return review_dir.exists() and any(review_dir.glob("deviations_merged_*.json"))

    def _pd_spec_map_done(self, study_id: str) -> bool:
        manifest = self._read_upload_manifest_obj(study_id)
        mode = str(manifest.get("pdSpecImportMode") or "")
        return mode in {"map", "enrich_stub", "enrich"} or self._has_import_snapshot(study_id)

    def _pd_spec_enrich_done(self, study_id: str) -> bool:
        manifest = self._read_upload_manifest_obj(study_id)
        mode = str(manifest.get("pdSpecImportMode") or "")
        if mode == "enrich":
            return True
        enriched_path = paths.local_deviations_review_enriched_pd_spec_json(study_id, self.output_dir)
        if enriched_path.is_file():
            try:
                obj = read_json(enriched_path)
                return str(obj.get("pd_spec_import_mode") or "") == "enrich"
            except (OSError, ValueError):
                pass
        return False

    def _step_artifact_complete(self, study_id: str, step_id: str) -> bool:
        p = self._study_paths(study_id)
        if step_id == "extract-inputs":
            return p.protocol_source.exists() and p.acrf_source.exists()
        if step_id == "index-protocol":
            return p.paragraph_index.exists()
        if step_id == "acrf-split-toc":
            return p.acrf_sections_toc_dir.exists() and any(p.acrf_sections_toc_dir.glob("*.md"))
        if step_id == "acrf-summary-text":
            return p.acrf_summary_text_merged.exists()
        if step_id == "extract-rules":
            return p.rules_parsed.exists()
        if step_id == "extract-deviations":
            if not p.deviations_parsed.exists():
                return False
            try:
                parsed_obj = read_json(p.deviations_parsed)
                if parsed_obj.get("partial") is True:
                    return False
            except (json.JSONDecodeError, OSError, ValueError, TypeError):
                return False
            return p.deviations_review_state.exists()
        if step_id == "import-pd-spec-ground":
            return self._has_import_snapshot(study_id)
        if step_id == "import-pd-spec-map":
            return self._pd_spec_map_done(study_id)
        if step_id == "import-pd-spec-enrich":
            return self._pd_spec_enrich_done(study_id)
        if step_id == "merge-pd-spec-imports":
            return self._has_merged_snapshot(study_id)
        if step_id == "review-and-finalize":
            return p.final_json.exists() and p.final_xlsx.exists()
        return False

    def _processing_core_complete(self, study_id: str) -> bool:
        return all(self._step_artifact_complete(study_id, step_id) for step_id in PROCESSING_CORE_STEP_IDS)

    def _processing_complete(self, study_id: str) -> bool:
        return all(self._step_artifact_complete(study_id, step_id) for step_id in PROCESSING_BACKEND_STEP_IDS)

    def _step_statuses(self, study_id: str) -> Dict[str, str]:
        statuses: Dict[str, str] = {
            step_id: "done" if self._step_artifact_complete(study_id, step_id) else "pending"
            for step_id in STEP_ORDER
        }
        versions = pipeline_v2.list_import_versions(study_id, self.output_dir)
        if len(versions.get("imports", [])) < 2:
            statuses["merge-pd-spec-imports"] = "skipped"
        return statuses

    def _read_excerpt(self, file_path: Path, max_chars: int = 2500) -> str:
        if not file_path.exists() or not file_path.is_file():
            return ""
        return file_path.read_text(encoding="utf-8")[:max_chars]

    def _clear_deviation_extraction_artifacts(self, study_id: str) -> None:
        p = self._study_paths(study_id)
        for artifact in (
            p.deviations_parsed,
            p.deviations_review_state,
            p.deviations_validated,
            paths.local_deviations_raw_txt(study_id, self.output_dir),
            paths.local_deviations_review_generated_json(study_id, self.output_dir),
        ):
            if artifact.is_file():
                artifact.unlink()

    def get_extraction_live(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        p = self._study_paths(study_id)
        run_state = self._read_pipeline_run_state(study_id)
        run_status = str(run_state.get("status", "idle"))
        current_sub_step = str(run_state.get("currentSubStepId", ""))
        is_running = run_status == "running"
        llm_progress = run_state.get("llmProgress")
        llm_progress_out = llm_progress if isinstance(llm_progress, dict) else None

        rules: List[Dict[str, Any]] = []
        if p.rules_parsed.is_file():
            try:
                rules_obj = read_json(p.rules_parsed)
                for rule in rules_obj.get("rules", []):
                    if not isinstance(rule, dict):
                        continue
                    rules.append(
                        {
                            "rule_id": str(rule.get("rule_id", "")),
                            "title": str(rule.get("title", "")),
                            "text": str(rule.get("text", "")),
                            "paragraph_refs": list(rule.get("paragraph_refs", [])),
                        }
                    )
            except (json.JSONDecodeError, OSError, ValueError, TypeError):
                rules = []

        deviations: List[Dict[str, Any]] = []
        completed_rule_ids: List[str] = []
        file_partial = False
        if p.deviations_parsed.is_file():
            try:
                dev_obj = read_json(p.deviations_parsed)
                file_partial = dev_obj.get("partial") is True
                completed_rule_ids = [
                    str(rule_id)
                    for rule_id in dev_obj.get("completed_rule_ids", [])
                    if str(rule_id).strip()
                ]
                for dev in dev_obj.get("deviations", []):
                    if not isinstance(dev, dict):
                        continue
                    deviations.append(
                        {
                            "deviation_id": str(dev.get("deviation_id", "")),
                            "rule_id": str(dev.get("rule_id", "")),
                            "text": str(dev.get("text", "")),
                            "paragraph_refs": list(dev.get("paragraph_refs", [])),
                            "data_support_note": str(dev.get("data_support_note", "")),
                            "status": str(dev.get("status", "pending")),
                        }
                    )
            except (json.JSONDecodeError, OSError, ValueError, TypeError):
                deviations = []
                completed_rule_ids = []
                file_partial = False

        partial = file_partial or (
            is_running and current_sub_step in {"extract-rules", "extract-deviations"}
        )

        return {
            "studyId": study_id,
            "rules": rules,
            "deviations": deviations,
            "ruleCount": len(rules),
            "deviationCount": len(deviations),
            "partial": partial,
            "completedRuleIds": completed_rule_ids,
            "llmProgress": llm_progress_out,
            "runStatus": run_status,
        }

    def _ui_upload_manifest_path(self, study_id: str) -> Path:
        return paths.local_ui_upload_manifest(study_id, self.output_dir)

    def _read_upload_filenames(self, study_id: str) -> Dict[str, str]:
        obj = self._read_upload_manifest_obj(study_id)
        protocol = str(obj.get("protocolFileName") or "").strip()
        acrf = str(obj.get("acrfFileName") or "").strip()
        return {
            "protocolFileName": protocol or "protocol.pdf",
            "acrfFileName": acrf or "acrf.pdf",
        }

    def _sanitize_reference_filename(self, file_name: str) -> str:
        base = Path(file_name).name.strip()
        safe = re.sub(r"[^\w.\- ()]", "_", base)
        return safe or "document.pdf"

    def _read_upload_manifest_obj(self, study_id: str) -> Dict[str, Any]:
        manifest_path = self._ui_upload_manifest_path(study_id)
        if manifest_path.is_file():
            return read_json(manifest_path)
        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            blob_path = paths.ui_upload_manifest_blob(study_id)
            if blob_io.blob_exists(
                blob_service=blob_service,
                container_name=container,
                blob_path=blob_path,
            ):
                raw = blob_io.download_blob_bytes(
                    blob_service=blob_service,
                    container_name=container,
                    blob_path=blob_path,
                )
                return json.loads(raw.decode("utf-8"))
        except Exception:  # noqa: BLE001
            pass
        return {}

    def _write_upload_manifest(
        self,
        study_id: str,
        *,
        protocol_file_name: str | None = None,
        acrf_file_name: str | None = None,
        protocol_size: int | None = None,
        acrf_size: int | None = None,
        entry_mode: str | None = None,
        active_deviations_source: str | None = None,
        pd_spec_file_name: str | None = None,
        pd_spec_size: int | None = None,
        pd_spec_import_mode: str | None = None,
        coding_phase_accepted: bool | None = None,
        review_display_source: str | None = None,
        workflow_choice: str | None = None,
        ui_stage: str | None = None,
        clear_pd_spec_import_mode: bool = False,
    ) -> Dict[str, Any]:
        existing = self._read_upload_manifest_obj(study_id)
        resolved_pd_spec_import_mode = (
            None
            if clear_pd_spec_import_mode
            else (
                pd_spec_import_mode
                if pd_spec_import_mode is not None
                else existing.get("pdSpecImportMode")
            )
        )
        manifest = {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "protocolFileName": protocol_file_name
            or existing.get("protocolFileName")
            or "protocol.pdf",
            "acrfFileName": acrf_file_name or existing.get("acrfFileName") or "acrf.pdf",
            "uploadedAt": datetime.now(timezone.utc).isoformat(),
            "entryMode": entry_mode or existing.get("entryMode") or ENTRY_MODE_EXTRACTED,
            "activeDeviationsSource": active_deviations_source
            if active_deviations_source is not None
            else existing.get("activeDeviationsSource"),
            "pdSpecFileName": pd_spec_file_name or existing.get("pdSpecFileName"),
            "pdSpecImportMode": resolved_pd_spec_import_mode,
            "codingPhaseAccepted": coding_phase_accepted
            if coding_phase_accepted is not None
            else existing.get("codingPhaseAccepted", False),
            "reviewDisplaySource": review_display_source
            if review_display_source is not None
            else existing.get("reviewDisplaySource"),
            "workflowChoice": workflow_choice
            if workflow_choice is not None
            else existing.get("workflowChoice"),
            "uiStage": ui_stage if ui_stage is not None else existing.get("uiStage"),
        }
        if coding_phase_accepted:
            manifest["codingPhaseAcceptedAt"] = datetime.now(timezone.utc).isoformat()
        elif "codingPhaseAcceptedAt" in existing:
            manifest["codingPhaseAcceptedAt"] = existing["codingPhaseAcceptedAt"]
        if protocol_size is not None:
            manifest["protocolSize"] = protocol_size
        elif "protocolSize" in existing:
            manifest["protocolSize"] = existing["protocolSize"]
        if acrf_size is not None:
            manifest["acrfSize"] = acrf_size
        elif "acrfSize" in existing:
            manifest["acrfSize"] = existing["acrfSize"]
        if pd_spec_size is not None:
            manifest["pdSpecSize"] = pd_spec_size
        elif "pdSpecSize" in existing:
            manifest["pdSpecSize"] = existing["pdSpecSize"]

        manifest_path = self._ui_upload_manifest_path(study_id)
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        write_json(manifest_path, manifest)
        self._mirror_upload(study_id, manifest_path)

        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            blob_io.upload_blob_bytes(
                blob_service=blob_service,
                container_name=container,
                blob_path=paths.ui_upload_manifest_blob(study_id),
                data=json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"),
                content_type="application/json",
            )
        except Exception:  # noqa: BLE001
            pass
        return manifest

    def _blob_has_upload(self, study_id: str, role: str) -> bool:
        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            blob_path = paths.raw_protocol_blob(study_id) if role == "protocol" else paths.raw_acrf_blob(study_id)
            return blob_io.blob_exists(
                blob_service=blob_service,
                container_name=container,
                blob_path=blob_path,
            )
        except Exception:  # noqa: BLE001
            return False

    def _upload_reference_copy(self, study_id: str, role: str, data: bytes, file_name: str) -> str:
        safe_name = self._sanitize_reference_filename(file_name)
        blob_path = (
            paths.raw_protocol_reference_blob(study_id, safe_name)
            if role == "protocol"
            else paths.raw_acrf_reference_blob(study_id, safe_name)
        )
        blob_service = blob_io.blob_service_from_env()
        container = blob_io.container_from_env()
        blob_io.upload_blob_bytes(
            blob_service=blob_service,
            container_name=container,
            blob_path=blob_path,
            data=data,
            content_type="application/pdf",
        )
        return blob_path

    def _blob_has_pd_spec_workbook(self, study_id: str) -> bool:
        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            return blob_io.blob_exists(
                blob_service=blob_service,
                container_name=container,
                blob_path=paths.pd_spec_workbook_blob(study_id),
            )
        except Exception:  # noqa: BLE001
            return False

    def _read_pd_spec_workbook_bytes(self, study_id: str) -> bytes | None:
        """Return workbook bytes from local cache, downloading from blob when needed."""
        local_path = paths.local_pd_spec_workbook(study_id, self.output_dir)
        if local_path.is_file():
            return local_path.read_bytes()
        if not self._blob_has_pd_spec_workbook(study_id):
            return None
        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            data = blob_io.download_blob_bytes(
                blob_service=blob_service,
                container_name=container,
                blob_path=paths.pd_spec_workbook_blob(study_id),
            )
        except Exception:  # noqa: BLE001
            return None
        local_path.parent.mkdir(parents=True, exist_ok=True)
        local_path.write_bytes(data)
        self._mirror_upload(study_id, local_path)
        return data

    def get_step1_upload_status(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        manifest = self._read_upload_manifest_obj(study_id)
        protocol_uploaded = self._blob_has_upload(study_id, "protocol")
        acrf_uploaded = self._blob_has_upload(study_id, "acrf")
        pd_spec_uploaded = self._blob_has_pd_spec_workbook(study_id)
        if pd_spec_uploaded:
            self._read_pd_spec_workbook_bytes(study_id)

        def slot(role: str, uploaded: bool) -> Dict[str, Any]:
            if role == "pdSpec":
                return {
                    "uploaded": uploaded,
                    "fileName": str(manifest.get("pdSpecFileName") or "pd_specifications.xlsx"),
                    "size": int(manifest.get("pdSpecSize") or 0) if uploaded else 0,
                    "blob": paths.pd_spec_workbook_blob(study_id),
                }
            name_key = "protocolFileName" if role == "protocol" else "acrfFileName"
            size_key = "protocolSize" if role == "protocol" else "acrfSize"
            default_name = "protocol.pdf" if role == "protocol" else "acrf.pdf"
            return {
                "uploaded": uploaded,
                "fileName": str(manifest.get(name_key) or default_name),
                "size": int(manifest.get(size_key) or 0) if uploaded else 0,
                "blob": paths.raw_protocol_blob(study_id) if role == "protocol" else paths.raw_acrf_blob(study_id),
            }

        p = self._study_paths(study_id)
        return {
            "studyId": study_id,
            "protocol": slot("protocol", protocol_uploaded),
            "acrf": slot("acrf", acrf_uploaded),
            "pdSpec": slot("pdSpec", pd_spec_uploaded),
            "bothUploaded": protocol_uploaded and acrf_uploaded,
            "allThreeUploaded": protocol_uploaded and acrf_uploaded and pd_spec_uploaded,
            "protocolPreprocessed": p.paragraph_index.exists(),
            "acrfPreprocessed": p.acrf_summary_text_merged.exists(),
            "processingCoreComplete": self._processing_core_complete(study_id),
            "processingComplete": self._processing_complete(study_id),
            "stepStatuses": self._step_statuses(study_id),
        }

    def _assert_protocol_upload_ready(self, study_id: str) -> None:
        if not self._blob_has_upload(study_id, "protocol"):
            raise UiApiError(
                "UPLOAD_REQUIRED",
                "Upload the protocol PDF before preprocessing.",
                409,
            )

    def _assert_acrf_upload_ready(self, study_id: str) -> None:
        if not self._blob_has_upload(study_id, "acrf"):
            raise UiApiError(
                "UPLOAD_REQUIRED",
                "Upload the aCRF PDF before preprocessing.",
                409,
            )

    def _assert_both_uploads_ready(self, study_id: str) -> None:
        status = self.get_step1_upload_status(study_id)
        if not status["bothUploaded"]:
            raise UiApiError(
                "UPLOAD_REQUIRED",
                "Upload both protocol and aCRF PDFs before running extraction.",
                409,
            )

    def _pipeline_run_state_path(self, study_id: str) -> Path:
        return paths.local_ui_pipeline_run_state(study_id, self.output_dir)

    def _read_pipeline_run_state(self, study_id: str) -> Dict[str, Any]:
        path = self._pipeline_run_state_path(study_id)
        if path.is_file():
            return read_json(path)
        return {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "status": "idle",
            "currentStage": "",
            "currentSubStepId": "",
            "message": "",
            "error": "",
            "startedAt": "",
            "finishedAt": "",
            "logs": [],
        }

    def _append_pipeline_log(
        self,
        study_id: str,
        text: str,
        *,
        level: str = "info",
    ) -> None:
        state = self._read_pipeline_run_state(study_id)
        logs = list(state.get("logs", []))
        logs.append(
            {
                "ts": datetime.now(timezone.utc).isoformat(),
                "level": level,
                "text": text,
            }
        )
        state["logs"] = logs[-500:]
        path = self._pipeline_run_state_path(study_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        write_json(path, state)
        self._mirror_upload(study_id, path)

    def _write_pipeline_run_state(self, study_id: str, **updates: Any) -> Dict[str, Any]:
        state = self._read_pipeline_run_state(study_id)
        state.update(updates)
        path = self._pipeline_run_state_path(study_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        write_json(path, state)
        self._mirror_upload(study_id, path)
        return state

    def _report_llm_progress(
        self,
        study_id: str,
        *,
        phase: str,
        current: int,
        total: int,
        unit: str,
        label: str = "",
    ) -> None:
        self._write_pipeline_run_state(
            study_id,
            llmProgress={
                "phase": phase,
                "current": current,
                "total": total,
                "unit": unit,
                "label": label,
            },
        )
        self._append_pipeline_log(
            study_id,
            f"llm:{phase}:{current}/{total}:{label or unit}",
        )

    def _make_llm_progress_callback(self, study_id: str):
        def callback(*, phase: str, current: int, total: int, unit: str, label: str = "") -> None:
            self._report_llm_progress(
                study_id,
                phase=phase,
                current=current,
                total=total,
                unit=unit,
                label=label,
            )

        return callback

    def _run_manifest_path(self, study_id: str) -> Path:
        return paths.local_ui_run_manifest(study_id, self.output_dir)

    def _read_run_manifest(self, study_id: str) -> Dict[str, Any]:
        path = self._run_manifest_path(study_id)
        if path.is_file():
            return read_json(path)
        return {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "activeRunId": "",
            "runs": [],
        }

    def _write_run_manifest(self, study_id: str, manifest: Dict[str, Any]) -> Dict[str, Any]:
        path = self._run_manifest_path(study_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        write_json(path, manifest)
        self._mirror_upload(study_id, path)
        return manifest

    @staticmethod
    def _run_fingerprint(
        *,
        workflow: str,
        uploads: Dict[str, Any],
        settings: Dict[str, Any],
    ) -> str:
        parts = [
            workflow,
            str(uploads.get("protocolFileName") or ""),
            str(uploads.get("acrfFileName") or ""),
            str(uploads.get("pdSpecFileName") or ""),
            str(settings.get("extractorChoice") or ""),
            str(settings.get("extractionDeployment") or ""),
            str(settings.get("acrfSummaryDeployment") or ""),
            str(settings.get("extractionLlmInstructions") or ""),
        ]
        payload = "|".join(parts)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _active_run_entry(self, study_id: str) -> Dict[str, Any] | None:
        manifest = self._read_run_manifest(study_id)
        active_id = str(manifest.get("activeRunId") or "").strip()
        if not active_id:
            return None
        for entry in manifest.get("runs", []):
            if str(entry.get("runId")) == active_id:
                return entry
        return None

    def get_study_runs(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        manifest = self._read_run_manifest(study_id)
        return {
            "studyId": study_id,
            "activeRunId": manifest.get("activeRunId") or "",
            "runs": list(manifest.get("runs", [])),
        }

    def apply_study_run(self, study_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        workflow = str(payload.get("workflow") or "").strip()
        if workflow not in VALID_WORKFLOW_CHOICES:
            raise UiApiError(
                "VALIDATION_ERROR",
                f"workflow must be one of: {', '.join(sorted(VALID_WORKFLOW_CHOICES))}",
                400,
            )
        uploads = payload.get("uploads") or {}
        settings = payload.get("settings") or {}
        fingerprint = self._run_fingerprint(workflow=workflow, uploads=uploads, settings=settings)
        now = datetime.now(timezone.utc).isoformat()
        manifest = self._read_run_manifest(study_id)
        runs: List[Dict[str, Any]] = list(manifest.get("runs", []))
        existing = next((r for r in runs if str(r.get("fingerprint")) == fingerprint), None)
        if existing:
            existing["updatedAt"] = now
            existing["workflow"] = workflow
            existing["uploads"] = uploads
            existing["settings"] = settings
            run_id = str(existing["runId"])
        else:
            run_id = f"run-{uuid.uuid4().hex[:12]}"
            runs.append(
                {
                    "runId": run_id,
                    "fingerprint": fingerprint,
                    "createdAt": now,
                    "updatedAt": now,
                    "workflow": workflow,
                    "uploads": uploads,
                    "settings": settings,
                    "lastRunAt": None,
                    "stepStatusesSnapshot": {},
                }
            )
        manifest["activeRunId"] = run_id
        manifest["runs"] = runs
        self._write_run_manifest(study_id, manifest)
        active = next(r for r in runs if str(r.get("runId")) == run_id)
        return {
            "studyId": study_id,
            "runId": run_id,
            "fingerprint": fingerprint,
            "created": existing is None,
            "settings": active.get("settings", {}),
            "activeRunId": run_id,
            "runs": runs,
        }

    def activate_study_run(self, study_id: str, run_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        run_id = str(run_id).strip()
        if not run_id:
            raise UiApiError("VALIDATION_ERROR", "runId is required", 400)
        manifest = self._read_run_manifest(study_id)
        runs: List[Dict[str, Any]] = list(manifest.get("runs", []))
        active = next((r for r in runs if str(r.get("runId")) == run_id), None)
        if active is None:
            raise UiApiError("NOT_FOUND", f"Run '{run_id}' not found.", 404)
        manifest["activeRunId"] = run_id
        active["updatedAt"] = datetime.now(timezone.utc).isoformat()
        self._write_run_manifest(study_id, manifest)
        return {
            "studyId": study_id,
            "activeRunId": run_id,
            "settings": active.get("settings", {}),
            "run": active,
        }

    def _touch_active_run_after_step(self, study_id: str) -> None:
        active = self._active_run_entry(study_id)
        if not active:
            return
        now = datetime.now(timezone.utc).isoformat()
        active["lastRunAt"] = now
        active["stepStatusesSnapshot"] = self._step_statuses(study_id)
        active["updatedAt"] = now
        manifest = self._read_run_manifest(study_id)
        run_id = str(active.get("runId"))
        runs = []
        for entry in manifest.get("runs", []):
            if str(entry.get("runId")) == run_id:
                runs.append(active)
            else:
                runs.append(entry)
        manifest["runs"] = runs
        self._write_run_manifest(study_id, manifest)

    def get_step1_run_state(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        state = self._read_pipeline_run_state(study_id)
        stale_hours = 2
        if state.get("status") == "running" and state.get("startedAt"):
            try:
                started = datetime.fromisoformat(str(state["startedAt"]))
                age = datetime.now(timezone.utc) - started.replace(tzinfo=timezone.utc)
                if age.total_seconds() > stale_hours * 3600:
                    state = self._write_pipeline_run_state(
                        study_id,
                        status="failed",
                        message="Run may have been interrupted.",
                        error="Extraction run timed out in UI state.",
                    )
            except ValueError:
                pass
        llm_progress = state.get("llmProgress")
        return {
            "studyId": study_id,
            "status": state.get("status", "idle"),
            "currentStage": state.get("currentStage", ""),
            "currentSubStepId": state.get("currentSubStepId", ""),
            "message": state.get("message", ""),
            "error": state.get("error", ""),
            "startedAt": state.get("startedAt", ""),
            "finishedAt": state.get("finishedAt", ""),
            "logs": list(state.get("logs", [])),
            "llmProgress": llm_progress if isinstance(llm_progress, dict) else None,
        }

    def _resolve_review_source(self, study_id: str, review_source: str | None) -> str:
        if review_source:
            try:
                return review_sources.normalize_review_source(review_source)
            except ValueError as exc:
                raise UiApiError("VALIDATION_ERROR", str(exc), 400) from exc
        manifest = self._read_upload_manifest_obj(study_id)
        stored = str(manifest.get("reviewDisplaySource") or "").strip()
        if stored in review_sources.VALID_REVIEW_SOURCES:
            return stored
        return review_sources.REVIEW_SOURCE_GENERATED

    def _latest_import_snapshot_obj(
        self, study_id: str, *, pd_spec_import_mode: str | None
    ) -> Dict[str, Any] | None:
        review_dir = paths.local_review_dir(study_id, self.output_dir)
        if not review_dir.is_dir():
            return None
        mode_filter = str(pd_spec_import_mode or "").strip()
        for snapshot_path in sorted(review_dir.glob("deviations_import_*.json"), reverse=True):
            try:
                snapshot_obj = read_json(snapshot_path)
            except (OSError, json.JSONDecodeError):
                continue
            if mode_filter and str(snapshot_obj.get("pd_spec_import_mode", "")).strip() != mode_filter:
                continue
            return snapshot_obj
        return None

    def _seed_review_state_from_workbook(
        self, study_id: str, review_source: str, *, pd_spec_import_mode: str
    ) -> Dict[str, Any]:
        from pdcheck_factory import import_grounding

        workbook_bytes = self._read_pd_spec_workbook_bytes(study_id)
        if not workbook_bytes:
            return review_sources.empty_review_state(study_id)
        raw_deviations = parse_pd_spec_xlsx(workbook_bytes)
        snapshot = import_grounding.build_deviations_state(
            study_id=study_id,
            deviations=raw_deviations,
            import_version="seed",
            source_type="import",
            pd_spec_import_mode=pd_spec_import_mode,
        )
        return snapshot

    def _ensure_review_source_state(self, study_id: str, review_source: str) -> None:
        path = review_sources.review_state_path(study_id, self.output_dir, review_source)
        if path.is_file():
            return
        path.parent.mkdir(parents=True, exist_ok=True)

        if review_source == review_sources.REVIEW_SOURCE_GENERATED:
            legacy = paths.local_deviations_review_state(study_id, self.output_dir)
            parsed = paths.local_deviations_parsed_json(study_id, self.output_dir)
            if legacy.is_file():
                state_obj = read_json(legacy)
            elif parsed.is_file():
                state_obj = read_json(parsed)
            else:
                state_obj = review_sources.empty_review_state(study_id)
            write_json(path, state_obj)
            self._mirror_upload(study_id, path)
            return

        if review_source == review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC:
            snapshot = self._latest_import_snapshot_obj(study_id, pd_spec_import_mode="map")
            if snapshot is None:
                snapshot = self._latest_import_snapshot_obj(study_id, pd_spec_import_mode="ground")
            if snapshot is None:
                snapshot = self._seed_review_state_from_workbook(
                    study_id, review_source, pd_spec_import_mode="map"
                )
            write_json(path, snapshot)
            self._mirror_upload(study_id, path)
            return

        snapshot = self._latest_import_snapshot_obj(study_id, pd_spec_import_mode="enrich")
        if snapshot is None:
            snapshot = self._latest_import_snapshot_obj(study_id, pd_spec_import_mode="enrich_stub")
        if snapshot is None:
            snapshot = self._seed_review_state_from_workbook(
                study_id, review_source, pd_spec_import_mode="map"
            )
        write_json(path, snapshot)
        self._mirror_upload(study_id, path)

    def _imported_pd_spec_text_by_deviation_id(self, study_id: str) -> Dict[str, str]:
        path = review_sources.review_state_path(
            study_id, self.output_dir, review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC
        )
        if not path.is_file():
            return {}
        try:
            imported_obj = read_json(path)
        except (OSError, ValueError):
            return {}
        by_id: Dict[str, str] = {}
        for row in imported_obj.get("deviations", []):
            dev_id = str(row.get("deviation_id", "")).strip()
            if dev_id:
                by_id[dev_id] = str(row.get("text", "") or "")
        return by_id

    def _normalize_review_deviation_row(
        self,
        row: Dict[str, Any],
        *,
        study_id: str,
        review_source: str,
        imported_text_by_id: Dict[str, str] | None = None,
    ) -> Dict[str, Any]:
        normalized = dict(row)
        if has_flat_pd_spec_fields(normalized) or not isinstance(normalized.get("pd_spec_import"), dict):
            lift_pd_spec_row(normalized)
        if review_source == review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC:
            dev_id = str(normalized.get("deviation_id", "")).strip()
            if not str(normalized.get("original_deviation_text", "") or "").strip():
                backfill = (imported_text_by_id or {}).get(dev_id, "")
                normalized["original_deviation_text"] = backfill or str(normalized.get("text", "") or "")
        return normalized

    def _normalize_review_state_rows(
        self, study_id: str, state_obj: Dict[str, Any], review_source: str
    ) -> Dict[str, Any]:
        if review_source not in {
            review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC,
            review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC,
        }:
            return state_obj
        imported_by_id: Dict[str, str] | None = None
        if review_source == review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC:
            imported_by_id = self._imported_pd_spec_text_by_deviation_id(study_id)
        rows = [
            self._normalize_review_deviation_row(
                dict(row),
                study_id=study_id,
                review_source=review_source,
                imported_text_by_id=imported_by_id,
            )
            for row in state_obj.get("deviations", [])
        ]
        return {**state_obj, "deviations": rows}

    def _load_state(self, study_id: str, review_source: str | None = None) -> Dict[str, Any]:
        source = self._resolve_review_source(study_id, review_source)
        self._ensure_review_source_state(study_id, source)
        path = review_sources.review_state_path(study_id, self.output_dir, source)
        state_obj = read_json(path)
        return self._normalize_review_state_rows(study_id, state_obj, source)

    def _load_pseudo_state(self, study_id: str) -> Dict[str, Any]:
        path = paths.local_pseudo_logic_review_state(study_id, self.output_dir)
        if path.is_file():
            return read_json(path)
        return {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "generated_at": "",
            "items": [],
        }

    def _load_rules(self, study_id: str) -> Dict[str, Any]:
        path = paths.local_rules_parsed_json(study_id, self.output_dir)
        if path.is_file():
            return read_json(path)
        return {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "generated_at": "",
            "rules": [],
        }

    def _save_rules(self, study_id: str, rules_obj: Dict[str, Any]) -> None:
        rules_obj["schema_version"] = rules_obj.get("schema_version", "1.0.0")
        rules_obj["study_id"] = study_id
        if not rules_obj.get("generated_at"):
            rules_obj["generated_at"] = datetime.now(timezone.utc).isoformat()
        write_json(paths.local_rules_parsed_json(study_id, self.output_dir), rules_obj)
        self._mirror_upload(study_id, paths.local_rules_parsed_json(study_id, self.output_dir))

    def _load_paragraph_index(self, study_id: str) -> Dict[str, Dict[str, Any]]:
        path = paths.local_protocol_paragraph_index_json(study_id, self.output_dir)
        if not path.is_file():
            return {}
        obj = read_json(path)
        paragraphs = obj.get("paragraphs", [])
        by_ref: Dict[str, Dict[str, Any]] = {}
        if isinstance(paragraphs, list):
            for paragraph in paragraphs:
                if not isinstance(paragraph, dict):
                    continue
                ref = str(
                    paragraph.get("paragraph_id")
                    or paragraph.get("id")
                    or paragraph.get("ref")
                    or paragraph.get("paragraph_ref")
                    or ""
                )
                if ref:
                    by_ref[ref] = paragraph
        return by_ref

    def _chat_state_path(self, study_id: str) -> Path:
        return paths.local_review_dir(study_id, self.output_dir) / "deviation_chat_state.json"

    def _load_chat_state(self, study_id: str) -> Dict[str, Any]:
        chat_path = self._chat_state_path(study_id)
        if chat_path.is_file():
            return read_json(chat_path)
        return {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "updated_at": "",
            "deviations": {},
        }

    def _save_chat_state(self, study_id: str, chat_obj: Dict[str, Any]) -> None:
        chat_obj["updated_at"] = datetime.now(timezone.utc).isoformat()
        write_json(self._chat_state_path(study_id), chat_obj)
        self._mirror_upload(study_id, self._chat_state_path(study_id))

    def _append_chat_message(
        self,
        chat_obj: Dict[str, Any],
        deviation_id: str,
        *,
        role: str,
        text: str,
    ) -> None:
        dev_key = str(deviation_id)
        by_dev = dict(chat_obj.get("deviations", {}))
        cur = dict(by_dev.get(dev_key, {"messages": []}))
        msgs = list(cur.get("messages", []))
        msgs.append(
            {
                "role": role,
                "text": text,
                "ts": datetime.now(timezone.utc).isoformat(),
            }
        )
        cur["messages"] = msgs[-25:]
        by_dev[dev_key] = cur
        chat_obj["deviations"] = by_dev

    def _replace_row(self, state_obj: Dict[str, Any], updated_row: Dict[str, Any]) -> Dict[str, Any]:
        dev_id = str(updated_row.get("deviation_id", ""))
        rows = list(state_obj.get("deviations", []))
        for idx, row in enumerate(rows):
            if str(row.get("deviation_id", "")) == dev_id:
                rows[idx] = updated_row
                break
        state_obj["deviations"] = rows
        return state_obj

    def _persist_state(
        self,
        study_id: str,
        state_obj: Dict[str, Any],
        audit_obj: Dict[str, Any],
        *,
        review_source: str | None = None,
    ) -> None:
        source = self._resolve_review_source(study_id, review_source)
        state_obj["schema_version"] = state_obj.get("schema_version", "1.0.0")
        state_obj["study_id"] = study_id
        if not state_obj.get("generated_at"):
            state_obj["generated_at"] = datetime.now(timezone.utc).isoformat()
        per_source_path = review_sources.review_state_path(study_id, self.output_dir, source)
        write_json(per_source_path, state_obj)
        write_json(paths.local_deviations_review_audit_json(study_id, self.output_dir), audit_obj)
        mirror_paths: List[Path] = [
            per_source_path,
            paths.local_deviations_review_audit_json(study_id, self.output_dir),
        ]
        if source == review_sources.REVIEW_SOURCE_GENERATED:
            write_json(paths.local_deviations_review_state(study_id, self.output_dir), state_obj)
            write_json(paths.local_deviations_validated_json(study_id, self.output_dir), state_obj)
            mirror_paths.extend(
                [
                    paths.local_deviations_review_state(study_id, self.output_dir),
                    paths.local_deviations_validated_json(study_id, self.output_dir),
                ]
            )
        self._mirror_upload(study_id, *mirror_paths)

    def _audit(self, study_id: str, *, action: str, target_id: str, updated_rows: int) -> Dict[str, Any]:
        return {
            "study_id": study_id,
            "review_type": "deviations",
            "action": action,
            "target_id": target_id,
            "updated_rows": updated_rows,
            "revised_rows": 0,
            "run_revision_cycle": False,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }

    def _normalize_refs(self, value: Any, *, allow_empty: bool = False) -> List[str]:
        if isinstance(value, list):
            refs = [str(item).strip() for item in value if str(item).strip()]
        else:
            refs = [part.strip() for part in str(value or "").replace(";", ",").split(",") if part.strip()]
        if not refs and not allow_empty:
            raise UiApiError("VALIDATION_ERROR", "paragraph_refs is required", 400)
        invalid = [ref for ref in refs if not ref.startswith("p") or not ref[1:].isdigit()]
        if invalid:
            raise UiApiError("VALIDATION_ERROR", f"Invalid paragraph_refs: {', '.join(invalid)}", 400)
        return refs

    def _normalize_deviation_payload(self, payload: Dict[str, Any], *, default_source: str) -> Dict[str, Any]:
        deviation_id = str(payload.get("deviation_id") or payload.get("deviationId") or "").strip()
        rule_id = str(payload.get("rule_id") or payload.get("ruleId") or "").strip()
        text = str(payload.get("text") or payload.get("deviation_text") or payload.get("deviationText") or "").strip()
        if not deviation_id or not rule_id or not text:
            raise UiApiError("VALIDATION_ERROR", "deviation_id, rule_id, and text are required", 400)
        status = str(payload.get("status") or "pending").strip().lower()
        if status not in {"pending", "accepted", "to_review", "rejected"}:
            raise UiApiError("VALIDATION_ERROR", "status must be one of pending,accepted,to_review,rejected", 400)
        grounding_error = str(payload.get("grounding_error") or payload.get("groundingError") or "").strip()
        allow_empty_refs = bool(grounding_error) or default_source == "imported_pd_spec"
        row = {
            "deviation_id": deviation_id,
            "rule_id": rule_id,
            "text": text,
            "paragraph_refs": self._normalize_refs(
                payload.get("paragraph_refs") or payload.get("paragraphRefs"),
                allow_empty=allow_empty_refs,
            ),
            "data_support_note": str(payload.get("data_support_note") or payload.get("dataSupportNote") or ""),
            "status": status,
            "dm_comment": str(payload.get("dm_comment") or payload.get("dmComment") or ""),
            "entry_source": str(payload.get("entry_source") or payload.get("entrySource") or default_source),
        }
        for field in (
            "protocol_deviation_category",
            "protocol_deviation_sub_category",
            "classification",
            "data_source",
            "manual_or_programmable",
            "programming_status",
            "programmer_comments",
            "reviewer_comments",
            "aa_comment",
            "grounding_error",
        ):
            if field in payload or self._camel_field(field) in payload:
                row[field] = str(payload.get(field) or payload.get(self._camel_field(field)) or "").strip()
        return row

    @staticmethod
    def _camel_field(snake: str) -> str:
        parts = snake.split("_")
        return parts[0] + "".join(p.title() for p in parts[1:])

    def _normalized_rule_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        rule_id = str(payload.get("rule_id") or payload.get("ruleId") or "").strip()
        title = str(payload.get("title") or payload.get("rule_title") or payload.get("ruleTitle") or "").strip()
        text = str(payload.get("text") or payload.get("rule_text") or payload.get("ruleText") or "").strip()
        if not rule_id:
            raise UiApiError("VALIDATION_ERROR", "rule_id is required", 400)
        refs_value = payload.get("paragraph_refs") or payload.get("paragraphRefs")
        return {
            "rule_id": rule_id,
            "title": title,
            "text": text,
            "paragraph_refs": self._normalize_refs(refs_value) if refs_value else [],
        }

    def _load_protocol_enrichment_artifact(
        self, study_id: str, deviation_id: str
    ) -> Dict[str, Any] | None:
        path = paths.local_protocol_enrichment_json(study_id, self.output_dir, deviation_id)
        if not path.is_file():
            return None
        try:
            return read_json(path)
        except (OSError, ValueError):
            return None

    def _enrichment_row_by_deviation_id(
        self, study_id: str, deviation_id: str
    ) -> Dict[str, Any] | None:
        enriched_path = paths.local_deviations_review_enriched_pd_spec_json(
            study_id, self.output_dir
        )
        if not enriched_path.is_file():
            return None
        try:
            state_obj = read_json(enriched_path)
        except (OSError, ValueError):
            return None
        for row in state_obj.get("deviations", []):
            if str(row.get("deviation_id", "")) == deviation_id:
                return dict(row)
        return None

    def _enrichment_detail_from_artifact(self, artifact: Dict[str, Any]) -> Dict[str, Any]:
        merged = (artifact or {}).get("merged") or {}
        suggested = str(
            merged.get("suggested_deviation_text")
            or merged.get("improved_deviation_text")
            or ""
        )
        protocol_grounding = dict(artifact.get("protocol_grounding") or {})
        acrf_grounding = dict(artifact.get("acrf_grounding") or {})
        return {
            "enrichment_status": str(artifact.get("enrichment_status") or ""),
            "enrichment_summary": str(artifact.get("enrichment_summary") or ""),
            "enrichment_errors": dict(artifact.get("enrichment_errors") or {}),
            "original_deviation_text": str(merged.get("original_deviation_text") or ""),
            "suggested_deviation_text": suggested,
            "improved_deviation_text": suggested,
            "improved_pseudo_logic_plain_english": str(
                merged.get("improved_pseudo_logic_plain_english") or ""
            ),
            "paragraph_refs": list(merged.get("paragraph_refs") or []),
            "assumptions": list(merged.get("assumptions") or []),
            "caveats": list(merged.get("caveats") or []),
            "data_gaps": list(merged.get("data_gaps") or []),
            "weak_spots": list(merged.get("weak_spots") or []),
            "suggested_changes": list(merged.get("suggested_changes") or []),
            "protocol_conflicts": list(merged.get("protocol_conflicts") or []),
            "programmability_risk": str(merged.get("programmability_risk") or ""),
            "required_datasets": list(merged.get("required_datasets") or []),
            "required_fields": list(merged.get("required_fields") or []),
            "protocol_grounding": protocol_grounding,
            "acrf_grounding": acrf_grounding,
        }

    def _enrichment_detail_from_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        suggested = str(row.get("suggested_deviation_text", "") or "").strip()
        return {
            "enrichment_status": pd_spec_field(row, "enrichment_status"),
            "enrichment_summary": pd_spec_field(row, "enrichment_summary"),
            "enrichment_errors": {},
            "original_deviation_text": str(row.get("original_deviation_text", "") or ""),
            "suggested_deviation_text": suggested,
            "improved_deviation_text": suggested,
            "improved_pseudo_logic_plain_english": pd_spec_field(row, "pseudo_logic_seed"),
            "paragraph_refs": list(row.get("paragraph_refs") or []),
            "assumptions": [],
            "caveats": [],
            "data_gaps": [],
            "weak_spots": [],
            "suggested_changes": [],
            "protocol_conflicts": [],
            "programmability_risk": "",
            "required_datasets": [],
            "required_fields": [],
            "protocol_grounding": {},
            "acrf_grounding": {},
        }

    def _enrichment_api_fields(
        self, study_id: str, row: Dict[str, Any], *, review_source: str
    ) -> Dict[str, Any]:
        if review_source != review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC:
            return {}
        deviation_id = str(row.get("deviation_id", ""))
        artifact = self._load_protocol_enrichment_artifact(study_id, deviation_id)
        suggested = str(row.get("suggested_deviation_text", "") or "").strip()
        if not suggested and artifact:
            merged = (artifact or {}).get("merged") or {}
            suggested = str(
                merged.get("suggested_deviation_text")
                or merged.get("improved_deviation_text")
                or ""
            )
        return {
            "enrichment_status": pd_spec_field(
                row, "enrichment_status", default=str((artifact or {}).get("enrichment_status") or "")
            ),
            "enrichment_summary": pd_spec_field(
                row, "enrichment_summary", default=str((artifact or {}).get("enrichment_summary") or "")
            ),
            "original_deviation_text": str(row.get("original_deviation_text", "") or ""),
            "suggested_deviation_text": suggested,
        }

    def get_step7_enrichment_detail(self, study_id: str, deviation_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        dev_id = str(deviation_id).strip()
        if not dev_id:
            raise UiApiError("VALIDATION_ERROR", "deviationId is required", 400)
        artifact = self._load_protocol_enrichment_artifact(study_id, dev_id)
        if artifact is not None:
            fields = self._enrichment_detail_from_artifact(artifact)
        else:
            row = self._enrichment_row_by_deviation_id(study_id, dev_id)
            if row is None:
                raise UiApiError(
                    "NOT_FOUND",
                    f"No enrichment artifact for deviation '{dev_id}'",
                    404,
                )
            has_summary = bool(
                str(row.get("suggested_deviation_text", "") or "").strip()
                or pd_spec_field(row, "enrichment_summary")
                or pd_spec_field(row, "enrichment_status")
            )
            if not has_summary:
                raise UiApiError(
                    "NOT_FOUND",
                    f"No enrichment artifact for deviation '{dev_id}'",
                    404,
                )
            fields = self._enrichment_detail_from_row(row)
        return {
            "studyId": study_id,
            "deviationId": dev_id,
            **fields,
        }

    def _normalized_step7_row(
        self,
        row: Dict[str, Any],
        pseudo_by_dev: Dict[str, Dict[str, Any]],
        rule_by_id: Dict[str, Dict[str, Any]],
        paragraph_by_ref: Dict[str, Dict[str, Any]] | None = None,
        *,
        study_id: str = "",
        review_source: str = "",
    ) -> Dict[str, Any]:
        deviation_id = str(row.get("deviation_id", ""))
        rule_id = str(row.get("rule_id", ""))
        pseudo = pseudo_by_dev.get(deviation_id, {})
        rule = rule_by_id.get(rule_id, {})
        refs = list(row.get("paragraph_refs", []))
        paragraph_lookup = paragraph_by_ref or {}
        supporting_sentences = []
        for ref in refs:
            paragraph = paragraph_lookup.get(str(ref), {})
            text = str(paragraph.get("text") or paragraph.get("content") or paragraph.get("paragraph_text") or "")
            supporting_sentences.append({"ref": str(ref), "text": text})
        category = pd_spec_field(row, "protocol_deviation_category").strip()
        sub_category = pd_spec_field(row, "protocol_deviation_sub_category").strip()
        rule_title = str(rule.get("title", "")).strip()
        entry_source = pd_spec_field(row, "entry_source", default=str(row.get("entry_source", "extracted")))
        if not rule_title:
            if sub_category:
                rule_title = sub_category
            elif category or sub_category:
                rule_title = f"{category} / {sub_category}".strip(" /")
        programmable = pseudo.get("programmable")
        if programmable is None:
            programmable = programmable_from_manual_or_programmable(
                pd_spec_field(row, "manual_or_programmable").strip()
            )
        rule_text = str(rule.get("text") or rule.get("rule_text") or rule.get("description") or "")
        if not rule_text and category and entry_source == "imported_pd_spec":
            rule_text = category
        result = {
            "rule_id": rule_id,
            "deviation_id": deviation_id,
            "rule_title": rule_title,
            "rule_text": rule_text,
            "deviation_text": str(row.get("text", "")),
            "paragraph_refs": refs,
            "paragraph_refs_text": ", ".join(refs),
            "supporting_sentences": supporting_sentences,
            "data_support_note": str(row.get("data_support_note", "")),
            "pseudo_logic": str(pseudo.get("pseudo_logic", "")),
            "status": str(row.get("status", "pending")),
            "dm_comment": str(row.get("dm_comment", "")),
            "entry_source": entry_source,
            "programmable": programmable,
            "programmability_note": str(pseudo.get("programmability_note", "")),
            "protocol_deviation_category": category,
            "protocol_deviation_sub_category": sub_category,
        }
        if study_id and review_source:
            result.update(self._enrichment_api_fields(study_id, row, review_source=review_source))
        return result

    def _collect_study_ids_from_blob(self) -> set[str]:
        blob_service = blob_io.blob_service_from_env()
        container = blob_io.container_from_env()
        study_ids: set[str] = set()
        for prefix in ("raw/", "extractions/", "pipeline/", "review/"):
            names = blob_io.list_blob_names_with_prefix(
                blob_service=blob_service,
                container_name=container,
                prefix=prefix,
            )
            for name in names:
                parts = name.strip("/").split("/")
                if len(parts) < 2 or parts[0] != prefix.rstrip("/"):
                    continue
                study_id = parts[1]
                if study_id:
                    study_ids.add(study_id)
        return study_ids

    def _study_exists(self, study_id: str) -> bool:
        if self._ui_upload_manifest_path(study_id).is_file():
            return True
        try:
            blob_service = blob_io.blob_service_from_env()
            container = blob_io.container_from_env()
            blob_path = paths.ui_upload_manifest_blob(study_id)
            blob_io.download_blob_bytes(
                blob_service=blob_service,
                container_name=container,
                blob_path=blob_path,
            )
            return True
        except Exception:  # noqa: BLE001
            pass
        return study_id in self._collect_study_ids_from_blob()

    def _explicit_workflow_choice(self, manifest: Dict[str, Any]) -> str | None:
        explicit = str(manifest.get("workflowChoice") or "").strip()
        if explicit in VALID_WORKFLOW_CHOICES:
            return explicit
        return None

    def _infer_workflow_choice(self, manifest: Dict[str, Any]) -> str | None:
        explicit = self._explicit_workflow_choice(manifest)
        if explicit:
            return explicit
        entry_mode = str(manifest.get("entryMode") or ENTRY_MODE_EXTRACTED).strip()
        import_mode = str(manifest.get("pdSpecImportMode") or "").strip()
        if entry_mode == ENTRY_MODE_IMPORTED_PD_SPEC:
            if import_mode == "enrich":
                return WORKFLOW_CHOICE_ENRICH
            if import_mode == "map":
                return WORKFLOW_CHOICE_MAP
        if entry_mode == ENTRY_MODE_EXTRACTED:
            return WORKFLOW_CHOICE_EXTRACT
        return None

    def _workflow_label(self, workflow: str | None) -> str:
        labels = {
            WORKFLOW_CHOICE_EXTRACT: "Extract PD from protocol + aCRF",
            WORKFLOW_CHOICE_MAP: "Map uploaded PD Specifications",
            WORKFLOW_CHOICE_ENRICH: "Enrich PD Specifications",
        }
        return labels.get(workflow or "", "Not selected")

    def _uploads_complete_for_workflow(
        self, workflow: str | None, upload_status: Dict[str, Any]
    ) -> bool:
        protocol = bool(upload_status.get("protocol", {}).get("uploaded"))
        acrf = bool(upload_status.get("acrf", {}).get("uploaded"))
        pd_spec = bool(upload_status.get("pdSpec", {}).get("uploaded"))
        if workflow == WORKFLOW_CHOICE_EXTRACT:
            return protocol and acrf
        if workflow in {WORKFLOW_CHOICE_MAP, WORKFLOW_CHOICE_ENRICH}:
            return protocol and acrf and pd_spec
        return False

    def _workflow_steps_started(
        self, workflow: str | None, statuses: Dict[str, str]
    ) -> bool:
        if not workflow:
            return False
        step_ids = WORKFLOW_STEPS.get(workflow, [])
        return any(statuses.get(step_id) in {"done", "skipped"} for step_id in step_ids)

    def _workflow_steps_complete(
        self, workflow: str | None, statuses: Dict[str, str]
    ) -> bool:
        if not workflow:
            return False
        step_ids = WORKFLOW_STEPS.get(workflow, [])
        if not step_ids:
            return False
        return all(statuses.get(step_id) in {"done", "skipped"} for step_id in step_ids)

    def derive_stage(
        self,
        *,
        workflow: str | None,
        upload_status: Dict[str, Any],
        statuses: Dict[str, str],
    ) -> str:
        if not workflow:
            return UI_STAGE_PROJECT
        if not self._uploads_complete_for_workflow(workflow, upload_status):
            return UI_STAGE_SETUP
        if not self._workflow_steps_started(workflow, statuses):
            return UI_STAGE_SETUP
        if not self._workflow_steps_complete(workflow, statuses):
            return UI_STAGE_PROCESSING
        return UI_STAGE_REVIEW

    def _deviation_summary(self, study_id: str) -> Dict[str, int] | None:
        try:
            state = self._load_state(study_id)
            deviations = list(state.get("deviations", []))
            if not deviations:
                return None
            accepted = sum(1 for row in deviations if str(row.get("status")) == "accepted")
            rejected = sum(1 for row in deviations if str(row.get("status")) == "rejected")
            to_review = sum(
                1
                for row in deviations
                if str(row.get("status", "pending")) in {"pending", "to_review"}
            )
            return {
                "total": len(deviations),
                "accepted": accepted,
                "toReview": to_review,
                "rejected": rejected,
            }
        except Exception:  # noqa: BLE001
            return None

    def create_study(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        self._assert_safe_study_id(study_id)
        if self._study_exists(study_id):
            raise UiApiError(
                "DUPLICATE_STUDY",
                f"Study '{study_id}' already exists.",
                409,
            )
        manifest = self._write_upload_manifest(
            study_id,
            entry_mode=ENTRY_MODE_EXTRACTED,
            workflow_choice=None,
            ui_stage=UI_STAGE_PROJECT,
        )
        return {
            "studyId": study_id,
            "manifestBlobPath": paths.ui_upload_manifest_blob(study_id),
            "manifest": manifest,
        }

    def patch_study_manifest(self, study_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        workflow_choice = payload.get("workflowChoice")
        ui_stage = payload.get("uiStage")

        entry_mode: str | None = None
        pd_spec_import_mode: str | None = None
        clear_pd_spec_import_mode = False

        if workflow_choice is not None:
            choice = str(workflow_choice).strip()
            if choice not in VALID_WORKFLOW_CHOICES:
                raise UiApiError(
                    "VALIDATION_ERROR",
                    f"workflowChoice must be one of: {', '.join(sorted(VALID_WORKFLOW_CHOICES))}",
                    400,
                )
            if choice == WORKFLOW_CHOICE_EXTRACT:
                entry_mode = ENTRY_MODE_EXTRACTED
                clear_pd_spec_import_mode = True
            elif choice == WORKFLOW_CHOICE_MAP:
                entry_mode = ENTRY_MODE_IMPORTED_PD_SPEC
                pd_spec_import_mode = "map"
            elif choice == WORKFLOW_CHOICE_ENRICH:
                entry_mode = ENTRY_MODE_IMPORTED_PD_SPEC
                pd_spec_import_mode = "enrich"

        if ui_stage is not None:
            stage = str(ui_stage).strip()
            if stage not in VALID_UI_STAGES:
                raise UiApiError(
                    "VALIDATION_ERROR",
                    f"uiStage must be one of: {', '.join(sorted(VALID_UI_STAGES))}",
                    400,
                )

        manifest = self._write_upload_manifest(
            study_id,
            entry_mode=entry_mode,
            pd_spec_import_mode=pd_spec_import_mode,
            workflow_choice=str(workflow_choice).strip() if workflow_choice is not None else None,
            ui_stage=str(ui_stage).strip() if ui_stage is not None else None,
            clear_pd_spec_import_mode=clear_pd_spec_import_mode,
        )
        summary = self.get_study_summary(study_id)
        return {
            "studyId": study_id,
            "manifest": manifest,
            "stage": summary["stage"],
            "workflow": summary["workflow"],
        }

    def get_study_summary(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        manifest = self._read_upload_manifest_obj(study_id)
        upload_status = self.get_step1_upload_status(study_id)
        run_state = self.get_step1_run_state(study_id)
        status_payload = self.get_status(study_id)
        statuses = {item["stepId"]: item["status"] for item in status_payload["steps"]}
        workflow = self._explicit_workflow_choice(manifest)
        inferred_workflow = self._infer_workflow_choice(manifest)
        stage = self.derive_stage(
            workflow=workflow,
            upload_status=upload_status,
            statuses=statuses,
        )
        deviation_summary = self._deviation_summary(study_id)
        return {
            "studyId": study_id,
            "workflow": workflow,
            "inferredWorkflow": inferred_workflow,
            "workflowLabel": self._workflow_label(workflow or inferred_workflow),
            "stage": stage,
            "entryMode": manifest.get("entryMode") or ENTRY_MODE_EXTRACTED,
            "workflowChoice": manifest.get("workflowChoice"),
            "pdSpecImportMode": manifest.get("pdSpecImportMode"),
            "lastModified": manifest.get("uploadedAt"),
            "uiStage": manifest.get("uiStage") or stage,
            "uploads": {
                "protocol": upload_status["protocol"],
                "acrf": upload_status["acrf"],
                "pdSpec": upload_status["pdSpec"],
            },
            "bothUploaded": upload_status["bothUploaded"],
            "allThreeUploaded": upload_status["allThreeUploaded"],
            "preprocess": {
                "protocol": upload_status["protocolPreprocessed"],
                "acrf": upload_status["acrfPreprocessed"],
            },
            "processingComplete": upload_status["processingComplete"],
            "runState": run_state,
            "steps": status_payload["steps"],
            "stepStatuses": statuses,
            "nextStepId": status_payload["nextStepId"],
            "importVersions": status_payload.get("importVersions"),
            "codingPhaseAccepted": status_payload.get("codingPhaseAccepted", False),
            "deviationSummary": deviation_summary,
        }

    def list_studies(self) -> Dict[str, Any]:
        study_ids = self._collect_study_ids_from_blob()
        studies = []
        for study_id in sorted(study_ids):
            manifest = self._read_upload_manifest_obj(study_id)
            workflow = self._explicit_workflow_choice(manifest)
            inferred_workflow = self._infer_workflow_choice(manifest)
            upload_status = {
                "protocol": {"uploaded": self._blob_has_upload(study_id, "protocol")},
                "acrf": {"uploaded": self._blob_has_upload(study_id, "acrf")},
                "pdSpec": {"uploaded": self._blob_has_pd_spec_workbook(study_id)},
            }
            statuses = self._step_statuses(study_id)
            stage = self.derive_stage(
                workflow=workflow,
                upload_status=upload_status,
                statuses=statuses,
            )
            studies.append(
                {
                    "studyId": study_id,
                    "workflow": workflow or inferred_workflow,
                    "workflowLabel": self._workflow_label(workflow or inferred_workflow),
                    "stage": stage,
                    "lastModified": manifest.get("uploadedAt"),
                }
            )
        return {"studies": studies}

    def list_openai_deployments(self) -> Dict[str, Any]:
        from pdcheck_factory import azure_openai_config

        return azure_openai_config.list_openai_deployments()

    def get_pd_taxonomy(self) -> Dict[str, Any]:
        from pdcheck_factory.pd_taxonomy import all_sub_category_options, category_options, load_taxonomy

        return {
            "categories": load_taxonomy(),
            "categoryOptions": category_options(),
            "subCategoryOptions": all_sub_category_options(),
        }

    def delete_study(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        self._assert_safe_study_id(study_id)

        blob_service = blob_io.blob_service_from_env()
        container = blob_io.container_from_env()
        blob_names: List[str] = []
        prefixes_scanned: List[str] = []
        for prefix in paths.study_blob_list_prefixes(study_id):
            names = blob_io.list_blob_names_with_prefix(
                blob_service=blob_service,
                container_name=container,
                prefix=prefix,
            )
            if names:
                prefixes_scanned.append(prefix)
            blob_names.extend(names)

        unique_blob_names = sorted(set(blob_names))
        deleted_blob_count = 0
        if unique_blob_names:
            deleted_blob_count = blob_io.delete_blobs(
                blob_service=blob_service,
                container_name=container,
                blob_paths=unique_blob_names,
            )

        local_root = paths.local_study_root(study_id, self.output_dir)
        local_output_removed = False
        if local_root.exists():
            shutil.rmtree(local_root)
            local_output_removed = True

        return {
            "studyId": study_id,
            "deletedBlobCount": deleted_blob_count,
            "totalBlobCount": len(unique_blob_names),
            "blobPrefixes": prefixes_scanned,
            "localOutputRemoved": local_output_removed,
            "message": (
                f"Deleted {deleted_blob_count} blob object(s) for study {study_id!r}."
                if unique_blob_names
                else f"No blob objects found for study {study_id!r}."
            ),
        }

    def upload_step1_files(
        self,
        study_id: str,
        protocol_bytes: bytes | None = None,
        acrf_bytes: bytes | None = None,
        *,
        protocol_file_name: str | None = None,
        acrf_file_name: str | None = None,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        if not protocol_bytes and not acrf_bytes:
            raise UiApiError(
                "VALIDATION_ERROR",
                "At least one of protocolFile or acrfFile must be provided",
                400,
            )

        max_mb = int(os.getenv("UI_UPLOAD_MAX_MB", "100"))
        max_bytes = max_mb * 1024 * 1024
        if protocol_bytes and len(protocol_bytes) > max_bytes:
            raise UiApiError("VALIDATION_ERROR", f"Protocol file must be <= {max_mb}MB", 400)
        if acrf_bytes and len(acrf_bytes) > max_bytes:
            raise UiApiError("VALIDATION_ERROR", f"aCRF file must be <= {max_mb}MB", 400)

        blob_service = blob_io.blob_service_from_env()
        container = blob_io.container_from_env()
        protocol_blob = paths.raw_protocol_blob(study_id)
        acrf_blob = paths.raw_acrf_blob(study_id)
        protocol_name = (protocol_file_name or "").strip() or "protocol.pdf"
        acrf_name = (acrf_file_name or "").strip() or "acrf.pdf"
        protocol_size: int | None = None
        acrf_size: int | None = None

        if protocol_bytes:
            blob_io.upload_blob_bytes(
                blob_service=blob_service,
                container_name=container,
                blob_path=protocol_blob,
                data=protocol_bytes,
                content_type="application/pdf",
            )
            self._upload_reference_copy(study_id, "protocol", protocol_bytes, protocol_name)
            protocol_size = len(protocol_bytes)

        if acrf_bytes:
            blob_io.upload_blob_bytes(
                blob_service=blob_service,
                container_name=container,
                blob_path=acrf_blob,
                data=acrf_bytes,
                content_type="application/pdf",
            )
            self._upload_reference_copy(study_id, "acrf", acrf_bytes, acrf_name)
            acrf_size = len(acrf_bytes)

        manifest = self._write_upload_manifest(
            study_id,
            protocol_file_name=protocol_name if protocol_bytes else None,
            acrf_file_name=acrf_name if acrf_bytes else None,
            protocol_size=protocol_size,
            acrf_size=acrf_size,
        )

        upload_status = self.get_step1_upload_status(study_id)
        return {
            "studyId": study_id,
            "protocolBlob": protocol_blob,
            "acrfBlob": acrf_blob,
            "protocolFileName": manifest["protocolFileName"],
            "acrfFileName": manifest["acrfFileName"],
            "protocolSize": int(manifest.get("protocolSize") or 0),
            "acrfSize": int(manifest.get("acrfSize") or 0),
            "bothUploaded": upload_status["bothUploaded"],
            "stepStatuses": self._step_statuses(study_id),
        }

    def run_step1_extract(
        self,
        study_id: str,
        extractor: str | None = None,
        *,
        force: bool = False,
    ) -> Dict[str, Any]:
        from pdcheck_factory.cli import run_extract

        study_id = self._require_study_id(study_id)
        self._assert_both_uploads_ready(study_id)

        if not force and self._step_artifact_complete(study_id, "extract-inputs"):
            mode = extraction_resolve.read_ui_extractor_choice(study_id, self.output_dir)
            return {
                "studyId": study_id,
                "message": "Extraction already complete",
                "skipped": True,
                "extractor": mode or extraction_resolve.UI_EXTRACTOR_BOTH,
                "stepStatuses": self._step_statuses(study_id),
            }

        raw = (extractor or "").strip().lower()
        if not raw:
            mode = extraction_resolve.UI_EXTRACTOR_BOTH
        elif raw in extraction_resolve.VALID_UI_EXTRACTORS:
            mode = raw
        else:
            raise UiApiError(
                "VALIDATION_ERROR",
                "extractor must be 'opendataloader', 'document_intelligence', or 'both'.",
                400,
            )

        run_odl = mode != extraction_resolve.UI_EXTRACTOR_DI
        odl_only = mode == extraction_resolve.UI_EXTRACTOR_OPEN

        started_at = datetime.now(timezone.utc).isoformat()
        active_run = self._active_run_entry(study_id)
        active_run_id = str(active_run.get("runId")) if active_run else ""
        self._write_pipeline_run_state(
            study_id,
            status="running",
            currentStage="extract",
            currentSubStepId="extract-inputs",
            message="Extracting PDFs — this may take several minutes.",
            error="",
            startedAt=started_at,
            finishedAt="",
            logs=[],
            activeRunId=active_run_id,
        )
        self._append_pipeline_log(study_id, f"Starting extraction (extractor={mode})")

        def _extract_log(message: str) -> None:
            self._append_pipeline_log(study_id, message)

        from pdcheck_factory import llm

        try:
            with llm.use_pipeline_log(_extract_log):
                run_extract(
                    study_id=study_id,
                    protocol_blob=None,
                    acrf_blob=None,
                    output_dir=self.output_dir,
                    model_id=None,
                    sas_ttl=int(os.getenv("DI_SAS_TTL_MINUTES", "15")),
                    upload=True,
                    skip_acrf=False,
                    skip_protocol=False,
                    upload_only=False,
                    run_opendataloader_ocr=run_odl,
                    opendataloader_only=odl_only,
                    debug_blob=False,
                    log_callback=_extract_log,
                )
            extraction_resolve.write_ui_extractor_choice(study_id, self.output_dir, mode)
            self._mirror_upload(study_id, extraction_resolve.local_ui_extractor_choice_json(study_id, self.output_dir))
            extractions_root = paths.local_study_root(study_id, self.output_dir) / "extractions"
            study_artifact_sync.mirror_upload_directory(study_id, self.output_dir, extractions_root)
            self._append_pipeline_log(study_id, "PDF extraction completed")
            self._write_pipeline_run_state(
                study_id,
                status="done",
                currentStage="complete",
                currentSubStepId="extract-inputs",
                message="Extraction completed",
                finishedAt=datetime.now(timezone.utc).isoformat(),
                activeRunId=active_run_id,
            )
            self._touch_active_run_after_step(study_id)
        except Exception as exc:  # noqa: BLE001
            self._append_pipeline_log(study_id, f"Extraction failed: {exc}", level="error")
            self._write_pipeline_run_state(
                study_id,
                status="failed",
                message="Extraction failed",
                error=str(exc),
                finishedAt=datetime.now(timezone.utc).isoformat(),
            )
            raise

        return {
            "studyId": study_id,
            "message": "Extraction completed",
            "extractor": mode,
            "stepStatuses": self._step_statuses(study_id),
        }

    def _run_partial_extract(
        self,
        study_id: str,
        *,
        skip_protocol: bool,
        skip_acrf: bool,
        log_prefix: str,
        force: bool = False,
    ) -> None:
        from pdcheck_factory.cli import run_extract

        if not force:
            p = self._study_paths(study_id)
            if skip_acrf and p.protocol_source.exists():
                return
            if skip_protocol and p.acrf_source.exists():
                return

        mode = extraction_resolve.read_ui_extractor_choice(study_id, self.output_dir)
        if not mode:
            mode = extraction_resolve.UI_EXTRACTOR_BOTH
        run_odl = mode != extraction_resolve.UI_EXTRACTOR_DI
        odl_only = mode == extraction_resolve.UI_EXTRACTOR_OPEN

        def _extract_log(message: str) -> None:
            self._append_pipeline_log(study_id, f"{log_prefix}: {message}")

        run_extract(
            study_id=study_id,
            protocol_blob=None,
            acrf_blob=None,
            output_dir=self.output_dir,
            model_id=None,
            sas_ttl=int(os.getenv("DI_SAS_TTL_MINUTES", "15")),
            upload=True,
            skip_acrf=skip_acrf,
            skip_protocol=skip_protocol,
            upload_only=False,
            run_opendataloader_ocr=run_odl,
            opendataloader_only=odl_only,
            debug_blob=False,
            log_callback=_extract_log,
        )
        extractions_root = paths.local_study_root(study_id, self.output_dir) / "extractions"
        study_artifact_sync.mirror_upload_directory(study_id, self.output_dir, extractions_root)

    def preprocess_protocol(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        self._assert_protocol_upload_ready(study_id)

        p = self._study_paths(study_id)
        if p.paragraph_index.exists():
            summary = "Protocol already indexed."
            status = self.get_step1_upload_status(study_id)
            return {
                "studyId": study_id,
                "role": "protocol",
                "message": summary,
                "skipped": True,
                "protocolPreprocessed": status["protocolPreprocessed"],
                "stepStatuses": status["stepStatuses"],
            }

        self._write_pipeline_run_state(
            study_id,
            status="running",
            currentStage="index",
            currentSubStepId="preprocess-protocol",
            message="Preparing protocol (extract + index)…",
            error="",
            startedAt=datetime.now(timezone.utc).isoformat(),
            finishedAt="",
        )
        self._append_pipeline_log(study_id, "Starting protocol preprocess")

        try:
            p = self._study_paths(study_id)
            if not p.protocol_source.exists():
                self._run_partial_extract(
                    study_id,
                    skip_protocol=False,
                    skip_acrf=True,
                    log_prefix="protocol",
                )
            if not p.paragraph_index.exists():
                result = pipeline_v2.step2_protocol_paragraph_index(study_id, self.output_dir)
                index_path = paths.local_protocol_paragraph_index_json(study_id, self.output_dir)
                study_artifact_sync.mirror_upload_path(study_id, self.output_dir, index_path)
                summary = f"Protocol ready: indexed {len(result.get('paragraphs', []))} paragraphs."
            else:
                summary = "Protocol already indexed."
            self._append_pipeline_log(study_id, summary)
            self._write_pipeline_run_state(
                study_id,
                status="done",
                currentStage="complete",
                currentSubStepId="preprocess-protocol",
                message=summary,
                finishedAt=datetime.now(timezone.utc).isoformat(),
            )
        except Exception as exc:  # noqa: BLE001
            self._append_pipeline_log(study_id, f"Protocol preprocess failed: {exc}", level="error")
            self._write_pipeline_run_state(
                study_id,
                status="failed",
                message="Protocol preprocess failed",
                error=str(exc),
                finishedAt=datetime.now(timezone.utc).isoformat(),
            )
            raise

        status = self.get_step1_upload_status(study_id)
        return {
            "studyId": study_id,
            "role": "protocol",
            "message": summary,
            "protocolPreprocessed": status["protocolPreprocessed"],
            "stepStatuses": status["stepStatuses"],
        }

    def preprocess_acrf(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        self._assert_acrf_upload_ready(study_id)

        p = self._study_paths(study_id)
        if p.acrf_summary_text_merged.exists():
            summary = "aCRF already summarized."
            status = self.get_step1_upload_status(study_id)
            return {
                "studyId": study_id,
                "role": "acrf",
                "message": summary,
                "skipped": True,
                "acrfPreprocessed": status["acrfPreprocessed"],
                "stepStatuses": status["stepStatuses"],
            }

        self._write_pipeline_run_state(
            study_id,
            status="running",
            currentStage="acrf_split",
            currentSubStepId="preprocess-acrf",
            message="Preparing aCRF (extract + split + summary)…",
            error="",
            startedAt=datetime.now(timezone.utc).isoformat(),
            finishedAt="",
        )
        self._append_pipeline_log(study_id, "Starting aCRF preprocess")

        try:
            from pdcheck_factory.cli import run_acrf_split_toc

            p = self._study_paths(study_id)
            if not p.acrf_source.exists():
                self._run_partial_extract(
                    study_id,
                    skip_protocol=True,
                    skip_acrf=False,
                    log_prefix="acrf",
                )
                p = self._study_paths(study_id)

            if not p.acrf_sections_toc_dir.exists() or not any(p.acrf_sections_toc_dir.glob("*.md")):
                if not p.acrf_source.exists():
                    raise UiApiError("STEP_BLOCKED", f"Missing aCRF source markdown: {p.acrf_source}", 409)
                count, _manifest_path = run_acrf_split_toc(
                    source_md=p.acrf_source,
                    destination_dir=p.acrf_sections_toc_dir,
                    write_manifest=True,
                )
                study_artifact_sync.mirror_upload_directory(study_id, self.output_dir, p.acrf_sections_toc_dir)
                self._append_pipeline_log(study_id, f"Split aCRF into {count} sections")

            if not p.acrf_summary_text_merged.exists():
                result = pipeline_v2.step1_acrf_summary_text(study_id, self.output_dir)
                summary_path = paths.local_acrf_summary_text_merged(study_id, self.output_dir)
                study_artifact_sync.mirror_upload_path(study_id, self.output_dir, summary_path)
                summary = f"aCRF ready: merged summary with {len(result.get('datasets', []))} datasets."
            else:
                summary = "aCRF already summarized."
            self._append_pipeline_log(study_id, summary)
            self._write_pipeline_run_state(
                study_id,
                status="done",
                currentStage="complete",
                currentSubStepId="preprocess-acrf",
                message=summary,
                finishedAt=datetime.now(timezone.utc).isoformat(),
            )
        except Exception as exc:  # noqa: BLE001
            self._append_pipeline_log(study_id, f"aCRF preprocess failed: {exc}", level="error")
            self._write_pipeline_run_state(
                study_id,
                status="failed",
                message="aCRF preprocess failed",
                error=str(exc),
                finishedAt=datetime.now(timezone.utc).isoformat(),
            )
            raise

        status = self.get_step1_upload_status(study_id)
        return {
            "studyId": study_id,
            "role": "acrf",
            "message": summary,
            "acrfPreprocessed": status["acrfPreprocessed"],
            "stepStatuses": status["stepStatuses"],
        }

    def get_step1_preview(self, study_id: str, *, full: bool = False) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        p = self._study_paths(study_id)
        filenames = self._read_upload_filenames(study_id)
        preview_max = 500_000 if full else 8000
        return {
            "studyId": study_id,
            "protocolPreview": self._read_excerpt(p.protocol_source, max_chars=preview_max),
            "acrfPreview": self._read_excerpt(p.acrf_source, max_chars=preview_max),
            "protocolPreviewPath": str(p.protocol_source),
            "acrfPreviewPath": str(p.acrf_source),
            "protocolExists": p.protocol_source.exists(),
            "acrfExists": p.acrf_source.exists(),
            "protocolFileName": filenames["protocolFileName"],
            "acrfFileName": filenames["acrfFileName"],
            "extractor": extraction_resolve.read_ui_extractor_choice(study_id, self.output_dir),
            "stepStatuses": self._step_statuses(study_id),
        }

    def _preview_row_from_normalized(self, row: Dict[str, Any]) -> Dict[str, Any]:
        text = str(row.get("deviation_text", "") or row.get("text", ""))
        return {
            "deviation_id": str(row.get("deviation_id", "")),
            "rule_id": str(row.get("rule_id", "")),
            "rule_title": str(row.get("rule_title", "")),
            "deviation_text": text,
            "text": text,
            "entry_source": str(row.get("entry_source", "")),
            "status": str(row.get("status", "")),
        }

    def get_specifications_preview(self, study_id: str) -> Dict[str, Any]:
        from pdcheck_factory.pd_spec_import import parse_pd_spec_xlsx_table

        study_id = self._require_study_id(study_id)
        sources: List[Dict[str, Any]] = []
        workbook_bytes = self._read_pd_spec_workbook_bytes(study_id)
        if workbook_bytes:
            try:
                table = parse_pd_spec_xlsx_table(workbook_bytes)
                sources.append(
                    {
                        "key": "pd_spec_workbook",
                        "label": "PD Specifications workbook (parsed)",
                        "columns": table["headers"],
                        "rows": table["rows"],
                    }
                )
            except Exception:
                sources.append(
                    {
                        "key": "pd_spec_workbook",
                        "label": "PD Specifications workbook (parsed)",
                        "rows": [],
                    }
                )

        state_obj = self._load_state(study_id)
        pseudo_obj = self._load_pseudo_state(study_id)
        rules_obj = self._load_rules(study_id)
        paragraph_by_ref = self._load_paragraph_index(study_id)
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        review_rows = [
            self._preview_row_from_normalized(
                self._normalized_step7_row(row, pseudo_by_dev, rule_by_id, paragraph_by_ref)
            )
            for row in state_obj.get("deviations", [])
        ]
        if review_rows:
            sources.append(
                {
                    "key": "review_state",
                    "label": "Active review state",
                    "rows": review_rows,
                }
            )

        review_dir = paths.local_review_dir(study_id, self.output_dir)
        for snapshot_path in sorted(review_dir.glob("deviations_import_*.json")) if review_dir.exists() else []:
            try:
                snapshot_obj = json.loads(snapshot_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            version = snapshot_path.stem.replace("deviations_import_", "")
            snap_rows = [
                self._preview_row_from_normalized(
                    self._normalized_step7_row(row, pseudo_by_dev, rule_by_id, paragraph_by_ref)
                )
                for row in snapshot_obj.get("deviations", [])
            ]
            sources.append(
                {
                    "key": f"import_{version}",
                    "label": f"Import snapshot v{version}",
                    "rows": snap_rows,
                }
            )

        merged_paths = sorted(review_dir.glob("deviations_merged_*.json")) if review_dir.exists() else []
        for snapshot_path in merged_paths:
            try:
                snapshot_obj = json.loads(snapshot_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            version = snapshot_path.stem.replace("deviations_merged_", "")
            snap_rows = [
                self._preview_row_from_normalized(
                    self._normalized_step7_row(row, pseudo_by_dev, rule_by_id, paragraph_by_ref)
                )
                for row in snapshot_obj.get("deviations", [])
            ]
            sources.append(
                {
                    "key": f"merged_{version}",
                    "label": f"Merged snapshot {version}",
                    "rows": snap_rows,
                }
            )

        return {
            "studyId": study_id,
            "sources": sources,
            "stepStatuses": self._step_statuses(study_id),
        }

    def get_status(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        statuses = self._step_statuses(study_id)
        manifest = self._read_upload_manifest_obj(study_id)
        return {
            "studyId": study_id,
            "entryMode": self._get_entry_mode(study_id),
            "activeDeviationsSource": manifest.get("activeDeviationsSource"),
            "codingPhaseAccepted": bool(manifest.get("codingPhaseAccepted")),
            "codingPhaseAcceptedAt": manifest.get("codingPhaseAcceptedAt"),
            "importVersions": pipeline_v2.list_import_versions(study_id, self.output_dir),
            "steps": [
                {"stepId": step_id, "status": statuses.get(step_id, "pending")}
                for step_id in self._effective_step_order(study_id)
            ],
            "nextStepId": next(
                (
                    step_id
                    for step_id in self._effective_step_order(study_id)
                    if statuses.get(step_id) not in {"done", "skipped"}
                ),
                None,
            ),
        }

    def set_study_entry_mode(self, study_id: str, entry_mode: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        mode = (entry_mode or "").strip()
        if mode not in {ENTRY_MODE_EXTRACTED, ENTRY_MODE_IMPORTED_PD_SPEC}:
            raise UiApiError(
                "VALIDATION_ERROR",
                f"entryMode must be '{ENTRY_MODE_EXTRACTED}' or '{ENTRY_MODE_IMPORTED_PD_SPEC}'",
                400,
            )
        manifest = self._write_upload_manifest(study_id, entry_mode=mode)
        return {
            "studyId": study_id,
            "entryMode": manifest.get("entryMode"),
            "stepStatuses": self._step_statuses(study_id),
        }

    def accept_coding_phase(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        state_obj = self._load_state(study_id)
        deviations = list(state_obj.get("deviations", []))
        if deviations:
            incomplete = [
                str(row.get("deviation_id", ""))
                for row in deviations
                if str(row.get("status", "pending")) not in {"accepted", "rejected"}
            ]
            if incomplete:
                raise UiApiError(
                    "VALIDATION_ERROR",
                    f"All deviations must be accepted or rejected before continuing to coding. "
                    f"{len(incomplete)} still pending or to review.",
                    400,
                )
        manifest = self._write_upload_manifest(study_id, coding_phase_accepted=True)
        return {
            "studyId": study_id,
            "codingPhaseAccepted": bool(manifest.get("codingPhaseAccepted")),
            "codingPhaseAcceptedAt": manifest.get("codingPhaseAcceptedAt"),
            "stepStatuses": self._step_statuses(study_id),
        }

    def upload_pd_spec_workbook(
        self,
        study_id: str,
        workbook_bytes: bytes,
        *,
        file_name: str | None = None,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        if not workbook_bytes:
            raise UiApiError("VALIDATION_ERROR", "Workbook must not be empty", 400)
        max_mb = int(os.getenv("UI_UPLOAD_MAX_MB", "100"))
        max_bytes = max_mb * 1024 * 1024
        if len(workbook_bytes) > max_bytes:
            raise UiApiError("VALIDATION_ERROR", f"Workbook must be <= {max_mb}MB", 400)

        out_path = paths.local_pd_spec_workbook(study_id, self.output_dir)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(workbook_bytes)

        blob_service = blob_io.blob_service_from_env()
        container = blob_io.container_from_env()
        pd_spec_blob = paths.pd_spec_workbook_blob(study_id)
        blob_io.upload_blob_bytes(
            blob_service=blob_service,
            container_name=container,
            blob_path=pd_spec_blob,
            data=workbook_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self._mirror_upload(study_id, out_path)

        safe_name = (file_name or "").strip() or "pd_specifications.xlsx"
        manifest = self._write_upload_manifest(
            study_id,
            pd_spec_file_name=safe_name,
            pd_spec_size=len(workbook_bytes),
        )
        return {
            "studyId": study_id,
            "pdSpecPath": str(out_path),
            "pdSpecBlob": pd_spec_blob,
            "pdSpecFileName": manifest.get("pdSpecFileName"),
            "pdSpecSize": int(manifest.get("pdSpecSize") or len(workbook_bytes)),
            "entryMode": manifest.get("entryMode"),
            "stepStatuses": self._step_statuses(study_id),
        }

    def set_active_deviations_source(self, study_id: str, source_key: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        key = (source_key or "").strip()
        if not key:
            raise UiApiError("VALIDATION_ERROR", "activeDeviationsSource is required", 400)
        try:
            result = pipeline_v2.apply_active_deviations_source(study_id, self.output_dir, key)
        except ValueError as exc:
            raise UiApiError("VALIDATION_ERROR", str(exc), 400) from exc
        self._write_upload_manifest(study_id, active_deviations_source=key)
        return {
            "studyId": study_id,
            "activeDeviationsSource": key,
            "deviationCount": result.get("deviation_count", 0),
            "stepStatuses": self._step_statuses(study_id),
        }

    def get_import_versions(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        versions = pipeline_v2.list_import_versions(study_id, self.output_dir)
        manifest = self._read_upload_manifest_obj(study_id)
        sources: List[Dict[str, str]] = []
        for version in versions.get("imports", []):
            sources.append({"key": f"import_{version}", "label": f"Import {version}", "type": "import"})
        for version in versions.get("merged", []):
            sources.append({"key": f"merged_{version}", "label": f"Merged {version}", "type": "merged"})
        return {
            "studyId": study_id,
            "activeDeviationsSource": manifest.get("activeDeviationsSource"),
            "importVersions": versions,
            "sources": sources,
        }

    def sync_study(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        self._assert_safe_study_id(study_id)
        report = study_artifact_sync.sync_study(study_id, self.output_dir)
        return {
            "studyId": study_id,
            "sync": report.to_dict(),
            "stepStatuses": self._step_statuses(study_id),
        }

    def run_step(
        self,
        study_id: str,
        step_id: str,
        *,
        llm_instructions: str | None = None,
        llm_deployment: str | None = None,
        force: bool = False,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        allowed_steps = set(STEP_ORDER) | set(IMPORT_STEP_ORDER)
        if step_id not in allowed_steps:
            raise UiApiError("NOT_FOUND", f"Unknown stepId '{step_id}'", 404)

        statuses = self._step_statuses(study_id)
        self._assert_step_dependencies(statuses, step_id, study_id)

        if not force and self._step_artifact_complete(study_id, step_id):
            summary = "Already complete (skipped)"
            return {
                "studyId": study_id,
                "stepId": step_id,
                "summary": summary,
                "skipped": True,
                "stepStatuses": self._step_statuses(study_id),
            }

        extra = (llm_instructions or "").strip()
        stage_labels = {
            "index-protocol": "index",
            "acrf-split-toc": "acrf_split",
            "acrf-summary-text": "acrf_merge",
            "extract-rules": "rules",
            "extract-deviations": "deviations",
            "import-pd-spec-ground": "import_ground",
            "import-pd-spec-map": "import_map",
            "import-pd-spec-enrich": "import_enrich",
            "merge-pd-spec-imports": "import_merge",
            "review-and-finalize": "finalize",
        }
        active_run = self._active_run_entry(study_id)
        active_run_id = str(active_run.get("runId")) if active_run else ""
        self._write_pipeline_run_state(
            study_id,
            status="running",
            currentStage=stage_labels.get(step_id, step_id),
            currentSubStepId=step_id,
            message=f"Running {step_id}…",
            error="",
            startedAt=datetime.now(timezone.utc).isoformat(),
            finishedAt="",
            llmProgress=None,
            activeRunId=active_run_id,
        )
        self._append_pipeline_log(study_id, f"Starting step {step_id}")

        from pdcheck_factory import llm

        def _pipeline_log(message: str) -> None:
            self._append_pipeline_log(study_id, message)

        try:
            with llm.use_deployment(llm_deployment), llm.use_pipeline_log(_pipeline_log):
                summary = self._execute_run_step(study_id, step_id, extra=extra, force=force)
        except Exception as exc:  # noqa: BLE001
            self._append_pipeline_log(study_id, f"Step {step_id} failed: {exc}", level="error")
            self._write_pipeline_run_state(
                study_id,
                status="failed",
                message=f"Step {step_id} failed",
                error=str(exc),
                finishedAt=datetime.now(timezone.utc).isoformat(),
                llmProgress=None,
            )
            raise

        self._append_pipeline_log(study_id, summary)
        self._write_pipeline_run_state(
            study_id,
            status="done",
            currentStage="complete",
            currentSubStepId=step_id,
            message=summary,
            finishedAt=datetime.now(timezone.utc).isoformat(),
            llmProgress=None,
            activeRunId=active_run_id,
        )
        self._touch_active_run_after_step(study_id)

        return {
            "studyId": study_id,
            "stepId": step_id,
            "summary": summary,
            "stepStatuses": self._step_statuses(study_id),
        }

    def _execute_run_step(self, study_id: str, step_id: str, *, extra: str, force: bool = False) -> str:
        if not force and self._step_artifact_complete(study_id, step_id):
            return "Already complete (skipped)"

        progress_callback = self._make_llm_progress_callback(study_id)

        if step_id == "index-protocol":
            result = pipeline_v2.step2_protocol_paragraph_index(study_id, self.output_dir)
            summary = f"Indexed {len(result.get('paragraphs', []))} protocol paragraphs."
        elif step_id == "acrf-split-toc":
            from pdcheck_factory.cli import run_acrf_split_toc

            p = self._study_paths(study_id)
            if not p.acrf_source.exists():
                raise UiApiError("STEP_BLOCKED", f"Missing aCRF source markdown: {p.acrf_source}", 409)
            count, _manifest_path = run_acrf_split_toc(
                source_md=p.acrf_source,
                destination_dir=p.acrf_sections_toc_dir,
                write_manifest=True,
            )
            summary = f"Split aCRF markdown into {count} TOC section files."
            study_artifact_sync.mirror_upload_directory(study_id, self.output_dir, p.acrf_sections_toc_dir)
        elif step_id == "acrf-summary-text":
            result = pipeline_v2.step1_acrf_summary_text(
                study_id,
                self.output_dir,
                progress_callback=progress_callback,
            )
            summary = f"Merged aCRF summary text with {len(result.get('datasets', []))} datasets."
        elif step_id == "extract-rules":
            def _rules_log(message: str) -> None:
                self._append_pipeline_log(study_id, message)

            result = pipeline_v2.step3_extract_rules(
                study_id,
                self.output_dir,
                additional_instructions=extra,
                progress_callback=progress_callback,
                log_callback=_rules_log,
            )
            summary = f"Extracted {len(result.get('rules', []))} rules."
        elif step_id == "extract-deviations":
            if force:
                self._clear_deviation_extraction_artifacts(study_id)
            result = pipeline_v2.step4_5_extract_deviations(
                study_id,
                self.output_dir,
                additional_instructions=extra,
                progress_callback=progress_callback,
                force=force,
            )
            pipeline_v2.initialize_review_states(study_id, self.output_dir)
            summary = f"Extracted {len(result.get('deviations', []))} deviations and initialized review state."
        elif step_id == "import-pd-spec-ground":
            workbook_bytes = self._read_pd_spec_workbook_bytes(study_id)
            if not workbook_bytes:
                raise UiApiError(
                    "STEP_BLOCKED",
                    "Upload PD Specifications workbook before running import-pd-spec-ground.",
                    409,
                )
            result = pipeline_v2.run_import_pd_spec_grounding(
                study_id,
                self.output_dir,
                workbook_bytes=workbook_bytes,
            )
            version = result.get("import_version", "")
            self._write_upload_manifest(
                study_id,
                active_deviations_source=f"import_{version}",
            )
            summary = (
                f"Imported and grounded {len(result.get('deviations', []))} deviations "
                f"(version {version})."
            )
        elif step_id == "import-pd-spec-map":
            workbook_bytes = self._read_pd_spec_workbook_bytes(study_id)
            if not workbook_bytes:
                raise UiApiError(
                    "STEP_BLOCKED",
                    "Upload PD Specifications workbook before mapping to review.",
                    409,
                )
            result = pipeline_v2.run_import_pd_spec_map(
                study_id,
                self.output_dir,
                workbook_bytes=workbook_bytes,
                pd_spec_import_mode="map",
            )
            version = result.get("import_version", "")
            self._write_upload_manifest(
                study_id,
                entry_mode=ENTRY_MODE_IMPORTED_PD_SPEC,
                active_deviations_source=f"import_{version}",
                pd_spec_import_mode="map",
                review_display_source=review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC,
            )
            summary = (
                f"Mapped {len(result.get('deviations', []))} imported deviations to review "
                f"(version {version})."
            )
        elif step_id == "import-pd-spec-enrich":
            workbook_bytes = self._read_pd_spec_workbook_bytes(study_id)
            if not workbook_bytes:
                raise UiApiError(
                    "STEP_BLOCKED",
                    "Upload PD Specifications workbook before enrich.",
                    409,
                )
            result = pipeline_v2.run_import_pd_spec_enrich(
                study_id,
                self.output_dir,
                workbook_bytes=workbook_bytes,
                progress_callback=progress_callback,
            )
            version = result.get("import_version", "")
            self._write_upload_manifest(
                study_id,
                entry_mode=ENTRY_MODE_IMPORTED_PD_SPEC,
                active_deviations_source=f"import_{version}",
                pd_spec_import_mode="enrich",
                review_display_source=review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC,
            )
            count = result.get("deviation_count", len(result.get("deviations", [])))
            summary = (
                f"Enriched {count} imported deviations with protocol and aCRF analysis "
                f"(version {version})."
            )
        elif step_id == "merge-pd-spec-imports":
            merge_result = pipeline_v2.merge_imported_deviation_snapshots(study_id, self.output_dir)
            merged_version = merge_result.get("merged_version", "")
            pipeline_v2.apply_active_deviations_source(
                study_id,
                self.output_dir,
                f"merged_{merged_version}",
            )
            self._write_upload_manifest(
                study_id,
                active_deviations_source=f"merged_{merged_version}",
            )
            summary = (
                f"Merged import snapshots into {merged_version} "
                f"({merge_result.get('deviation_count', 0)} deviations)."
            )
        elif step_id == "review-and-finalize":
            validated_path = paths.local_deviations_validated_json(study_id, self.output_dir)
            if not validated_path.exists():
                review_state_path = paths.local_deviations_review_state(study_id, self.output_dir)
                if review_state_path.exists():
                    validated_path.parent.mkdir(parents=True, exist_ok=True)
                    validated_path.write_text(review_state_path.read_text(encoding="utf-8"), encoding="utf-8")
                    self._mirror_upload(study_id, validated_path)
                else:
                    raise UiApiError(
                        "STEP_BLOCKED",
                        "Missing deviation review state; run extract-deviations first.",
                        409,
                    )
            pseudo = pipeline_v2.step8_generate_pseudo_logic(study_id, self.output_dir)
            final = pipeline_v2.step10_finalize(study_id, self.output_dir)
            summary = (
                f"Generated pseudo logic for {len(pseudo.get('items', []))} accepted deviations and "
                f"finalized {len(final.get('items', []))} output rows."
            )
        else:
            raise UiApiError("STEP_BLOCKED", f"Step '{step_id}' must be run via dedicated endpoint.", 409)

        return summary

    def get_step_preview(self, study_id: str, step_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        allowed_steps = set(STEP_ORDER) | set(IMPORT_STEP_ORDER)
        if step_id not in allowed_steps:
            raise UiApiError("NOT_FOUND", f"Unknown stepId '{step_id}'", 404)

        p = self._study_paths(study_id)
        previews: List[Dict[str, Any]] = []

        if step_id == "index-protocol":
            previews.append(
                {
                    "title": "Paragraph index preview",
                    "body": self._read_excerpt(p.paragraph_index),
                    "highlight": True,
                }
            )
        elif step_id == "acrf-split-toc":
            previews.append(
                {
                    "title": "aCRF sections_toc directory",
                    "body": str(p.acrf_sections_toc_dir),
                    "highlight": True,
                }
            )
            if p.acrf_sections_toc_dir.exists():
                section_files = sorted(p.acrf_sections_toc_dir.glob("*.md"))[:30]
                previews.append(
                    {
                        "title": "Section files",
                        "body": "\n".join(file.name for file in section_files)
                        or "No section markdown files found.",
                    }
                )
        elif step_id == "acrf-summary-text":
            previews.append(
                {
                    "title": "aCRF merged summary preview",
                    "body": self._read_excerpt(p.acrf_summary_text_merged),
                    "highlight": True,
                }
            )
        elif step_id == "extract-rules":
            previews.append(
                {
                    "title": "Rules preview",
                    "body": self._read_excerpt(p.rules_parsed),
                    "highlight": True,
                }
            )
        elif step_id == "extract-deviations":
            previews.append(
                {
                    "title": "Deviations preview",
                    "body": self._read_excerpt(p.deviations_parsed),
                    "highlight": True,
                }
            )
            previews.append(
                {
                    "title": "Review state preview",
                    "body": self._read_excerpt(p.deviations_review_state),
                }
            )
        elif step_id == "import-pd-spec-ground":
            review_dir = paths.local_review_dir(study_id, self.output_dir)
            import_files = sorted(review_dir.glob("deviations_import_*.json")) if review_dir.exists() else []
            previews.append(
                {
                    "title": "Import snapshots",
                    "body": "\n".join(file.name for file in import_files) or "No import snapshots yet.",
                    "highlight": True,
                }
            )
            previews.append(
                {
                    "title": "Active review state",
                    "body": self._read_excerpt(p.deviations_review_state),
                }
            )
        elif step_id == "merge-pd-spec-imports":
            review_dir = paths.local_review_dir(study_id, self.output_dir)
            merged_files = sorted(review_dir.glob("deviations_merged_*.json")) if review_dir.exists() else []
            previews.append(
                {
                    "title": "Merged snapshots",
                    "body": "\n".join(file.name for file in merged_files) or "No merged snapshots yet.",
                    "highlight": True,
                }
            )
        elif step_id == "review-and-finalize":
            previews.append(
                {
                    "title": "Final JSON preview",
                    "body": self._read_excerpt(p.final_json),
                    "highlight": True,
                }
            )
            previews.append(
                {
                    "title": "Final XLSX path",
                    "body": str(p.final_xlsx) if p.final_xlsx.exists() else "No final workbook generated yet.",
                }
            )

        run_state = self._read_pipeline_run_state(study_id)
        is_running = run_state.get("status") == "running"
        partial = is_running and run_state.get("currentSubStepId") == step_id
        item_count = 0
        if step_id == "extract-rules" and p.rules_parsed.is_file():
            try:
                rules_obj = read_json(p.rules_parsed)
                item_count = len(rules_obj.get("rules", []))
            except (json.JSONDecodeError, OSError, ValueError):
                item_count = 0
        elif step_id == "extract-deviations" and p.deviations_parsed.is_file():
            try:
                dev_obj = read_json(p.deviations_parsed)
                item_count = len(dev_obj.get("deviations", []))
            except (json.JSONDecodeError, OSError, ValueError):
                item_count = 0

        return {
            "studyId": study_id,
            "stepId": step_id,
            "previews": previews,
            "stepStatuses": self._step_statuses(study_id),
            "partial": partial,
            "itemCount": item_count,
        }

    def _pd_spec_workbook_available(self, study_id: str) -> bool:
        if paths.local_pd_spec_workbook(study_id, self.output_dir).is_file():
            return True
        try:
            return bool(self._read_pd_spec_workbook_bytes(study_id))
        except Exception:  # noqa: BLE001
            return False

    def get_step7_review_sources(self, study_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        manifest = self._read_upload_manifest_obj(study_id)
        selected = self._resolve_review_source(study_id, None)
        sources: List[Dict[str, Any]] = []

        generated_path = paths.local_deviations_parsed_json(study_id, self.output_dir)
        if generated_path.is_file():
            self._ensure_review_source_state(study_id, review_sources.REVIEW_SOURCE_GENERATED)
            gen_state = read_json(
                review_sources.review_state_path(
                    study_id, self.output_dir, review_sources.REVIEW_SOURCE_GENERATED
                )
            )
            sources.append(
                {
                    "key": review_sources.REVIEW_SOURCE_GENERATED,
                    "label": review_sources.REVIEW_SOURCE_LABELS[review_sources.REVIEW_SOURCE_GENERATED],
                    "available": True,
                    "rowCount": len(gen_state.get("deviations", [])),
                }
            )

        if self._pd_spec_workbook_available(study_id):
            self._ensure_review_source_state(study_id, review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC)
            imp_state = read_json(
                review_sources.review_state_path(
                    study_id, self.output_dir, review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC
                )
            )
            sources.append(
                {
                    "key": review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC,
                    "label": review_sources.REVIEW_SOURCE_LABELS[
                        review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC
                    ],
                    "available": True,
                    "rowCount": len(imp_state.get("deviations", [])),
                }
            )

        if self._pd_spec_enrich_done(study_id):
            self._ensure_review_source_state(study_id, review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC)
            enr_state = read_json(
                review_sources.review_state_path(
                    study_id, self.output_dir, review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
                )
            )
            sources.append(
                {
                    "key": review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC,
                    "label": review_sources.REVIEW_SOURCE_LABELS[
                        review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
                    ],
                    "available": True,
                    "rowCount": len(enr_state.get("deviations", [])),
                }
            )

        available_keys = {item["key"] for item in sources}
        if selected not in available_keys and sources:
            selected = str(sources[0]["key"])

        return {
            "studyId": study_id,
            "sources": sources,
            "selectedSource": selected,
            "stepStatuses": self._step_statuses(study_id),
        }

    def set_step7_review_display_source(self, study_id: str, review_source: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        self._ensure_review_source_state(study_id, source)
        self._write_upload_manifest(study_id, review_display_source=source)
        payload = self.get_step7_review_sources(study_id)
        payload["selectedSource"] = source
        return payload

    def get_step7_deviations(self, study_id: str, *, review_source: str | None = None) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        state_obj = self._load_state(study_id, source)
        pseudo_obj = self._load_pseudo_state(study_id)
        rules_obj = self._load_rules(study_id)
        paragraph_by_ref = self._load_paragraph_index(study_id)
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        rows = [
            self._normalized_step7_row(
                row,
                pseudo_by_dev,
                rule_by_id,
                paragraph_by_ref,
                study_id=study_id,
                review_source=source,
            )
            for row in state_obj.get("deviations", [])
        ]
        return {
            "studyId": study_id,
            "reviewSource": source,
            "columns": [
                "rule_id",
                "deviation_id",
                "rule_title",
                "deviation_text",
                "paragraph_refs",
                "pseudo_logic",
            ],
            "rows": rows,
            "stepStatuses": self._step_statuses(study_id),
        }

    def create_step7_deviation(
        self, study_id: str, payload: Dict[str, Any], *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        new_row = self._normalize_deviation_payload(payload, default_source="imported")
        state_obj = self._load_state(study_id, source)
        rows = list(state_obj.get("deviations", []))
        if any(str(row.get("deviation_id", "")) == new_row["deviation_id"] for row in rows):
            raise UiApiError("VALIDATION_ERROR", f"Duplicate deviation_id '{new_row['deviation_id']}'", 400)
        rows.append(new_row)
        state_obj["deviations"] = rows
        self._persist_state(
            study_id,
            state_obj,
            self._audit(study_id, action="create_deviation", target_id=new_row["deviation_id"], updated_rows=1),
            review_source=source,
        )
        return self.get_step7_deviations(study_id, review_source=source)

    def patch_step7_deviation_fields(
        self,
        study_id: str,
        deviation_id: str,
        payload: Dict[str, Any],
        *,
        review_source: str | None = None,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        dev_id = str(deviation_id).strip()
        state_obj = self._load_state(study_id, source)
        rows = list(state_obj.get("deviations", []))
        row = next((item for item in rows if str(item.get("deviation_id", "")) == dev_id), None)
        if row is None:
            raise UiApiError("NOT_FOUND", f"Unknown deviationId '{dev_id}'", 404)
        merged = dict(row)
        for source_key, target_key in [
            ("rule_id", "rule_id"),
            ("ruleId", "rule_id"),
            ("text", "text"),
            ("deviation_text", "text"),
            ("deviationText", "text"),
            ("data_support_note", "data_support_note"),
            ("dataSupportNote", "data_support_note"),
            ("dm_comment", "dm_comment"),
            ("dmComment", "dm_comment"),
            ("status", "status"),
        ]:
            if source_key in payload:
                merged[target_key] = payload[source_key]
        if "paragraph_refs" in payload or "paragraphRefs" in payload:
            merged["paragraph_refs"] = payload.get("paragraph_refs") or payload.get("paragraphRefs")
        default_source = pd_spec_field(row, "entry_source", default=str(row.get("entry_source", "extracted")))
        normalized = self._normalize_deviation_payload(merged, default_source=default_source)
        normalized["deviation_id"] = dev_id
        if "original_deviation_text" in row:
            normalized["original_deviation_text"] = row["original_deviation_text"]
        if "suggested_deviation_text" in row:
            normalized["suggested_deviation_text"] = row["suggested_deviation_text"]
        if isinstance(row.get("pd_spec_import"), dict):
            normalized["pd_spec_import"] = dict(row["pd_spec_import"])
            for field in (
                "protocol_deviation_category",
                "protocol_deviation_sub_category",
                "classification",
                "data_source",
                "manual_or_programmable",
                "programming_status",
                "programmer_comments",
                "reviewer_comments",
                "aa_comment",
                "grounding_error",
                "pseudo_logic_seed",
                "enrichment_status",
                "enrichment_summary",
                "entry_source",
            ):
                if field in row and field not in normalized.get("pd_spec_import", {}):
                    normalized.setdefault("pd_spec_import", {})[field] = row[field]
        state_obj = self._replace_row(state_obj, normalized)
        self._persist_state(
            study_id,
            state_obj,
            self._audit(study_id, action="update_deviation", target_id=dev_id, updated_rows=1),
            review_source=source,
        )
        return self._single_step7_deviation_response(study_id, normalized, review_source=source)

    def delete_step7_deviation(
        self, study_id: str, deviation_id: str, *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        dev_id = str(deviation_id).strip()
        state_obj = self._load_state(study_id, source)
        rows = list(state_obj.get("deviations", []))
        next_rows = [row for row in rows if str(row.get("deviation_id", "")) != dev_id]
        if len(next_rows) == len(rows):
            raise UiApiError("NOT_FOUND", f"Unknown deviationId '{dev_id}'", 404)
        state_obj["deviations"] = next_rows
        self._persist_state(
            study_id,
            state_obj,
            self._audit(study_id, action="delete_deviation", target_id=dev_id, updated_rows=1),
            review_source=source,
        )

        pseudo_obj = self._load_pseudo_state(study_id)
        pseudo_items = [item for item in pseudo_obj.get("items", []) if str(item.get("deviation_id", "")) != dev_id]
        if len(pseudo_items) != len(pseudo_obj.get("items", [])):
            pseudo_obj["items"] = pseudo_items
            write_json(paths.local_pseudo_logic_review_state(study_id, self.output_dir), pseudo_obj)
            write_json(paths.local_pseudo_logic_validated_json(study_id, self.output_dir), pseudo_obj)
            self._mirror_upload(
                study_id,
                paths.local_pseudo_logic_review_state(study_id, self.output_dir),
                paths.local_pseudo_logic_validated_json(study_id, self.output_dir),
            )
        return self.get_step7_deviations(study_id, review_source=source)

    def import_step7_deviations_xlsx(
        self, study_id: str, workbook_bytes: bytes, *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        if not workbook_bytes:
            raise UiApiError("VALIDATION_ERROR", "Workbook must not be empty", 400)
        try:
            workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as exc:  # noqa: BLE001
            raise UiApiError("VALIDATION_ERROR", "Workbook must be a readable .xlsx file", 400) from exc

        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise UiApiError("VALIDATION_ERROR", "Workbook must include a header row", 400)
        header_map = {str(value or "").strip().lower(): index for index, value in enumerate(headers)}

        def cell(row_values: tuple[Any, ...], *names: str) -> Any:
            for name in names:
                idx = header_map.get(name)
                if idx is not None and idx < len(row_values):
                    return row_values[idx]
            return ""

        imported: List[Dict[str, Any]] = []
        for row_values in rows_iter:
            if not row_values or not any(value is not None and str(value).strip() for value in row_values):
                continue
            imported.append(
                self._normalize_deviation_payload(
                    {
                        "deviation_id": cell(row_values, "deviation_id", "deviationid"),
                        "rule_id": cell(row_values, "rule_id", "ruleid"),
                        "text": cell(row_values, "text", "deviation_text", "deviationtext"),
                        "paragraph_refs": cell(row_values, "paragraph_refs", "paragraphrefs"),
                        "data_support_note": cell(row_values, "data_support_note", "datasupportnote"),
                        "dm_comment": cell(row_values, "dm_comment", "dmcomment"),
                        "status": cell(row_values, "status") or "pending",
                    },
                    default_source="imported",
                )
            )

        if not imported:
            raise UiApiError("VALIDATION_ERROR", "Workbook did not contain any deviation rows", 400)

        source = self._resolve_review_source(study_id, review_source)
        state_obj = self._load_state(study_id, source)
        existing_ids = {str(row.get("deviation_id", "")) for row in state_obj.get("deviations", [])}
        imported_ids = [row["deviation_id"] for row in imported]
        duplicate_ids = sorted({dev_id for dev_id in imported_ids if imported_ids.count(dev_id) > 1 or dev_id in existing_ids})
        if duplicate_ids:
            raise UiApiError("VALIDATION_ERROR", f"Duplicate deviation_id values: {', '.join(duplicate_ids)}", 400)

        state_obj["deviations"] = list(state_obj.get("deviations", [])) + imported
        self._persist_state(
            study_id,
            state_obj,
            self._audit(study_id, action="import_deviations", target_id="xlsx", updated_rows=len(imported)),
            review_source=source,
        )
        payload = self.get_step7_deviations(study_id, review_source=source)
        payload["imported"] = len(imported)
        return payload

    @staticmethod
    def _format_supporting_sentences(sentences: List[Dict[str, Any]]) -> str:
        parts: List[str] = []
        for item in sentences:
            ref = str(item.get("ref", "")).strip()
            text = str(item.get("text", "")).strip()
            if ref and text:
                parts.append(f"{ref}: {text}")
            elif ref:
                parts.append(ref)
            elif text:
                parts.append(text)
        return " | ".join(parts)

    @staticmethod
    def _format_programmable(value: Any) -> str:
        if value is True:
            return "true"
        if value is False:
            return "false"
        return ""

    def export_step7_deviations_xlsx(self, study_id: str, *, review_source: str | None = None) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        payload = self.get_step7_deviations(study_id, review_source=review_source)
        rows = list(payload.get("rows", []))
        rules_obj = self._load_rules(study_id)
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        exported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        timestamp_slug = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        file_name = f"{study_id}_deviations_review_{timestamp_slug}.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Deviations"
        sheet.append(STEP7_EXPORT_COLUMNS)
        for row in rows:
            rule = rule_by_id.get(str(row.get("rule_id", "")), {})
            rule_refs = rule.get("paragraph_refs") or []
            sheet.append(
                [
                    study_id,
                    exported_at,
                    row.get("rule_id", ""),
                    row.get("rule_title", ""),
                    row.get("rule_text", ""),
                    ", ".join(str(ref) for ref in rule_refs),
                    row.get("deviation_id", ""),
                    row.get("deviation_text", ""),
                    row.get("paragraph_refs_text", "") or ", ".join(str(ref) for ref in row.get("paragraph_refs", [])),
                    self._format_supporting_sentences(list(row.get("supporting_sentences", []))),
                    row.get("data_support_note", ""),
                    row.get("status", ""),
                    row.get("dm_comment", ""),
                    row.get("entry_source", ""),
                    self._format_programmable(row.get("programmable")),
                    row.get("programmability_note", ""),
                    row.get("pseudo_logic", ""),
                ]
            )

        summary = workbook.create_sheet("Summary")
        status_counts = {"pending": 0, "to_review": 0, "accepted": 0, "rejected": 0}
        for row in rows:
            status = str(row.get("status", "pending"))
            if status in status_counts:
                status_counts[status] += 1
        summary.append(["field", "value"])
        summary.append(["study_id", study_id])
        summary.append(["exported_at", exported_at])
        summary.append(["total_deviations", len(rows)])
        summary.append(["accepted", status_counts["accepted"]])
        summary.append(["to_review", status_counts["to_review"]])
        summary.append(["rejected", status_counts["rejected"]])
        summary.append(["pending", status_counts["pending"]])

        out_path = paths.local_deviations_review_export_xlsx(study_id, self.output_dir)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_path)

        buffer = BytesIO()
        workbook.save(buffer)
        content = buffer.getvalue()
        return {
            "studyId": study_id,
            "fileName": file_name,
            "filePath": str(out_path),
            "rowCount": len(rows),
            "exportedAt": exported_at,
            "content": content,
        }

    def export_step7_deviations_coding_xlsx(
        self, study_id: str, *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        payload = self.get_step7_deviations(study_id, review_source=review_source)
        rows = list(payload.get("rows", []))
        exported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        timestamp_slug = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        file_name = f"{study_id}_company_pds_{timestamp_slug}.xlsx"

        out_path = paths.local_deviations_coding_export_xlsx(study_id, self.output_dir)
        coding_workbook_export.write_coding_workbook_xlsx(
            rows,
            out_path,
            study_id=study_id,
            exported_at=exported_at,
        )

        buffer = BytesIO()
        buffer.write(out_path.read_bytes())
        content = buffer.getvalue()
        return {
            "studyId": study_id,
            "fileName": file_name,
            "filePath": str(out_path),
            "rowCount": len(rows),
            "exportedAt": exported_at,
            "content": content,
        }

    def _single_step7_deviation_response(
        self, study_id: str, row: Dict[str, Any], *, review_source: str | None = None
    ) -> Dict[str, Any]:
        pseudo_obj = self._load_pseudo_state(study_id)
        rules_obj = self._load_rules(study_id)
        paragraph_by_ref = self._load_paragraph_index(study_id)
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        dev_id = str(row.get("deviation_id", ""))
        source = self._resolve_review_source(study_id, review_source)
        return {
            "studyId": study_id,
            "reviewSource": source,
            "deviationId": dev_id,
            "row": self._normalized_step7_row(
                row,
                pseudo_by_dev,
                rule_by_id,
                paragraph_by_ref,
                study_id=study_id,
                review_source=source,
            ),
            "stepStatuses": self._step_statuses(study_id),
        }

    def create_step7_rule(self, study_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        rule = self._normalized_rule_payload(payload)
        rules_obj = self._load_rules(study_id)
        rules = list(rules_obj.get("rules", []))
        if any(str(item.get("rule_id", "")) == rule["rule_id"] for item in rules):
            raise UiApiError("VALIDATION_ERROR", f"Duplicate rule_id '{rule['rule_id']}'", 400)
        rules.append(rule)
        rules_obj["rules"] = rules
        self._save_rules(study_id, rules_obj)
        return {"studyId": study_id, "rule": rule, "stepStatuses": self._step_statuses(study_id)}

    def update_step7_rule(self, study_id: str, rule_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        rid = str(rule_id).strip()
        rules_obj = self._load_rules(study_id)
        rules = list(rules_obj.get("rules", []))
        idx = next((i for i, item in enumerate(rules) if str(item.get("rule_id", "")) == rid), None)
        if idx is None:
            raise UiApiError("NOT_FOUND", f"Unknown ruleId '{rid}'", 404)
        merged = dict(rules[idx])
        merged.update(payload)
        updated = self._normalized_rule_payload({**merged, "rule_id": rid})
        rules[idx] = updated
        rules_obj["rules"] = rules
        self._save_rules(study_id, rules_obj)
        return {"studyId": study_id, "rule": updated, "stepStatuses": self._step_statuses(study_id)}

    def delete_step7_rule(
        self, study_id: str, rule_id: str, *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        rid = str(rule_id).strip()
        source = self._resolve_review_source(study_id, review_source)
        state_obj = self._load_state(study_id, source)
        if any(str(row.get("rule_id", "")) == rid for row in state_obj.get("deviations", [])):
            raise UiApiError("VALIDATION_ERROR", f"Rule '{rid}' is used by one or more deviations", 400)
        rules_obj = self._load_rules(study_id)
        rules = list(rules_obj.get("rules", []))
        next_rules = [rule for rule in rules if str(rule.get("rule_id", "")) != rid]
        if len(next_rules) == len(rules):
            raise UiApiError("NOT_FOUND", f"Unknown ruleId '{rid}'", 404)
        rules_obj["rules"] = next_rules
        self._save_rules(study_id, rules_obj)
        return {"studyId": study_id, "deletedRuleId": rid, "stepStatuses": self._step_statuses(study_id)}

    def get_step7_deviation_chat(self, study_id: str, deviation_id: str) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        dev_id = str(deviation_id).strip()
        if not dev_id:
            raise UiApiError("VALIDATION_ERROR", "deviationId is required", 400)
        chat_obj = self._load_chat_state(study_id)
        dev_chat = chat_obj.get("deviations", {}).get(dev_id, {})
        return {
            "studyId": study_id,
            "deviationId": dev_id,
            "messages": list(dev_chat.get("messages", []))[-25:],
        }

    def refine_step7_deviation(
        self,
        *,
        study_id: str,
        deviation_id: str,
        dm_comment: str,
        run_revision_cycle: bool = True,
        also_generate_pseudo: bool = False,
        review_source: str | None = None,
        llm_deployment: str | None = None,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        dev_id = str(deviation_id).strip()
        if not dev_id:
            raise UiApiError("VALIDATION_ERROR", "deviationId is required", 400)
        comment = str(dm_comment or "")

        state_obj = self._load_state(study_id, source)
        rows = list(state_obj.get("deviations", []))
        row = next((item for item in rows if str(item.get("deviation_id", "")) == dev_id), None)
        if row is None:
            raise UiApiError("NOT_FOUND", f"Unknown deviationId '{dev_id}'", 404)

        chat_obj = self._load_chat_state(study_id)
        self._append_chat_message(chat_obj, dev_id, role="dm", text=comment.strip() or "(empty)")
        dev_chat = chat_obj.get("deviations", {}).get(dev_id, {})
        prior_messages = list(dev_chat.get("messages", []))[:-1]
        chat_history = [
            {"role": str(m.get("role", "")), "text": str(m.get("text", ""))}
            for m in prior_messages[-10:]
        ]
        from pdcheck_factory import llm

        try:
            with llm.use_deployment(llm_deployment):
                revised_row, audit = pipeline_v2.refine_single_deviation_with_comment(
                    study_id=study_id,
                    output_dir=self.output_dir,
                    row=row,
                    dm_comment=comment,
                    run_revision_cycle=run_revision_cycle,
                    chat_history=chat_history,
                    also_generate_pseudo=also_generate_pseudo,
                )
            assistant_text = str(audit.get("assistant_message", "")).strip()
            if not assistant_text:
                assistant_text = "Processed your message."
            self._append_chat_message(chat_obj, dev_id, role="assistant", text=assistant_text)
        except Exception as exc:
            self._append_chat_message(chat_obj, dev_id, role="assistant", text=f"Refinement failed: {exc}")
            self._save_chat_state(study_id, chat_obj)
            raise UiApiError("REFINE_FAILED", str(exc), 500) from exc

        state_obj = self._replace_row(state_obj, revised_row)
        self._persist_state(study_id, state_obj, audit, review_source=source)
        self._save_chat_state(study_id, chat_obj)

        pseudo_obj = self._load_pseudo_state(study_id)
        pseudo_item = audit.get("pseudo_item")
        if isinstance(pseudo_item, dict) and pseudo_item.get("deviation_id"):
            items = list(pseudo_obj.get("items", []))
            replaced = False
            for idx, existing in enumerate(items):
                if str(existing.get("deviation_id", "")) == dev_id:
                    items[idx] = pseudo_item
                    replaced = True
                    break
            if not replaced:
                items.append(pseudo_item)
            pseudo_obj["schema_version"] = pseudo_obj.get("schema_version", "1.0.0")
            pseudo_obj["study_id"] = study_id
            pseudo_obj["generated_at"] = datetime.now(timezone.utc).isoformat()
            pseudo_obj["items"] = items
            write_json(paths.local_pseudo_logic_review_state(study_id, self.output_dir), pseudo_obj)
            write_json(paths.local_pseudo_logic_validated_json(study_id, self.output_dir), pseudo_obj)
            self._mirror_upload(
                study_id,
                paths.local_pseudo_logic_review_state(study_id, self.output_dir),
                paths.local_pseudo_logic_validated_json(study_id, self.output_dir),
            )

        rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}

        agent_reason = ""
        agent_block = audit.get("agent") or {}
        decision_block = agent_block.get("decision") if isinstance(agent_block, dict) else None
        if isinstance(decision_block, dict):
            agent_reason = str(decision_block.get("reason", "")).strip()

        return {
            "studyId": study_id,
            "reviewSource": source,
            "deviationId": dev_id,
            "row": self._normalized_step7_row(
                revised_row,
                pseudo_by_dev,
                rule_by_id,
                study_id=study_id,
                review_source=source,
            ),
            "messages": list(chat_obj.get("deviations", {}).get(dev_id, {}).get("messages", []))[-25:],
            "audit": audit,
            "responseType": str(audit.get("response_type", "")),
            "agentReason": agent_reason,
            "missingCaveats": list(audit.get("missing_caveats", [])),
            "stepStatuses": self._step_statuses(study_id),
        }

    def update_step7_deviation(
        self,
        *,
        study_id: str,
        deviation_id: str,
        status: str,
        dm_comment: str | None = None,
        review_source: str | None = None,
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        dev_id = str(deviation_id).strip()
        next_status = str(status).strip().lower()
        if next_status not in {"pending", "to_review", "accepted", "rejected"}:
            raise UiApiError("VALIDATION_ERROR", "Invalid status value", 400)

        state_obj = self._load_state(study_id, source)
        row = next((item for item in state_obj.get("deviations", []) if str(item.get("deviation_id", "")) == dev_id), None)
        if row is None:
            raise UiApiError("NOT_FOUND", f"Unknown deviationId '{dev_id}'", 404)
        updated = dict(row)
        updated["status"] = next_status
        if dm_comment is not None:
            updated["dm_comment"] = dm_comment
        audit = {
            "study_id": study_id,
            "review_type": "deviations",
            "deviation_id": dev_id,
            "updated_rows": 1,
            "revised_rows": 0,
            "run_revision_cycle": False,
        }
        state_obj = self._replace_row(state_obj, updated)
        self._persist_state(study_id, state_obj, audit, review_source=source)

        pseudo_obj = self._load_pseudo_state(study_id)
        rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        return {
            "studyId": study_id,
            "reviewSource": source,
            "deviationId": dev_id,
            "row": self._normalized_step7_row(
                updated,
                pseudo_by_dev,
                rule_by_id,
                study_id=study_id,
                review_source=source,
            ),
            "stepStatuses": self._step_statuses(study_id),
        }

    def generate_step7_pseudo_logic_for_deviation(
        self, study_id: str, deviation_id: str, *, review_source: str | None = None
    ) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        dev_id = str(deviation_id).strip()
        if not dev_id:
            raise UiApiError("VALIDATION_ERROR", "deviationId is required", 400)

        state_obj = self._load_state(study_id, source)
        row = next(
            (item for item in state_obj.get("deviations", []) if str(item.get("deviation_id", "")) == dev_id),
            None,
        )
        if row is None:
            raise UiApiError("NOT_FOUND", f"Unknown deviationId '{dev_id}'", 404)
        if str(row.get("status", "")) != "accepted":
            raise UiApiError(
                "STEP_BLOCKED",
                "Pseudo logic can only be generated for deviations with status='accepted'.",
                409,
            )

        try:
            pseudo_item = pipeline_v2.generate_pseudo_logic_for_deviation(
                study_id=study_id,
                output_dir=self.output_dir,
                deviation=row,
            )
        except Exception as exc:  # noqa: BLE001
            raise UiApiError("PSEUDO_LOGIC_FAILED", str(exc), 500) from exc

        pseudo_obj = self._load_pseudo_state(study_id)
        items = list(pseudo_obj.get("items", []))
        replaced = False
        for idx, existing in enumerate(items):
            if str(existing.get("deviation_id", "")) == dev_id:
                items[idx] = pseudo_item
                replaced = True
                break
        if not replaced:
            items.append(pseudo_item)
        pseudo_obj["schema_version"] = pseudo_obj.get("schema_version", "1.0.0")
        pseudo_obj["study_id"] = study_id
        pseudo_obj["generated_at"] = datetime.now(timezone.utc).isoformat()
        pseudo_obj["items"] = items
        write_json(paths.local_pseudo_logic_review_state(study_id, self.output_dir), pseudo_obj)
        write_json(paths.local_pseudo_logic_validated_json(study_id, self.output_dir), pseudo_obj)
        self._mirror_upload(
            study_id,
            paths.local_pseudo_logic_review_state(study_id, self.output_dir),
            paths.local_pseudo_logic_validated_json(study_id, self.output_dir),
        )

        rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in items}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        return {
            "studyId": study_id,
            "reviewSource": source,
            "deviationId": dev_id,
            "row": self._normalized_step7_row(
                row,
                pseudo_by_dev,
                rule_by_id,
                study_id=study_id,
                review_source=source,
            ),
            "stepStatuses": self._step_statuses(study_id),
        }

    def accept_step7_deviations_bulk(self, study_id: str, *, review_source: str | None = None) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)
        state_obj = self._load_state(study_id, source)
        rows = list(state_obj.get("deviations", []))
        accepted_count = 0
        for row in rows:
            status = str(row.get("status", "pending"))
            if status in {"accepted", "rejected"}:
                continue
            row["status"] = "accepted"
            accepted_count += 1
        if accepted_count == 0:
            pseudo_obj = self._load_pseudo_state(study_id)
            rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
            pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
            rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
            normalized = [
                self._normalized_step7_row(
                    row,
                    pseudo_by_dev,
                    rule_by_id,
                    study_id=study_id,
                    review_source=source,
                )
                for row in state_obj.get("deviations", [])
            ]
            return {
                "studyId": study_id,
                "reviewSource": source,
                "accepted": 0,
                "rows": normalized,
                "stepStatuses": self._step_statuses(study_id),
            }

        state_obj["deviations"] = rows
        self._persist_state(
            study_id,
            state_obj,
            self._audit(study_id, action="accept_all_deviations", target_id="bulk", updated_rows=accepted_count),
            review_source=source,
        )

        pseudo_obj = self._load_pseudo_state(study_id)
        rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in pseudo_obj.get("items", [])}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        normalized = [
            self._normalized_step7_row(
                row,
                pseudo_by_dev,
                rule_by_id,
                study_id=study_id,
                review_source=source,
            )
            for row in state_obj.get("deviations", [])
        ]
        return {
            "studyId": study_id,
            "reviewSource": source,
            "accepted": accepted_count,
            "rows": normalized,
            "stepStatuses": self._step_statuses(study_id),
        }

    def generate_step7_pseudo_logic_bulk(self, study_id: str, *, review_source: str | None = None) -> Dict[str, Any]:
        study_id = self._require_study_id(study_id)
        source = self._resolve_review_source(study_id, review_source)

        validated_path = paths.local_deviations_validated_json(study_id, self.output_dir)
        if source == review_sources.REVIEW_SOURCE_GENERATED:
            generated_review = review_sources.review_state_path(
                study_id, self.output_dir, review_sources.REVIEW_SOURCE_GENERATED
            )
            if not validated_path.exists():
                if not generated_review.is_file():
                    legacy = paths.local_deviations_review_state(study_id, self.output_dir)
                    if not legacy.is_file():
                        raise UiApiError(
                            "STEP_BLOCKED",
                            "Missing deviation review state; run extract-deviations first.",
                            409,
                        )
                    seed_path = legacy
                else:
                    seed_path = generated_review
                validated_path.parent.mkdir(parents=True, exist_ok=True)
                validated_path.write_text(seed_path.read_text(encoding="utf-8"), encoding="utf-8")
                self._mirror_upload(study_id, validated_path)

        self._write_pipeline_run_state(
            study_id,
            status="running",
            currentStage="pseudo_logic",
            currentSubStepId="pseudo-logic-bulk",
            message="Generating pseudo logic for accepted deviations…",
            error="",
            llmProgress=None,
        )
        progress_callback = self._make_llm_progress_callback(study_id)
        try:
            pseudo_out = pipeline_v2.step8_generate_pseudo_logic(
                study_id,
                self.output_dir,
                progress_callback=progress_callback,
            )
        except Exception as exc:  # noqa: BLE001
            self._write_pipeline_run_state(
                study_id,
                status="failed",
                message="Pseudo logic generation failed",
                error=str(exc),
                llmProgress=None,
            )
            raise UiApiError("PSEUDO_LOGIC_FAILED", str(exc), 500) from exc
        self._write_pipeline_run_state(
            study_id,
            status="done",
            currentStage="complete",
            currentSubStepId="pseudo-logic-bulk",
            message=f"Generated pseudo logic for {len(pseudo_out.get('items', []))} deviations.",
            llmProgress=None,
        )

        items = list(pseudo_out.get("items", []))
        state_obj = self._load_state(study_id, source)
        rules_obj = read_json(paths.local_rules_parsed_json(study_id, self.output_dir))
        pseudo_by_dev = {str(item.get("deviation_id", "")): item for item in items}
        rule_by_id = {str(rule.get("rule_id", "")): rule for rule in rules_obj.get("rules", [])}
        rows = [
            self._normalized_step7_row(
                row,
                pseudo_by_dev,
                rule_by_id,
                study_id=study_id,
                review_source=source,
            )
            for row in state_obj.get("deviations", [])
        ]
        return {
            "studyId": study_id,
            "reviewSource": source,
            "generated": len(items),
            "rows": rows,
            "stepStatuses": self._step_statuses(study_id),
        }


def parse_json_body(raw: bytes) -> Dict[str, Any]:
    try:
        parsed = json.loads(raw.decode("utf-8"))
    except Exception as exc:  # noqa: BLE001
        raise UiApiError("BAD_JSON", "Request body must be valid JSON", 400) from exc
    if not isinstance(parsed, dict):
        raise UiApiError("BAD_JSON", "JSON body must be an object", 400)
    return parsed
