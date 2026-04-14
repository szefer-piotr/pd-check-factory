"""Step 2 DM review workbook round-trip and finalization helpers."""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pdcheck_factory.json_util import load_schema, read_json, validate, write_json

VALIDATION_STATUSES = {"accepted", "to_review", "rejected"}
YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

HEADERS = [
    "row_key",
    "rule_id",
    "deviation_id",
    "rule_title",
    "atomic_requirement",
    "scenario_description",
    "example_violation_narrative",
    "rule_sentence_refs",
    "deviation_sentence_refs",
    "source_section_ids",
    "validation_status",
    "dm_comments",
    "llm_updated",
    "llm_notes",
]


VALIDATION_STATUS_OPTIONS = '"accepted,to_review,rejected"'


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def build_row_key(rule_id: str, deviation_id: str) -> str:
    return f"{rule_id}::{deviation_id}"


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _read_header_map(ws: Worksheet) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    m: Dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        key = _norm_header(str(cell))
        if key:
            m[key] = idx
    return m


def _get_cell(row: Tuple[Any, ...], colmap: Dict[str, int], *names: str) -> str:
    for n in names:
        idx = colmap.get(n)
        if idx is None or idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            return ""
        return str(value).strip()
    return ""


def _csv(values: List[str]) -> str:
    return ", ".join(values)


def _normalize_status(raw: str) -> Optional[str]:
    status = (raw or "").strip().lower()
    if not status:
        return None
    if status in VALIDATION_STATUSES:
        return status
    return None


def _build_row(
    *,
    rule: Dict[str, Any],
    deviation: Dict[str, Any],
    validation_status: str,
    dm_comments: str = "",
    llm_updated: str = "",
    llm_notes: str = "",
) -> List[str]:
    return [
        build_row_key(rule.get("rule_id", ""), deviation.get("deviation_id", "")),
        rule.get("rule_id", ""),
        deviation.get("deviation_id", ""),
        rule.get("title", ""),
        rule.get("atomic_requirement", ""),
        deviation.get("scenario_description", ""),
        deviation.get("example_violation_narrative", ""),
        _csv(rule.get("sentence_refs", [])),
        _csv(deviation.get("sentence_refs", [])),
        _csv(deviation.get("source_section_ids", [])),
        validation_status,
        dm_comments,
        llm_updated,
        llm_notes,
    ]


def export_step2_review_workbook(*, step2_json_path: Path, workbook_path: Path) -> Path:
    step2_obj = read_json(step2_json_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "DM Review"
    ws.append(HEADERS)

    for rule in step2_obj.get("rules", []):
        for deviation in rule.get("candidate_deviations", []):
            ws.append(
                _build_row(
                    rule=rule,
                    deviation=deviation,
                    validation_status="",
                )
            )

    # Restrict validation_status to explicit DM choices via in-cell dropdown.
    status_col = HEADERS.index("validation_status") + 1
    status_col_letter = ws.cell(row=1, column=status_col).column_letter
    validation = DataValidation(
        type="list",
        formula1=VALIDATION_STATUS_OPTIONS,
        allow_blank=True,
    )
    validation.errorTitle = "Invalid validation_status"
    validation.error = "Allowed values: accepted, to_review, rejected."
    validation.promptTitle = "Select validation_status"
    validation.prompt = "Choose one: accepted, to_review, rejected."
    ws.add_data_validation(validation)
    validation.add(f"{status_col_letter}2:{status_col_letter}1048576")

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    return workbook_path


def read_step2_review_workbook(workbook_path: Path) -> Dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active
    colmap = _read_header_map(ws)
    required = {"row_key", "rule_id", "deviation_id", "validation_status", "dm_comments"}
    missing = sorted([h for h in required if h not in colmap])
    if missing:
        raise ValueError(f"Workbook missing required columns: {', '.join(missing)}")

    updates: Dict[str, Dict[str, str]] = {}
    warnings: List[str] = []
    errors: List[str] = []

    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        row_key = _get_cell(row, colmap, "row_key")
        rule_id = _get_cell(row, colmap, "rule_id")
        deviation_id = _get_cell(row, colmap, "deviation_id")
        if not row_key and rule_id and deviation_id:
            row_key = build_row_key(rule_id, deviation_id)
        if not row_key:
            warnings.append(f"Row {excel_row_idx}: missing row identity, skipped.")
            continue

        status_raw = _get_cell(row, colmap, "validation_status")
        status = _normalize_status(status_raw)
        if status_raw and not status:
            errors.append(
                f"Row {excel_row_idx} ({row_key}): invalid validation_status={status_raw!r}"
            )
            continue

        update = {
            "row_key": row_key,
            "rule_id": rule_id,
            "deviation_id": deviation_id,
            "validation_status": status or "",
            "dm_comments": _get_cell(row, colmap, "dm_comments"),
        }
        if row_key in updates:
            warnings.append(f"Row {excel_row_idx} ({row_key}): duplicate key, using last row.")
        updates[row_key] = update

    return {"updates": updates, "warnings": warnings, "errors": errors}


def apply_review_and_finalize(
    *,
    step2_obj: Dict[str, Any],
    review_rows: Dict[str, Any],
    revalidate_deviation: Optional[
        Callable[[Dict[str, Any], Dict[str, Any], str], List[Dict[str, Any]]]
    ],
    strict: bool = False,
) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]]]:
    updates = review_rows.get("updates", {})
    errors = list(review_rows.get("errors", []))
    warnings = list(review_rows.get("warnings", []))
    final_rows: List[Dict[str, Any]] = []
    unresolved: List[str] = []
    removed: List[str] = []
    unchanged: List[str] = []

    final_obj = deepcopy(step2_obj)
    final_obj["generated_at"] = _iso_now()

    final_rules: List[Dict[str, Any]] = []
    for rule in final_obj.get("rules", []):
        kept_deviations: List[Dict[str, Any]] = []
        for deviation in rule.get("candidate_deviations", []):
            row_key = build_row_key(rule.get("rule_id", ""), deviation.get("deviation_id", ""))
            update = updates.get(row_key, {})
            status = (update.get("validation_status") or "accepted").strip().lower()
            dm_comments = update.get("dm_comments", "")

            if status not in VALIDATION_STATUSES:
                errors.append(f"{row_key}: unsupported status {status!r}.")
                continue
            if status == "rejected":
                removed.append(row_key)
                continue
            if status == "accepted":
                unchanged.append(row_key)
                kept_deviations.append(deviation)
                final_rows.append(
                    {
                        "row": _build_row(
                            rule=rule,
                            deviation=deviation,
                            validation_status="accepted",
                            dm_comments=dm_comments,
                            llm_updated="no",
                        ),
                        "color": "green",
                    }
                )
                continue

            if not dm_comments.strip():
                warnings.append(f"{row_key}: to_review without dm_comments.")
            if revalidate_deviation is None:
                unresolved.append(f"{row_key}: no revalidation function provided.")
                kept_deviations.append(deviation)
                final_rows.append(
                    {
                        "row": _build_row(
                            rule=rule,
                            deviation=deviation,
                            validation_status="accepted",
                            dm_comments=dm_comments,
                            llm_updated="no",
                            llm_notes="unresolved_to_review",
                        ),
                        "color": "green",
                    }
                )
                continue
            try:
                updated_list = revalidate_deviation(rule, deviation, dm_comments)
            except Exception as ex:  # pragma: no cover - defensive around external API errors
                unresolved.append(f"{row_key}: LLM revalidation failed ({ex}).")
                kept_deviations.append(deviation)
                final_rows.append(
                    {
                        "row": _build_row(
                            rule=rule,
                            deviation=deviation,
                            validation_status="accepted",
                            dm_comments=dm_comments,
                            llm_updated="no",
                            llm_notes=f"llm_error: {ex}",
                        ),
                        "color": "green",
                    }
                )
                continue
            if not isinstance(updated_list, list) or not updated_list:
                unresolved.append(f"{row_key}: LLM returned no deviations.")
                kept_deviations.append(deviation)
                final_rows.append(
                    {
                        "row": _build_row(
                            rule=rule,
                            deviation=deviation,
                            validation_status="accepted",
                            dm_comments=dm_comments,
                            llm_updated="no",
                            llm_notes="llm_empty_response",
                        ),
                        "color": "green",
                    }
                )
                continue

            for idx, updated in enumerate(updated_list):
                if not isinstance(updated, dict):
                    continue
                if idx == 0:
                    new_dev_id = str(updated.get("deviation_id") or deviation.get("deviation_id", ""))
                else:
                    base = deviation.get("deviation_id", "dev")
                    new_dev_id = str(updated.get("deviation_id") or f"{base}-r{idx + 1}")
                updated["deviation_id"] = new_dev_id
                updated["source_section_ids"] = deviation.get("source_section_ids", [])
                updated["source_section_paths"] = deviation.get("source_section_paths", [])
                kept_deviations.append(updated)
                final_rows.append(
                    {
                        "row": _build_row(
                            rule=rule,
                            deviation=updated,
                            validation_status="accepted",
                            dm_comments=dm_comments,
                            llm_updated="yes",
                        ),
                        "color": "yellow",
                    }
                )

        if kept_deviations:
            rule["candidate_deviations"] = kept_deviations
            final_rules.append(rule)
    final_obj["rules"] = final_rules

    schema = load_schema("protocol_sections_step2_merged.schema.json")
    schema_errors = validate(final_obj, schema)
    if schema_errors:
        errors.extend([f"Final output schema error: {e}" for e in schema_errors[:25]])
    if strict and (errors or unresolved):
        joined = "; ".join((errors + unresolved)[:30])
        raise ValueError(f"Review apply failed in strict mode: {joined}")

    audit = {
        "generated_at": _iso_now(),
        "warnings": warnings,
        "errors": errors,
        "unresolved": unresolved,
        "counts": {
            "updated": len([r for r in final_rows if r["color"] == "yellow"]),
            "removed": len(removed),
            "unchanged": len(unchanged),
            "unresolved": len(unresolved),
            "errors": len(errors),
            "warnings": len(warnings),
        },
    }
    return final_obj, audit, final_rows


def write_final_review_workbook(
    *,
    output_workbook: Path,
    rows: List[Dict[str, Any]],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "DM Review"
    ws.append(HEADERS)
    for row_meta in rows:
        ws.append(row_meta["row"])
        row_idx = ws.max_row
        fill = GREEN_FILL if row_meta.get("color") == "green" else YELLOW_FILL
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_workbook)
    return output_workbook


def write_finalized_step2_outputs(
    *,
    final_obj: Dict[str, Any],
    audit_obj: Dict[str, Any],
    final_json_path: Path,
    audit_json_path: Path,
) -> Tuple[Path, Path]:
    write_json(final_json_path, final_obj)
    write_json(audit_json_path, audit_obj)
    return final_json_path, audit_json_path
