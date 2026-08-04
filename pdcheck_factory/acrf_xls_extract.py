"""Deterministic aCRF extraction from Rave Architect Loader spreadsheets.

This module converts uploaded `.xls/.xlsx` aCRF workbooks into the canonical
`acrf_summary_text_merged.json` structure consumed by downstream pipeline
steps (notably `pipeline_v2.step_acrf_field_dictionary`).

Unlike the PDF+TOC+LLM route, this path is deterministic and does not require
LLM calls.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple, TypedDict
from pathlib import Path

from pdcheck_factory.json_util import write_json


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_header(value: Any) -> str:
    # Normalize header labels for fuzzy matching:
    # - case-insensitive
    # - drop spaces/punctuation
    s = str(value or "").strip().upper()
    return re.sub(r"[^A-Z0-9]+", "", s)


def _cell_to_str(value: Any) -> str:
    """Convert workbook cell values to stable string IDs.

    Rave Architect Loader tends to store OIDs as strings (e.g. `VIST`) or
    sometimes as numbers; handle both with minimal formatting.
    """

    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Use a conservative formatting that avoids scientific notation.
        return str(value).rstrip("0").rstrip(".")
    return str(value).strip()


def _maybe_int(value: Any) -> Optional[int]:
    s = _cell_to_str(value)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _map_control_type_to_column_type(control_type: str) -> str:
    ct = _norm_header(control_type)

    # The output type must be one of the schema-enumerated values used by
    # `acrf_field_dictionary._infer_type()` / schema validation.
    if "DATETIME" in ct:
        return "datetime"
    if ct == "CHECKBOX":
        return "boolean"
    if "CHECKBOX" in ct:
        # Conservative: checkboxes are boolean-ish, but could also be multi.
        return "boolean"
    if "RADIOBUTTON" in ct:
        return "categorical"
    if "DROPDOWNLIST" in ct:
        return "categorical"
    if "SEARCHLIST" in ct:
        return "categorical"
    if "LONGTEXT" in ct or ct == "TEXT":
        return "text"
    if ct == "TEXT":
        return "text"

    return "unknown"


class _SummaryColumn(TypedDict, total=False):
    column_name: str
    column_description: str
    column_values: str
    column_type: str


class _SummaryDataset(TypedDict, total=False):
    dataset_name: str
    columns: List[_SummaryColumn]


class _SummaryMerged(TypedDict):
    schema_version: str
    study_id: str
    generated_at: str
    datasets: List[_SummaryDataset]


@dataclass(frozen=True)
class _HeaderMatch:
    header_row_index: int
    # Mapping from semantic column key -> column index
    columns: Dict[str, int]


def _sheet_header_rows_values(
    get_row: Callable[[int], Sequence[Any]],
    *,
    max_rows: int = 25,
    max_cols: int = 60,
) -> Iterable[Tuple[int, List[Any]]]:
    for r in range(0, max_rows):
        try:
            row = list(get_row(r)[:max_cols])
        except IndexError:
            # Some sheets are internally ragged; stop scanning headers for this sheet.
            break
        if any(str(c or "").strip() for c in row):
            yield r, row


def _find_header_row(
    *,
    get_row: Callable[[int], Sequence[Any]],
    required_keys: Sequence[Tuple[str, Sequence[str]]],
    max_rows: int = 30,
    max_cols: int = 60,
) -> Optional[_HeaderMatch]:
    """Find the row that looks like the required table header.

    `required_keys` is a list of (semantic_key, header_candidates) where each
    candidate is either a normalized header substring or an exact normalized header.
    """

    for header_row_index, row in _sheet_header_rows_values(get_row, max_rows=max_rows, max_cols=max_cols):
        norm_row = [_norm_header(c) for c in row]
        columns: Dict[str, int] = {}

        ok = True
        for semantic_key, candidates in required_keys:
            chosen = None
            for i, hv in enumerate(norm_row):
                if not hv:
                    continue
                for cand in candidates:
                    cand_n = _norm_header(cand)
                    if cand_n and (hv == cand_n or cand_n in hv):
                        chosen = i
                        break
                if chosen is not None:
                    break
            if chosen is None:
                ok = False
                break
            columns[semantic_key] = chosen

        if ok:
            return _HeaderMatch(header_row_index=header_row_index, columns=columns)

    return None


def _build_acrf_summary_from_rows(
    *,
    study_id: str,
    form_rows: Sequence[Dict[str, Any]],
    field_rows: Sequence[Dict[str, Any]],
    data_dict_entries: Dict[str, List[str]],
    unit_dict_entries: Dict[str, str],
    unit_dict_standard_name: Dict[str, str],
) -> _SummaryMerged:
    forms_by_oid: Dict[str, str] = {}
    for fr in form_rows:
        oid = _cell_to_str(fr.get("oid"))
        if not oid:
            continue
        forms_by_oid[oid] = _cell_to_str(fr.get("name"))

    # Merge columns by (dataset_oid, column_oid) to stay deterministic.
    col_by_key: Dict[Tuple[str, str], _SummaryColumn] = {}

    for f in field_rows:
        form_oid = _cell_to_str(f.get("form_oid"))
        field_oid = _cell_to_str(f.get("field_oid"))
        if not form_oid or not field_oid:
            continue

        draft_field_name = _cell_to_str(f.get("label"))
        data_format = _cell_to_str(f.get("data_format"))
        data_dict_name = _cell_to_str(f.get("data_dictionary_name"))
        unit_dict_name = _cell_to_str(f.get("unit_dictionary_name"))
        control_type = _cell_to_str(f.get("control_type"))

        col_type = _map_control_type_to_column_type(control_type)

        allowed_values = []
        if data_dict_name:
            allowed_values = list(data_dict_entries.get(data_dict_name, []) or [])

        column_values = ""
        if allowed_values:
            column_values = ", ".join(str(x) for x in allowed_values if str(x).strip())
        elif data_format:
            column_values = data_format

        unit_standard = unit_dict_standard_name.get(unit_dict_name, "") if unit_dict_name else ""
        unit_entry_string = unit_dict_entries.get(unit_dict_name, "") if unit_dict_name else ""
        unit_hint = unit_standard or unit_entry_string

        col_desc_parts = [draft_field_name] if draft_field_name else []
        if unit_hint:
            # Keep it short: deterministic hints for the downstream UI/table.
            col_desc_parts.append(f"Unit: {unit_hint}")
        column_description = " · ".join(col_desc_parts).strip() or draft_field_name or field_oid

        key = (form_oid, field_oid)
        existing = col_by_key.get(key)
        if existing is None or len(column_description) > len(existing.get("column_description", "")):
            col_by_key[key] = {
                "column_name": field_oid,
                "column_description": column_description,
                "column_values": column_values,
                "column_type": col_type,
            }

    # Expand back to dataset list with deterministic ordering.
    datasets_by_oid: Dict[str, Dict[str, Any]] = {}
    for (form_oid, field_oid), col in col_by_key.items():
        bucket = datasets_by_oid.setdefault(
            form_oid,
            {
                "dataset_name": form_oid,
                "columns": [],
            },
        )
        bucket["columns"].append(col)

    datasets_out: List[_SummaryDataset] = []
    for form_oid in sorted(datasets_by_oid.keys()):
        bucket = datasets_by_oid[form_oid]
        columns = sorted(bucket["columns"], key=lambda c: str(c.get("column_name") or ""))
        datasets_out.append(
            {
                "dataset_name": form_oid,
                "columns": columns,
            }
        )

    return {
        "schema_version": "1.0.0",
        "study_id": study_id,
        "generated_at": _iso_now(),
        "datasets": datasets_out,
    }


def _extract_from_openpyxl(
    *,
    workbook_bytes: bytes,
    study_id: str,
) -> _SummaryMerged:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    # Locate tables by header detection; don’t rely on sheet names.
    forms_match: Optional[_HeaderMatch] = None
    fields_match: Optional[_HeaderMatch] = None
    data_dict_entries_match: Optional[_HeaderMatch] = None
    unit_dict_entries_match: Optional[_HeaderMatch] = None
    unit_dicts_match: Optional[_HeaderMatch] = None

    form_rows: List[Dict[str, Any]] = []
    field_rows: List[Dict[str, Any]] = []

    data_dict_entries: Dict[str, List[str]] = {}
    unit_dict_entries: Dict[str, str] = {}
    unit_dict_standard_name: Dict[str, str] = {}

    for sheet in workbook.worksheets:
        # Build a lightweight getter for header detection.
        def get_row(r: int) -> Sequence[Any]:
            return [sheet.cell(row=r + 1, column=c + 1).value for c in range(min(sheet.max_column or 0, 60))]

        if forms_match is None:
            forms_match = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("oid", ["OID"]),
                    ("name", ["DraftFormName", "FormName", "FormNameLabel", "FORMNAME"]),
                ],
                max_rows=15,
            )

        if fields_match is None:
            fields_match = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("form_oid", ["FormOID", "FORMOID", "Form OID"]),
                    ("field_oid", ["FieldOID", "FIELDOID", "Field OID"]),
                    ("label", ["DraftFieldName", "SASLabel", "SAS Label", "DraftFieldName", "FieldName", "LABEL"]),
                    ("data_format", ["DataFormat", "DATAFORMAT", "Data Format"]),
                    ("data_dictionary_name", ["DataDictionaryName", "DATADICTIONARYNAME", "DictionaryName", "CODEDDICT"]),
                    ("unit_dictionary_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("control_type", ["ControlType", "CONTROLTYPE"]),
                ],
                max_rows=20,
            )

        if data_dict_entries_match is None:
            data_dict_entries_match = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["DataDictionaryName", "DATADICTIONARYNAME"]),
                    ("coded", ["CodedData", "CODEDDATA"]),
                    ("user", ["UserDataString", "USERDATASTRING"]),
                ],
                max_rows=20,
            )

        if unit_dicts_match is None:
            unit_dicts_match = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("standard", ["StandardUnitName", "STANDARDUNITNAME"]),
                ],
                max_rows=20,
            )

        if unit_dict_entries_match is None:
            unit_dict_entries_match = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("unit_string", ["UnitString", "UNITSTRING"]),
                ],
                max_rows=20,
            )

        # Parse forms
        if forms_match is not None:
            # Re-detect within this sheet specifically to get correct row indices/columns.
            h = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("oid", ["OID"]),
                    ("name", ["DraftFormName", "FormName", "FORMNAME"]),
                ],
                max_rows=15,
            )
            if h is not None:
                header = h.header_row_index
                cols = h.columns
                # xl values are 1-indexed for openpyxl cell; our get_row uses r+1.
                for r in range(header + 1, min((sheet.max_row or 0), header + 20000)):
                    oid = _cell_to_str(sheet.cell(row=r + 1, column=cols["oid"] + 1).value)
                    name = _cell_to_str(sheet.cell(row=r + 1, column=cols["name"] + 1).value)
                    if not oid:
                        continue
                    form_rows.append({"oid": oid, "name": name})

        # Parse fields
        if fields_match is not None:
            h = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("form_oid", ["FormOID", "FORMOID"]),
                    ("field_oid", ["FieldOID", "FIELDOID"]),
                    ("label", ["DraftFieldName", "SASLabel", "SAS Label", "FieldName"]),
                    ("data_format", ["DataFormat", "DATAFORMAT"]),
                    ("data_dictionary_name", ["DataDictionaryName", "DATADICTIONARYNAME"]),
                    ("unit_dictionary_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("control_type", ["ControlType", "CONTROLTYPE"]),
                ],
                max_rows=20,
            )
            if h is not None:
                header = h.header_row_index
                cols = h.columns
                for r in range(header + 1, min((sheet.max_row or 0), header + 500000)):
                    form_oid = _cell_to_str(sheet.cell(row=r + 1, column=cols["form_oid"] + 1).value)
                    field_oid = _cell_to_str(sheet.cell(row=r + 1, column=cols["field_oid"] + 1).value)
                    if not form_oid or not field_oid:
                        continue
                    draft_field_name = sheet.cell(row=r + 1, column=cols["label"] + 1).value
                    data_format = sheet.cell(row=r + 1, column=cols["data_format"] + 1).value
                    data_dict_name = sheet.cell(row=r + 1, column=cols["data_dictionary_name"] + 1).value
                    unit_dict_name = sheet.cell(row=r + 1, column=cols["unit_dictionary_name"] + 1).value
                    control_type = sheet.cell(row=r + 1, column=cols["control_type"] + 1).value
                    field_rows.append(
                        {
                            "form_oid": form_oid,
                            "field_oid": field_oid,
                            "label": _cell_to_str(draft_field_name),
                            "data_format": _cell_to_str(data_format),
                            "data_dictionary_name": _cell_to_str(data_dict_name),
                            "unit_dictionary_name": _cell_to_str(unit_dict_name),
                            "control_type": _cell_to_str(control_type),
                        }
                    )

        # Parse data dictionary entries
        if data_dict_entries_match is not None:
            h = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["DataDictionaryName", "DATADICTIONARYNAME"]),
                    ("coded", ["CodedData", "CODEDDATA"]),
                    ("user", ["UserDataString", "USERDATASTRING"]),
                ],
                max_rows=20,
            )
            if h is not None:
                header = h.header_row_index
                cols = h.columns
                for r in range(header + 1, min((sheet.max_row or 0), header + 500000)):
                    dname = _cell_to_str(sheet.cell(row=r + 1, column=cols["dict_name"] + 1).value)
                    coded = _cell_to_str(sheet.cell(row=r + 1, column=cols["coded"] + 1).value)
                    user = _cell_to_str(sheet.cell(row=r + 1, column=cols["user"] + 1).value)
                    if not dname:
                        continue
                    chosen = user if user and user.upper() != "SPECIFY" else coded
                    chosen = chosen.strip()
                    if not chosen:
                        continue
                    data_dict_entries.setdefault(dname, []).append(chosen)

        # Parse unit dictionaries (to map unit dictionary name -> standard unit)
        if unit_dicts_match is not None:
            h = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("standard", ["StandardUnitName", "STANDARDUNITNAME"]),
                ],
                max_rows=20,
            )
            if h is not None:
                header = h.header_row_index
                cols = h.columns
                for r in range(header + 1, min((sheet.max_row or 0), header + 500000)):
                    dname = _cell_to_str(sheet.cell(row=r + 1, column=cols["dict_name"] + 1).value)
                    standard = _cell_to_str(sheet.cell(row=r + 1, column=cols["standard"] + 1).value)
                    if not dname or not standard:
                        continue
                    unit_dict_standard_name[dname] = standard

        # Parse unit dictionary entries (fallback unit string per dictionary)
        if unit_dict_entries_match is not None:
            h = _find_header_row(
                get_row=get_row,
                required_keys=[
                    ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                    ("unit_string", ["UnitString", "UNITSTRING"]),
                ],
                max_rows=20,
            )
            if h is not None:
                header = h.header_row_index
                cols = h.columns
                for r in range(header + 1, min((sheet.max_row or 0), header + 500000)):
                    dname = _cell_to_str(sheet.cell(row=r + 1, column=cols["dict_name"] + 1).value)
                    unit_string = _cell_to_str(sheet.cell(row=r + 1, column=cols["unit_string"] + 1).value)
                    if not dname or not unit_string:
                        continue
                    unit_dict_entries[dname] = unit_string

    return _build_acrf_summary_from_rows(
        study_id=study_id,
        form_rows=form_rows,
        field_rows=field_rows,
        data_dict_entries=data_dict_entries,
        unit_dict_entries=unit_dict_entries,
        unit_dict_standard_name=unit_dict_standard_name,
    )


def _extract_from_xlrd(
    *,
    workbook_bytes: bytes,
    study_id: str,
) -> _SummaryMerged:
    try:
        import xlrd  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError(
            "xlrd is required to parse `.xls` aCRF workbooks. Add `xlrd` to project dependencies."
        ) from exc

    workbook = xlrd.open_workbook(file_contents=workbook_bytes)

    form_rows: List[Dict[str, Any]] = []
    field_rows: List[Dict[str, Any]] = []
    data_dict_entries: Dict[str, List[str]] = {}
    unit_dict_entries: Dict[str, str] = {}
    unit_dict_standard_name: Dict[str, str] = {}

    # Scan sheets and parse the relevant tables by header detection.
    for sheet in workbook.sheets():
        def get_row(r: int) -> Sequence[Any]:
            # Use row_values to avoid IndexError on ragged rows.
            end_col = min(getattr(sheet, "ncols", 0) or 0, 60)
            return list(sheet.row_values(r, start_colx=0, end_colx=end_col))

        forms_match = _find_header_row(
            get_row=get_row,
            required_keys=[
                ("oid", ["OID"]),
                ("name", ["DraftFormName", "FormName", "FORMNAME"]),
            ],
            max_rows=15,
        )
        if forms_match is not None:
            cols = forms_match.columns
            header = forms_match.header_row_index
            for r in range(header + 1, min(sheet.nrows, header + 500000)):
                oid = _cell_to_str(sheet.cell_value(r, cols["oid"]))
                name = _cell_to_str(sheet.cell_value(r, cols["name"]))
                if not oid:
                    continue
                form_rows.append({"oid": oid, "name": name})

        fields_match = _find_header_row(
            get_row=get_row,
            required_keys=[
                ("form_oid", ["FormOID", "FORMOID"]),
                ("field_oid", ["FieldOID", "FIELDOID"]),
                ("label", ["DraftFieldName", "SASLabel", "SAS Label", "FieldName", "LABEL"]),
                ("data_format", ["DataFormat", "DATAFORMAT"]),
                ("data_dictionary_name", ["DataDictionaryName", "DATADICTIONARYNAME"]),
                ("unit_dictionary_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                ("control_type", ["ControlType", "CONTROLTYPE"]),
            ],
            max_rows=20,
        )
        if fields_match is not None:
            cols = fields_match.columns
            header = fields_match.header_row_index
            for r in range(header + 1, min(sheet.nrows, header + 500000)):
                form_oid = _cell_to_str(sheet.cell_value(r, cols["form_oid"]))
                field_oid = _cell_to_str(sheet.cell_value(r, cols["field_oid"]))
                if not form_oid or not field_oid:
                    continue
                draft_field_name = sheet.cell_value(r, cols["label"])
                data_format = sheet.cell_value(r, cols["data_format"])
                data_dict_name = sheet.cell_value(r, cols["data_dictionary_name"])
                unit_dict_name = sheet.cell_value(r, cols["unit_dictionary_name"])
                control_type = sheet.cell_value(r, cols["control_type"])
                field_rows.append(
                    {
                        "form_oid": form_oid,
                        "field_oid": field_oid,
                        "label": _cell_to_str(draft_field_name),
                        "data_format": _cell_to_str(data_format),
                        "data_dictionary_name": _cell_to_str(data_dict_name),
                        "unit_dictionary_name": _cell_to_str(unit_dict_name),
                        "control_type": _cell_to_str(control_type),
                    }
                )

        data_entries_match = _find_header_row(
            get_row=get_row,
            required_keys=[
                ("dict_name", ["DataDictionaryName", "DATADICTIONARYNAME"]),
                ("coded", ["CodedData", "CODEDDATA"]),
                ("user", ["UserDataString", "USERDATASTRING"]),
            ],
            max_rows=20,
        )
        if data_entries_match is not None:
            cols = data_entries_match.columns
            header = data_entries_match.header_row_index
            for r in range(header + 1, min(sheet.nrows, header + 500000)):
                dname = _cell_to_str(sheet.cell_value(r, cols["dict_name"]))
                coded = _cell_to_str(sheet.cell_value(r, cols["coded"]))
                user = _cell_to_str(sheet.cell_value(r, cols["user"]))
                if not dname:
                    continue
                chosen = user if user and user.upper() != "SPECIFY" else coded
                chosen = chosen.strip()
                if not chosen:
                    continue
                data_dict_entries.setdefault(dname, []).append(chosen)

        unit_dicts_match = _find_header_row(
            get_row=get_row,
            required_keys=[
                ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                ("standard", ["StandardUnitName", "STANDARDUNITNAME"]),
            ],
            max_rows=20,
        )
        if unit_dicts_match is not None:
            cols = unit_dicts_match.columns
            header = unit_dicts_match.header_row_index
            for r in range(header + 1, min(sheet.nrows, header + 500000)):
                dname = _cell_to_str(sheet.cell_value(r, cols["dict_name"]))
                standard = _cell_to_str(sheet.cell_value(r, cols["standard"]))
                if not dname or not standard:
                    continue
                unit_dict_standard_name[dname] = standard

        unit_entries_match = _find_header_row(
            get_row=get_row,
            required_keys=[
                ("dict_name", ["UnitDictionaryName", "UNITDICTIONARYNAME"]),
                ("unit_string", ["UnitString", "UNITSTRING"]),
            ],
            max_rows=20,
        )
        if unit_entries_match is not None:
            cols = unit_entries_match.columns
            header = unit_entries_match.header_row_index
            for r in range(header + 1, min(sheet.nrows, header + 500000)):
                dname = _cell_to_str(sheet.cell_value(r, cols["dict_name"]))
                unit_string = _cell_to_str(sheet.cell_value(r, cols["unit_string"]))
                if not dname or not unit_string:
                    continue
                unit_dict_entries[dname] = unit_string

    return _build_acrf_summary_from_rows(
        study_id=study_id,
        form_rows=form_rows,
        field_rows=field_rows,
        data_dict_entries=data_dict_entries,
        unit_dict_entries=unit_dict_entries,
        unit_dict_standard_name=unit_dict_standard_name,
    )


def build_acrf_summary_text_merged_from_workbook_bytes(
    *,
    workbook_bytes: bytes,
    file_format: str,
    study_id: str,
) -> _SummaryMerged:
    """Build deterministic `acrf_summary_text_merged.json` from an uploaded workbook."""

    fmt = str(file_format or "").strip().lower()
    if not workbook_bytes:
        raise ValueError("workbook_bytes is required")
    if fmt in {"xlsx"}:
        return _extract_from_openpyxl(workbook_bytes=workbook_bytes, study_id=study_id)
    if fmt in {"xls"}:
        return _extract_from_xlrd(workbook_bytes=workbook_bytes, study_id=study_id)
    raise ValueError(f"Unsupported workbook file_format: {file_format!r}")


def write_acrf_summary_text_merged_json(
    *,
    out_path: Path,
    workbook_bytes: bytes,
    file_format: str,
    study_id: str,
) -> _SummaryMerged:
    """Helper that writes output JSON and returns it."""

    out_obj = build_acrf_summary_text_merged_from_workbook_bytes(
        workbook_bytes=workbook_bytes,
        file_format=file_format,
        study_id=study_id,
    )
    # Preserve existing project JSON formatting by using write_json (UTF-8, ensure_ascii=False).
    write_json(out_path, out_obj)
    return out_obj

