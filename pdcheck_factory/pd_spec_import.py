"""Parse company PD Specifications workbooks into deviation review rows."""

from __future__ import annotations

import hashlib
import re
from io import BytesIO
from typing import Any, Dict, List, Mapping, Tuple

from openpyxl import load_workbook

from pdcheck_factory.deviation_contract import split_pd_spec_row
from pdcheck_factory.pd_spec_export import PD_SPEC_SHEET_TITLE


def _normalize_header(value: str) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    return re.sub(r"\s+", " ", text)

_PD_FIELDS = [
    "protocol_deviation_category",
    "protocol_deviation_sub_category",
    "text",
    "occurrence_date",
    "classification",
    "manual_or_programmable",
    "additional_information",
    "programming_status",
    "data_source",
    "pseudo_logic_seed",
    "programmer_comments",
    "reviewer_comments",
    "programmer_check_number",
]

_LEGACY_AA_FIELD = "aa_comment"

_HEADER_ALIASES = {
    _normalize_header(h): field
    for h, field in zip(
        [
            "Protocol Deviation Category",
            "Protocol Deviation Sub-Category",
            "Protocol Deviation Description\n250 Character Limit",
            "Protocol Deviation Occurrence Date",
            "Protocol Deviation Classification",
            "Manual or Programmable Deviation",
            "Additional Information / Comments",
            "Programming Status",
            "Data Source (e.g., RAVE, Clario, LabConnect)\n30 Character Limit",
            "Programming Information",
            "Programmer Comments",
            "Reviewer Comments",
            "Programmer Check Number",
        ],
        _PD_FIELDS,
    )
}
_HEADER_ALIASES[_normalize_header("AA comment")] = _LEGACY_AA_FIELD


def stable_deviation_id(category: str, sub_category: str, description: str) -> str:
    """Stable id from PD spec identity fields (row index excluded)."""
    normalized = "|".join(part.strip().lower() for part in (category, sub_category, description))
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]
    return f"dev-import-{digest}"


def stable_pd_spec_rule_id(category: str, sub_category: str) -> str:
    """Stable rule id per PD spec category + sub-category (protocol deviation sub-category)."""
    normalized = "|".join(part.strip().lower() for part in (category, sub_category))
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]
    return f"pd-spec-rule-{digest}"


def synthetic_rule_id(deviation_id: str) -> str:
    """Deprecated: per-deviation rule id; kept for tests referencing old behavior."""
    return f"pd-spec-{deviation_id}"


def programmable_from_manual_or_programmable(value: str) -> bool | None:
    text = str(value or "").strip().lower()
    if text == "manual":
        return False
    if text == "programmable":
        return True
    return None


def build_pd_spec_deviation_text(row_values: Mapping[str, str]) -> str:
    """Compose display text from PD spec columns (description + supplemental fields)."""
    parts: List[str] = []
    description = str(row_values.get("text", "") or "").strip()
    if description:
        parts.append(description)
    for label, key in (
        ("Additional information", "additional_information"),
        ("Data source", "data_source"),
        ("Programming information", "pseudo_logic_seed"),
        ("Programmer comments", "programmer_comments"),
        ("Reviewer comments", "reviewer_comments"),
    ):
        value = str(row_values.get(key, "") or "").strip()
        if value:
            parts.append(f"{label}: {value}")
    return "\n\n".join(parts)


def _build_header_map(headers: Tuple[Any, ...]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for index, value in enumerate(headers):
        key = _normalize_header(str(value or ""))
        if key:
            header_map[key] = index
    return header_map


def _row_values_from_pd_spec(
    row: Tuple[Any, ...],
    header_map: Mapping[str, int],
) -> Dict[str, str]:
    values: Dict[str, str] = {field: "" for field in _PD_FIELDS}
    values[_LEGACY_AA_FIELD] = ""
    for header_key, col_index in header_map.items():
        field = _HEADER_ALIASES.get(header_key)
        if not field or col_index >= len(row):
            continue
        cell_val = row[col_index]
        values[field] = str(cell_val).strip() if cell_val is not None else ""
    if values.get("text"):
        return values
    description_idx = header_map.get(_normalize_header("Protocol Deviation Description 250 Character Limit"))
    if description_idx is not None and description_idx < len(row):
        cell_val = row[description_idx]
        values["text"] = str(cell_val).strip() if cell_val is not None else ""
    return values


def map_pd_spec_row_to_deviation(
    row_values: Mapping[str, str],
    *,
    row_index: int,
) -> Dict[str, Any]:
    """Map one PD Specifications row to a deviations_review_state deviation."""
    category = str(row_values.get("protocol_deviation_category", "") or "").strip()
    sub_category = str(row_values.get("protocol_deviation_sub_category", "") or "").strip()
    text = str(row_values.get("text", "") or "").strip()
    if not category or not sub_category or not text:
        raise ValueError(
            f"Row {row_index}: category, sub-category, and description are required"
        )

    deviation_id = stable_deviation_id(category, sub_category, text)
    rule_id = stable_pd_spec_rule_id(category, sub_category)
    composite_text = build_pd_spec_deviation_text(row_values)
    data_support = str(row_values.get("data_source", "") or "").strip()
    manual_prog = str(row_values.get("manual_or_programmable", "") or "").strip()

    row: Dict[str, Any] = {
        "deviation_id": deviation_id,
        "rule_id": rule_id,
        "text": composite_text,
        "paragraph_refs": [],
        "data_support_note": data_support,
        "status": "pending",
        "dm_comment": "",
        "entry_source": "imported_pd_spec",
        "protocol_deviation_category": category,
        "protocol_deviation_sub_category": sub_category,
        "classification": str(row_values.get("classification", "") or "").strip(),
        "data_source": data_support,
        "manual_or_programmable": manual_prog,
        "programming_status": str(row_values.get("programming_status", "") or "").strip(),
        "programmer_comments": str(row_values.get("programmer_comments", "") or "").strip(),
        "reviewer_comments": str(row_values.get("reviewer_comments", "") or "").strip(),
        "aa_comment": str(row_values.get(_LEGACY_AA_FIELD, "") or "").strip(),
        "programmer_check_number": str(row_values.get("programmer_check_number", "") or "").strip(),
        "occurrence_date": str(row_values.get("occurrence_date", "") or "").strip(),
        "additional_information": str(row_values.get("additional_information", "") or "").strip(),
        "pseudo_logic_seed": str(row_values.get("pseudo_logic_seed", "") or "").strip(),
        "grounding_error": "",
    }
    return split_pd_spec_row(row)


def _unique_column_headers(raw_headers: Tuple[Any, ...]) -> List[str]:
    """Preserve workbook header labels; disambiguate duplicates."""
    labels: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        label = str(value or "").strip() if value is not None else ""
        if not label:
            label = f"Column {index + 1}"
        count = seen.get(label, 0) + 1
        seen[label] = count
        labels.append(label if count == 1 else f"{label} ({count})")
    return labels


def parse_pd_spec_xlsx_table(workbook_bytes: bytes) -> Dict[str, Any]:
    """Parse workbook into header labels and row dicts keyed by those labels."""
    if not workbook_bytes:
        raise ValueError("Workbook must not be empty")
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    for name in workbook.sheetnames:
        if name.strip().lower() == PD_SPEC_SHEET_TITLE.lower():
            sheet = workbook[name]
            break

    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("Workbook must include a header row")

    column_headers = _unique_column_headers(tuple(headers))
    table_rows: List[Dict[str, str]] = []
    for row_values in rows_iter:
        if not row_values or not any(v is not None and str(v).strip() for v in row_values):
            continue
        cells = tuple(row_values)
        table_rows.append(
            {
                column_headers[index]: (
                    str(cells[index]).strip() if index < len(cells) and cells[index] is not None else ""
                )
                for index in range(len(column_headers))
            }
        )

    if not table_rows:
        raise ValueError("Workbook did not contain any PD specification data rows")
    return {"headers": column_headers, "rows": table_rows}


def parse_pd_spec_xlsx(workbook_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse active sheet of a PD Specifications workbook into deviation dicts."""
    if not workbook_bytes:
        raise ValueError("Workbook must not be empty")
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    for name in workbook.sheetnames:
        if name.strip().lower() == PD_SPEC_SHEET_TITLE.lower():
            sheet = workbook[name]
            break

    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("Workbook must include a header row")

    header_map = _build_header_map(tuple(headers))
    deviations: List[Dict[str, Any]] = []
    row_index = 1
    for row_values in rows_iter:
        row_index += 1
        if not row_values or not any(v is not None and str(v).strip() for v in row_values):
            continue
        parsed = _row_values_from_pd_spec(tuple(row_values), header_map)
        try:
            deviations.append(map_pd_spec_row_to_deviation(parsed, row_index=row_index))
        except ValueError:
            continue

    if not deviations:
        raise ValueError("Workbook did not contain any valid PD specification rows")
    return deviations
