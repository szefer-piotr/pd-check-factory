#!/usr/bin/env python3
"""Extract PD category/sub-category taxonomy and template metadata from NAL00-107."""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = REPO_ROOT / "docs/refernce-files/NAL00-107 PD Specifications.xlsx"
TAXONOMY_OUT = REPO_ROOT / "pdcheck_factory/data/pd_category_subcategory.json"
META_OUT = REPO_ROOT / "pdcheck_factory/data/pd_spec_template_meta.json"

PROGRAMMING_STATUS_OPTIONS = [
    "Specd for CTL Review",
    "Not Applicable",
    "Question - Pending",
    "Ready for Programming",
    "Programmed",
    "Programmed - Ready for Review",
    "Review Failed",
    "Completed",
]

MANUAL_OR_PROGRAMMABLE_OPTIONS = ["Manual", "Programmable"]

EXPORT_HEADERS = [
    "Protocol Deviation Category",
    "Protocol Deviation Sub-Category",
    "Protocol Deviation Description\n250 Character Limit",
    "Protocol Deviation Occurrence Date",
    "Protocol Deviation Classification",
    "Manual or Programmable Deviation",
    "Additional Information / Comments",
    "Programming Status",
    "Data Source (e.g., RAVE, Clario, LabConnect)\n30 Character Limit",
    "Programming Information",
    "Programmer Comments",
    "Reviewer Comments",
    "Programmer Check Number",
]

COLUMN_WIDTHS = {
    1: 28,
    2: 28,
    3: 48,
    4: 22,
    5: 24,
    6: 26,
    7: 36,
    8: 22,
    9: 32,
    10: 48,
    11: 28,
    12: 28,
    13: 24,
}


def _extract_categories_from_dictionaries(ws) -> dict[str, list[str]]:
    categories: dict[str, list[str]] = {}
    for col in range(3, 13):
        cat = ws.cell(1, col).value
        if not cat:
            continue
        subs: list[str] = []
        seen: set[str] = set()
        for row in range(2, 50):
            value = ws.cell(row, col).value
            if not value:
                continue
            sub = str(value).strip()
            if sub in seen:
                continue
            seen.add(sub)
            subs.append(sub)
        categories[str(cat).strip()] = subs
    return categories


def extract_taxonomy(template_path: Path) -> dict:
    wb = load_workbook(template_path, data_only=True)
    categories = _extract_categories_from_dictionaries(wb["Dictionaries"])
    wb.close()
    return {
        "source": str(template_path.relative_to(REPO_ROOT)),
        "schema_version": "1.0.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "categories": categories,
    }


def extract_template_meta(template_path: Path, categories: dict[str, list[str]]) -> dict:
    wb = load_workbook(template_path, data_only=True)
    ws = wb["PD Specifications"]
    source_headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    wb.close()
    all_sub_categories: list[str] = []
    for subs in categories.values():
        for sub in subs:
            if sub not in all_sub_categories:
                all_sub_categories.append(sub)
    return {
        "source": str(template_path.relative_to(REPO_ROOT)),
        "schema_version": "1.0.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sheet_title": "PD Specifications",
        "source_headers": [str(h) if h is not None else "" for h in source_headers],
        "export_headers": EXPORT_HEADERS,
        "programming_status_options": PROGRAMMING_STATUS_OPTIONS,
        "manual_or_programmable_options": MANUAL_OR_PROGRAMMABLE_OPTIONS,
        "category_options": list(categories.keys()),
        "sub_category_options": all_sub_categories,
        "column_widths": COLUMN_WIDTHS,
        "freeze_panes": "A2",
    }


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    template_path = Path(args[0]) if args else DEFAULT_TEMPLATE
    if not template_path.is_file():
        print(f"Template not found: {template_path}", file=sys.stderr)
        return 1

    taxonomy = extract_taxonomy(template_path)
    meta = extract_template_meta(template_path, taxonomy["categories"])

    TAXONOMY_OUT.parent.mkdir(parents=True, exist_ok=True)
    TAXONOMY_OUT.write_text(json.dumps(taxonomy, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    META_OUT.write_text(json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    sub_count = sum(len(v) for v in taxonomy["categories"].values())
    print(f"Wrote {TAXONOMY_OUT} ({len(taxonomy['categories'])} categories, {sub_count} sub-categories)")
    print(f"Wrote {META_OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
