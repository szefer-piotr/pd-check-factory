from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pdcheck_factory.coding_workbook_export import (
    map_step7_row_to_pd_spec_row,
    write_coding_workbook_xlsx,
)
from pdcheck_factory.pd_spec_export import (
    DICTIONARIES_SHEET_TITLE,
    PD_SPEC_HEADERS,
    PD_SPEC_SHEET_TITLE,
)


class CodingWorkbookExportTests(unittest.TestCase):
    def test_map_step7_row_to_pd_spec_row(self) -> None:
        row = {
            "deviation_id": "dev-0001",
            "rule_id": "rule-001",
            "rule_title": "Visit window timing",
            "rule_text": "Visit must occur in window.",
            "deviation_text": "Visit outside allowed window.",
            "paragraph_refs": ["p1", "p2"],
            "paragraph_refs_text": "p1, p2",
            "supporting_sentences": [{"ref": "p1", "text": "Window is 7 days."}],
            "data_support_note": "Needs DM dataset.",
            "status": "accepted",
            "dm_comment": "Reviewed",
            "entry_source": "extracted",
            "programmable": True,
            "programmability_note": "Straightforward check",
            "pseudo_logic": "IF visit_date NOT IN window THEN flag",
        }
        mapped = map_step7_row_to_pd_spec_row(row, study_id="MY-STUDY", exported_at="2026-05-21T10:00:00Z")
        self.assertEqual(mapped[0], "Visit window timing")
        self.assertEqual(mapped[2], "Visit outside allowed window.")
        self.assertEqual(mapped[3], "2026-05-21")
        self.assertEqual(mapped[4], "accepted")
        self.assertEqual(mapped[5], "Programmable")
        self.assertIn("rule_id: rule-001", mapped[6])
        self.assertIn("rule_text: Visit must occur in window.", mapped[6])
        self.assertIn("supporting_sentences:", mapped[6])
        self.assertEqual(mapped[8], "Needs DM dataset.")
        self.assertEqual(mapped[9], "IF visit_date NOT IN window THEN flag")
        self.assertEqual(mapped[10], "Straightforward check")
        self.assertEqual(mapped[11], "Reviewed")
        self.assertEqual(mapped[12], "")

    def test_manual_or_programmable_manual(self) -> None:
        row = {
            "deviation_id": "dev-0002",
            "rule_id": "rule-002",
            "rule_title": "Dosing",
            "deviation_text": "Dose mismatch.",
            "status": "pending",
            "programmable": False,
            "entry_source": "imported",
        }
        mapped = map_step7_row_to_pd_spec_row(row, study_id="STUDY", exported_at="2026-05-21T12:00:00Z")
        self.assertEqual(mapped[5], "Manual")

    def test_description_truncated_to_250(self) -> None:
        row = {
            "rule_id": "rule-003",
            "deviation_id": "dev-0003",
            "deviation_text": "x" * 300,
        }
        mapped = map_step7_row_to_pd_spec_row(row, study_id="STUDY", exported_at="2026-05-21T12:00:00Z")
        self.assertEqual(len(mapped[2]), 250)
        self.assertTrue(mapped[2].endswith("..."))

    def test_write_coding_workbook_xlsx_structure(self) -> None:
        rows = [
            {
                "deviation_id": "dev-0001",
                "rule_id": "rule-001",
                "rule_title": "Timing",
                "rule_text": "Rule body",
                "deviation_text": "Deviation one.",
                "paragraph_refs": ["p1"],
                "paragraph_refs_text": "p1",
                "supporting_sentences": [],
                "data_support_note": "",
                "status": "accepted",
                "dm_comment": "",
                "entry_source": "extracted",
                "programmable": True,
                "programmability_note": "",
                "pseudo_logic": "pseudo one",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "coding.xlsx"
            write_coding_workbook_xlsx(
                rows,
                out_path,
                study_id="TEST-STUDY",
                exported_at="2026-05-21T10:00:00Z",
            )
            self.assertTrue(out_path.is_file())

            wb = load_workbook(out_path)
            self.assertEqual(wb.sheetnames[0], PD_SPEC_SHEET_TITLE)
            self.assertIn(DICTIONARIES_SHEET_TITLE, wb.sheetnames)

            ws = wb[PD_SPEC_SHEET_TITLE]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers, PD_SPEC_HEADERS)
            self.assertEqual(ws["A2"].value, "Timing")
            self.assertEqual(ws["C2"].value, "Deviation one.")
            self.assertIn("pseudo one", ws["J2"].value)
            self.assertIn("rule_id: rule-001", ws["G2"].value)
            self.assertGreater(len(ws.data_validations.dataValidation), 0)


if __name__ == "__main__":
    unittest.main()
