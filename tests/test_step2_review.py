from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pdcheck_factory.json_util import write_json
from pdcheck_factory.step2_review import (
    apply_review_and_finalize,
    export_step2_review_workbook,
    read_step2_review_workbook,
    write_final_review_workbook,
)


def _step2_fixture() -> dict:
    return {
        "schema_version": "2.1.1",
        "study_id": "study-x",
        "generated_at": "2026-01-01T00:00:00+00:00",
        "rules": [
            {
                "rule_id": "rule-001",
                "title": "Age requirement",
                "atomic_requirement": "Patient age must be at least 18.",
                "sentence_refs": ["sec:a#s1"],
                "source_section_ids": ["sec:a"],
                "source_section_paths": [["Section A"]],
                "candidate_deviations": [
                    {
                        "deviation_id": "dev-001a",
                        "scenario_description": "Subject is 16 years old.",
                        "example_violation_narrative": "Enrollment was performed before age 18.",
                        "sentence_refs": ["sec:a#s1"],
                        "programmable": True,
                        "pseudo_sql_logic": "SELECT subject_id FROM dm WHERE age < 18",
                        "source_section_ids": ["sec:a"],
                        "source_section_paths": [["Section A"]],
                    },
                    {
                        "deviation_id": "dev-001b",
                        "scenario_description": "Subject age unavailable.",
                        "example_violation_narrative": "Site could not verify age eligibility.",
                        "sentence_refs": ["sec:a#s2"],
                        "programmable": False,
                        "pseudo_sql_logic": "SELECT subject_id FROM dm WHERE age IS NULL",
                        "source_section_ids": ["sec:a"],
                        "source_section_paths": [["Section A"]],
                    },
                ],
            }
        ],
    }


class Step2ReviewTests(unittest.TestCase):
    def test_roundtrip_read_and_finalize(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            step2_path = root / "step2_merged.json"
            workbook_path = root / "dm_review.xlsx"
            reviewed_workbook_path = root / "dm_review.reviewed.xlsx"

            write_json(step2_path, _step2_fixture())
            export_step2_review_workbook(
                step2_json_path=step2_path,
                workbook_path=workbook_path,
            )

            wb = load_workbook(workbook_path)
            ws = wb.active
            # Row 2 -> to_review, Row 3 -> rejected.
            ws["M2"] = "to_review"
            ws["N2"] = "Keep same intent but rewrite using DM comment."
            ws["M3"] = "rejected"
            wb.save(workbook_path)

            parsed = read_step2_review_workbook(workbook_path)
            self.assertFalse(parsed["errors"])

            def _revalidate(rule: dict, deviation: dict, dm_comments: str) -> list[dict]:
                return [
                    {
                        "deviation_id": deviation["deviation_id"],
                        "scenario_description": f"{deviation['scenario_description']} (updated)",
                        "example_violation_narrative": dm_comments,
                        "sentence_refs": list(deviation["sentence_refs"]),
                        "programmable": True,
                        "pseudo_sql_logic": "SELECT subject_id FROM dm WHERE age < 18",
                        "source_section_ids": list(deviation["source_section_ids"]),
                        "source_section_paths": list(deviation["source_section_paths"]),
                    }
                ]

            final_obj, audit, final_rows = apply_review_and_finalize(
                step2_obj=_step2_fixture(),
                review_rows=parsed,
                revalidate_deviation=_revalidate,
                strict=True,
            )
            self.assertEqual(audit["counts"]["updated"], 1)
            self.assertEqual(audit["counts"]["removed"], 1)
            self.assertEqual(len(final_obj["rules"]), 1)
            self.assertEqual(len(final_obj["rules"][0]["candidate_deviations"]), 1)
            self.assertEqual(len(final_rows), 1)
            self.assertEqual(final_rows[0]["color"], "yellow")

            write_final_review_workbook(
                output_workbook=reviewed_workbook_path,
                rows=final_rows,
            )
            reviewed = load_workbook(reviewed_workbook_path)
            rws = reviewed.active
            self.assertEqual(rws["O2"].value, "yes")
            self.assertNotEqual(rws["A2"].fill.fill_type, None)
            self.assertEqual(rws.max_row, 2)

    def test_invalid_status_is_reported(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            step2_path = root / "step2_merged.json"
            workbook_path = root / "dm_review.xlsx"

            write_json(step2_path, _step2_fixture())
            export_step2_review_workbook(
                step2_json_path=step2_path,
                workbook_path=workbook_path,
            )
            wb = load_workbook(workbook_path)
            ws = wb.active
            ws["M2"] = "maybe"
            wb.save(workbook_path)

            parsed = read_step2_review_workbook(workbook_path)
            self.assertTrue(parsed["errors"])

    def test_export_adds_validation_status_dropdown(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            step2_path = root / "step2_merged.json"
            workbook_path = root / "dm_review.xlsx"
            write_json(step2_path, _step2_fixture())
            export_step2_review_workbook(
                step2_json_path=step2_path,
                workbook_path=workbook_path,
            )
            wb = load_workbook(workbook_path)
            ws = wb.active
            self.assertTrue(ws.data_validations.dataValidation)
            validation = ws.data_validations.dataValidation[0]
            self.assertEqual(validation.type, "list")
            self.assertEqual(validation.formula1, '"accepted,to_review,rejected"')

    def test_strict_mode_fails_on_unresolved_to_review(self) -> None:
        review_rows = {
            "updates": {
                "rule-001::dev-001a": {
                    "validation_status": "to_review",
                    "dm_comments": "Please reassess.",
                }
            },
            "warnings": [],
            "errors": [],
        }
        with self.assertRaises(ValueError):
            apply_review_and_finalize(
                step2_obj=_step2_fixture(),
                review_rows=review_rows,
                revalidate_deviation=None,
                strict=True,
            )

    def test_to_review_can_split_into_multiple_atomic_deviations(self) -> None:
        review_rows = {
            "updates": {
                "rule-001::dev-001a": {
                    "validation_status": "to_review",
                    "dm_comments": "Split procedure bundle into atomic deviations.",
                },
                "rule-001::dev-001b": {"validation_status": "rejected", "dm_comments": ""},
            },
            "warnings": [],
            "errors": [],
        }

        def _splitter(rule: dict, deviation: dict, dm_comments: str) -> list[dict]:
            return [
                {
                    "deviation_id": deviation["deviation_id"],
                    "scenario_description": "Procedure A missing.",
                    "example_violation_narrative": dm_comments,
                    "sentence_refs": list(deviation["sentence_refs"]),
                    "programmable": True,
                    "pseudo_sql_logic": "SELECT s FROM proc WHERE proc_a IS NULL",
                    "source_section_ids": list(deviation["source_section_ids"]),
                    "source_section_paths": list(deviation["source_section_paths"]),
                },
                {
                    "deviation_id": deviation["deviation_id"] + "-r2",
                    "scenario_description": "Procedure B missing.",
                    "example_violation_narrative": dm_comments,
                    "sentence_refs": list(deviation["sentence_refs"]),
                    "programmable": True,
                    "pseudo_sql_logic": "SELECT s FROM proc WHERE proc_b IS NULL",
                    "source_section_ids": list(deviation["source_section_ids"]),
                    "source_section_paths": list(deviation["source_section_paths"]),
                },
            ]

        final_obj, audit, final_rows = apply_review_and_finalize(
            step2_obj=_step2_fixture(),
            review_rows=review_rows,
            revalidate_deviation=_splitter,
            strict=True,
        )
        self.assertEqual(audit["counts"]["updated"], 2)
        self.assertEqual(audit["counts"]["removed"], 1)
        self.assertEqual(len(final_obj["rules"][0]["candidate_deviations"]), 2)
        self.assertEqual(len(final_rows), 2)
        self.assertTrue(all(r["color"] == "yellow" for r in final_rows))


if __name__ == "__main__":
    unittest.main()
