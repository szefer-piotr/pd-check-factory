"""Tests for PD Specifications workbook import parsing."""

from __future__ import annotations

import unittest

from pdcheck_factory.pd_spec_export import PD_SPEC_HEADERS, PD_SPEC_SHEET_TITLE, write_final_pd_spec_xlsx
from pdcheck_factory.pd_spec_import import (
    build_pd_spec_deviation_text,
    map_pd_spec_row_to_deviation,
    parse_pd_spec_xlsx,
    parse_pd_spec_xlsx_table,
    programmable_from_manual_or_programmable,
    stable_deviation_id,
    stable_pd_spec_rule_id,
    synthetic_rule_id,
)
from openpyxl import Workbook
from io import BytesIO


class TestPdSpecImport(unittest.TestCase):
    def test_stable_deviation_id_is_deterministic(self) -> None:
        a = stable_deviation_id("Eligibility", "Age", "Subject age below minimum")
        b = stable_deviation_id("Eligibility", "Age", "Subject age below minimum")
        self.assertEqual(a, b)
        self.assertTrue(a.startswith("dev-import-"))

    def test_stable_pd_spec_rule_id_groups_by_sub_category(self) -> None:
        rule_a = stable_pd_spec_rule_id("Eligibility Criteria", "Age")
        rule_b = stable_pd_spec_rule_id("Eligibility Criteria", "Age")
        rule_other = stable_pd_spec_rule_id("Eligibility Criteria", "Consent")
        self.assertEqual(rule_a, rule_b)
        self.assertNotEqual(rule_a, rule_other)
        self.assertTrue(rule_a.startswith("pd-spec-rule-"))

    def test_synthetic_rule_id(self) -> None:
        dev_id = "dev-import-abc123"
        self.assertEqual(synthetic_rule_id(dev_id), "pd-spec-dev-import-abc123")

    def test_programmable_from_manual_or_programmable(self) -> None:
        self.assertFalse(programmable_from_manual_or_programmable("Manual"))
        self.assertTrue(programmable_from_manual_or_programmable("Programmable"))
        self.assertIsNone(programmable_from_manual_or_programmable(""))

    def test_build_pd_spec_deviation_text_combines_fields(self) -> None:
        text = build_pd_spec_deviation_text(
            {
                "text": "Subject below minimum age",
                "additional_information": "extra note",
                "data_source": "RAVE",
                "pseudo_logic_seed": "seed logic",
                "programmer_comments": "prog",
                "reviewer_comments": "rev",
            }
        )
        self.assertIn("Subject below minimum age", text)
        self.assertIn("Data source: RAVE", text)
        self.assertIn("Reviewer comments: rev", text)

    def test_map_pd_spec_row_requires_core_fields(self) -> None:
        with self.assertRaises(ValueError):
            map_pd_spec_row_to_deviation(
                {"protocol_deviation_category": "", "protocol_deviation_sub_category": "x", "text": "y"},
                row_index=2,
            )

    def test_map_pd_spec_row_uses_sub_category_rule_and_composite_text(self) -> None:
        row = map_pd_spec_row_to_deviation(
            {
                "protocol_deviation_category": "Eligibility Criteria",
                "protocol_deviation_sub_category": "Age",
                "text": "Subject below minimum age",
                "manual_or_programmable": "Manual",
                "data_source": "RAVE",
                "additional_information": "note",
            },
            row_index=2,
        )
        self.assertEqual(row["rule_id"], stable_pd_spec_rule_id("Eligibility Criteria", "Age"))
        self.assertIn("Subject below minimum age", row["text"])
        self.assertIn("Data source: RAVE", row["text"])
        self.assertEqual(row["entry_source"], "imported_pd_spec")

    def test_parse_pd_spec_xlsx_table_returns_all_columns(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = PD_SPEC_SHEET_TITLE
        ws.append(PD_SPEC_HEADERS)
        ws.append(
            [
                "Eligibility Criteria",
                "Age",
                "Subject below minimum age",
                "",
                "Major",
                "Programmable",
                "extra note",
                "Ready for Programming",
                "RAVE",
                "pseudo seed",
                "prog comment",
                "review comment",
                "aa note",
            ]
        )
        buffer = BytesIO()
        wb.save(buffer)
        table = parse_pd_spec_xlsx_table(buffer.getvalue())
        assert table["headers"] == PD_SPEC_HEADERS
        assert len(table["rows"]) == 1
        row = table["rows"][0]
        assert row[PD_SPEC_HEADERS[6]] == "extra note"
        assert row[PD_SPEC_HEADERS[12]] == "aa note"

    def test_parse_pd_spec_xlsx_table_includes_extra_workbook_columns(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = PD_SPEC_SHEET_TITLE
        headers = list(PD_SPEC_HEADERS) + ["Custom PD Field"]
        ws.append(headers)
        ws.append(
            [
                "Eligibility Criteria",
                "Age",
                "Subject below minimum age",
                "",
                "Major",
                "Programmable",
                "",
                "",
                "RAVE",
                "",
                "",
                "",
                "",
                "custom value",
            ]
        )
        buffer = BytesIO()
        wb.save(buffer)
        table = parse_pd_spec_xlsx_table(buffer.getvalue())
        assert "Custom PD Field" in table["headers"]
        assert table["rows"][0]["Custom PD Field"] == "custom value"

    def test_parse_pd_spec_xlsx_round_trip_headers(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = PD_SPEC_SHEET_TITLE
        ws.append(PD_SPEC_HEADERS)
        ws.append(
            [
                "Eligibility Criteria",
                "Age",
                "Subject enrolled below minimum age",
                "",
                "Major",
                "Programmable",
                "",
                "",
                "RAVE",
                "",
                "",
                "",
                "",
            ]
        )
        buffer = BytesIO()
        wb.save(buffer)
        deviations = parse_pd_spec_xlsx(buffer.getvalue())
        self.assertEqual(len(deviations), 1)
        row = deviations[0]
        self.assertEqual(row["protocol_deviation_category"], "Eligibility Criteria")
        self.assertEqual(row["protocol_deviation_sub_category"], "Age")
        self.assertIn("Subject enrolled below minimum age", row["text"])
        self.assertEqual(row["entry_source"], "imported_pd_spec")
        self.assertEqual(row["data_source"], "RAVE")
        self.assertTrue(row["deviation_id"].startswith("dev-import-"))
        self.assertEqual(row["rule_id"], stable_pd_spec_rule_id("Eligibility Criteria", "Age"))

    def test_write_final_pd_spec_still_works(self) -> None:
        from pathlib import Path
        import tempfile

        final_obj = {
            "schema_version": "1.0.0",
            "study_id": "T",
            "generated_at": "2026-05-21T00:00:00Z",
            "items": [
                {
                    "rule_id": "r1",
                    "deviation_id": "d1",
                    "deviation_text": "text",
                    "paragraph_refs": ["p1"],
                    "pseudo_logic": "logic",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            write_final_pd_spec_xlsx(final_obj, out)
            self.assertTrue(out.is_file())


if __name__ == "__main__":
    unittest.main()
