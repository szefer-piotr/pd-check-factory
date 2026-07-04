from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pdcheck_factory.pd_spec_export import (
    PD_SPEC_HEADERS,
    PD_SPEC_SHEET_TITLE,
    map_deviation_to_pd_spec_row,
    write_pd_spec_xlsx,
)


class PdSpecExportTests(unittest.TestCase):
    def test_map_deviation_to_pd_spec_row(self) -> None:
        item = {
            "protocol_deviation_category": "Study Visit Related",
            "protocol_deviation_sub_category": "Study Visit Out of Window",
            "deviation_text": "Week 4 visit outside window.",
            "pseudo_logic": "SV: visit_date NOT BETWEEN anchor-2 AND anchor+2",
            "data_source": "Rave",
            "programmable": True,
        }
        row = map_deviation_to_pd_spec_row(item)
        self.assertEqual(row[0], "Study Visit Related")
        self.assertEqual(row[1], "Study Visit Out of Window")
        self.assertEqual(row[2], "Week 4 visit outside window.")
        self.assertEqual(row[3], "")
        self.assertEqual(row[6], "")
        self.assertEqual(row[8], "Rave")
        self.assertEqual(row[9], "SV: visit_date NOT BETWEEN anchor-2 AND anchor+2")
        self.assertEqual(row[10], "")
        self.assertEqual(row[11], "")
        self.assertEqual(row[12], "")
        self.assertNotIn("AA comment", PD_SPEC_HEADERS)
        self.assertIn("Programmer Check Number", PD_SPEC_HEADERS)

    def test_write_pd_spec_xlsx_single_sheet(self) -> None:
        rows = [
            {
                "protocol_deviation_category": "Eligibility Criteria",
                "protocol_deviation_sub_category": "Inclusion Criteria not met",
                "deviation_text": "Participant enrolled but failed inclusion criterion.",
                "pseudo_logic": "SUBJ: inclusion_flag = 0",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "final_deviations.xlsx"
            write_pd_spec_xlsx(rows, out_path)
            wb = load_workbook(out_path)
            self.assertEqual(wb.sheetnames, [PD_SPEC_SHEET_TITLE])
            ws = wb[PD_SPEC_SHEET_TITLE]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers, PD_SPEC_HEADERS)
            self.assertEqual(ws["C2"].value, "Participant enrolled but failed inclusion criterion.")
            self.assertGreater(len(ws.data_validations.dataValidation), 0)


if __name__ == "__main__":
    unittest.main()
