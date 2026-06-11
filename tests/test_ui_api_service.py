import json
from pathlib import Path
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from pdcheck_factory import blob_io, extraction_resolve, paths
from pdcheck_factory.json_util import read_json, write_json
from pdcheck_factory.ui_api.service import STEP_ORDER, UiApiError, UiStepService, parse_json_body


def _touch(path: Path, content: str = "x") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def test_parse_json_body_rejects_non_object() -> None:
    with pytest.raises(UiApiError) as exc:
        parse_json_body(b"[]")
    assert exc.value.code == "BAD_JSON"


def test_run_step_reports_llm_progress_for_extract_deviations(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from pdcheck_factory import pipeline_v2

    def fake_deviations(
        sid: str,
        output_dir: Path,
        *,
        additional_instructions: str = "",
        progress_callback=None,
    ) -> dict:
        assert progress_callback is not None
        progress_callback(
            phase="extract-deviations",
            current=2,
            total=5,
            unit="rules",
            label="rule-002",
        )
        out_path = paths.local_deviations_parsed_json(sid, output_dir)
        _touch(out_path, '{"deviations": []}')
        return {"deviations": []}

    monkeypatch.setattr(pipeline_v2, "step4_5_extract_deviations", fake_deviations)
    monkeypatch.setattr(pipeline_v2, "initialize_review_states", lambda *args, **kwargs: None)

    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"

    proto = extraction_resolve.resolve_protocol_rendered_source_md(study_id, tmp_path)
    acrf = extraction_resolve.resolve_acrf_rendered_source_md(study_id, tmp_path)
    _touch(proto)
    _touch(acrf)
    pindex = paths.local_protocol_paragraph_index_json(study_id, tmp_path)
    rules = paths.local_rules_parsed_json(study_id, tmp_path)
    acrf_summary = paths.local_acrf_summary_text_merged(study_id, tmp_path)
    _touch(pindex, '{"paragraphs": []}')
    _touch(rules, '{"rules": [{"rule_id": "rule-001"}]}')
    _touch(acrf_summary, "summary")

    service.run_step(study_id, "extract-deviations")

    run_state = service.get_step1_run_state(study_id)
    log_texts = [line["text"] for line in run_state["logs"]]
    assert "llm:extract-deviations:2/5:rule-002" in log_texts
    assert run_state["llmProgress"] is None


def test_run_step_forwards_llm_instructions_to_extract_rules(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from pdcheck_factory import pipeline_v2

    captured: dict[str, str] = {}

    def fake_rules(sid: str, output_dir: Path, *, additional_instructions: str = "") -> dict:
        captured["additional_instructions"] = additional_instructions
        out_path = paths.local_rules_parsed_json(sid, output_dir)
        _touch(out_path, '{"rules": []}')
        return {"rules": []}

    monkeypatch.setattr(pipeline_v2, "step3_extract_rules", fake_rules)

    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"

    proto = extraction_resolve.resolve_protocol_rendered_source_md(study_id, tmp_path)
    acrf = extraction_resolve.resolve_acrf_rendered_source_md(study_id, tmp_path)
    _touch(proto)
    _touch(acrf)
    pindex = paths.local_protocol_paragraph_index_json(study_id, tmp_path)
    _touch(pindex, '{"paragraphs": []}')

    service.run_step(study_id, "extract-rules", llm_instructions="  Focus oncology  ")
    assert captured["additional_instructions"] == "Focus oncology"


def test_status_progression_and_dependency_guard(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"

    status = service.get_status(study_id)
    assert status["steps"][0]["status"] == "pending"

    protocol = tmp_path / study_id / "extractions" / "protocol" / "opendataloader" / "rendered" / "source.md"
    acrf = tmp_path / study_id / "extractions" / "acrf" / "layout" / "rendered" / "source.md"
    _touch(protocol)
    _touch(acrf)

    status = service.get_status(study_id)
    assert {row["stepId"]: row["status"] for row in status["steps"]}["extract-inputs"] == "done"

    called = {"index": False, "split": False, "acrf": False, "rules": False, "dev": False, "init": False, "pseudo": False, "final": False}

    def fake_index(sid: str, output_dir: Path):
        called["index"] = True
        out = output_dir / sid / "pipeline" / "protocol_index" / "paragraph_index.json"
        _touch(out, '{"paragraphs": []}')
        return {"paragraphs": []}

    def fake_rules(sid: str, output_dir: Path, *args, **kwargs):
        called["rules"] = True
        out = output_dir / sid / "pipeline" / "rules" / "rules_parsed.json"
        _touch(out, '{"rules": []}')
        return {"rules": []}

    def fake_acrf_summary(sid: str, output_dir: Path, *args, **kwargs):
        called["acrf"] = True
        out = output_dir / sid / "pipeline" / "acrf_summary" / "acrf_summary_text_merged.json"
        _touch(out, '{"datasets": []}')
        return {"datasets": []}

    def fake_split_toc(source_md: Path, destination_dir: Path, write_manifest: bool):
        called["split"] = True
        _touch(destination_dir / "001_demo.md", "# demo")
        manifest = destination_dir / "sections_manifest.json"
        if write_manifest:
            _touch(manifest, '{"sections": []}')
        return 1, manifest

    def fake_dev(sid: str, output_dir: Path, *args, **kwargs):
        called["dev"] = True
        out = output_dir / sid / "pipeline" / "deviations" / "deviations_parsed.json"
        _touch(out, '{"deviations": []}')
        return {"deviations": []}

    def fake_init(sid: str, output_dir: Path):
        called["init"] = True
        review = output_dir / sid / "pipeline" / "review" / "deviations_review_state.json"
        _touch(review, '{"deviations": []}')

    def fake_pseudo(sid: str, output_dir: Path):
        called["pseudo"] = True
        out = output_dir / sid / "pipeline" / "pseudo_logic" / "pseudo_logic_validated.json"
        _touch(out, '{"items": []}')
        return {"items": []}

    def fake_final(sid: str, output_dir: Path):
        called["final"] = True
        final_json = output_dir / sid / "pipeline" / "final" / "final_deviations.json"
        final_xlsx = output_dir / sid / "pipeline" / "final" / "final_deviations.xlsx"
        _touch(final_json, '{"items": []}')
        _touch(final_xlsx, "xlsx")
        return {"items": []}

    from pdcheck_factory import pipeline_v2
    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(pipeline_v2, "step2_protocol_paragraph_index", fake_index)
    monkeypatch.setattr(pipeline_v2, "step1_acrf_summary_text", fake_acrf_summary)
    monkeypatch.setattr(pipeline_v2, "step3_extract_rules", fake_rules)
    monkeypatch.setattr(pipeline_v2, "step4_5_extract_deviations", fake_dev)
    monkeypatch.setattr(pipeline_v2, "initialize_review_states", fake_init)
    monkeypatch.setattr(pipeline_v2, "step8_generate_pseudo_logic", fake_pseudo)
    monkeypatch.setattr(pipeline_v2, "step10_finalize", fake_final)
    monkeypatch.setattr(cli_mod, "run_acrf_split_toc", fake_split_toc)

    with pytest.raises(UiApiError) as blocked:
        service.run_step(study_id, "extract-rules")
    assert blocked.value.code == "STEP_BLOCKED"

    extract_pipeline_steps = [
        "index-protocol",
        "acrf-split-toc",
        "acrf-summary-text",
        "extract-rules",
        "extract-deviations",
        "review-and-finalize",
    ]
    for step_id in extract_pipeline_steps:
        service.run_step(study_id, step_id)

    assert all(called.values())
    final_status = {row["stepId"]: row["status"] for row in service.get_status(study_id)["steps"]}
    assert final_status["review-and-finalize"] == "done"


def test_list_studies_discovers_raw_blob_pairs(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(
        blob_io,
        "list_blob_names_with_prefix",
        lambda **_kwargs: [
            "raw/STUDY-A/protocol.pdf",
            "raw/STUDY-A/acrf.pdf",
            "raw/STUDY-B/protocol.pdf",
            "raw/STUDY-C/acrf.pdf",
        ],
    )

    payload = service.list_studies()

    assert [study["studyId"] for study in payload["studies"]] == ["STUDY-A", "STUDY-B", "STUDY-C"]
    study_a = next(study for study in payload["studies"] if study["studyId"] == "STUDY-A")
    assert study_a["protocolBlob"] == "raw/STUDY-A/protocol.pdf"
    assert study_a["bothUploaded"] is True
    assert study_a["stepStatuses"]["extract-inputs"] == "pending"
    study_b = next(study for study in payload["studies"] if study["studyId"] == "STUDY-B")
    assert study_b["bothUploaded"] is False


def test_delete_study_removes_all_blob_prefixes_and_local_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "STUDY-DEL"

    def fake_list(**kwargs: object) -> list[str]:
        prefix = kwargs["prefix"]
        if prefix == f"raw/{study_id}/":
            return [
                f"raw/{study_id}/protocol.pdf",
                f"raw/{study_id}/reference/protocol_v2.pdf",
            ]
        if prefix == f"extractions/{study_id}/":
            return [f"extractions/{study_id}/protocol/opendataloader/rendered/source.md"]
        if prefix == f"pipeline/{study_id}/":
            return [f"pipeline/{study_id}/ui_upload_manifest.json"]
        return []

    deleted_paths: list[str] = []

    def fake_delete(**kwargs: object) -> int:
        paths_arg = kwargs["blob_paths"]
        deleted_paths.extend(paths_arg)
        return len(paths_arg)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "list_blob_names_with_prefix", fake_list)
    monkeypatch.setattr(blob_io, "delete_blobs", fake_delete)

    local_root = paths.local_study_root(study_id, tmp_path)
    _touch(local_root / "marker.txt", "local")

    result = service.delete_study(study_id)

    assert result["deletedBlobCount"] == 4
    assert result["totalBlobCount"] == 4
    assert result["localOutputRemoved"] is True
    assert not local_root.exists()
    assert sorted(deleted_paths) == sorted(
        [
            f"raw/{study_id}/protocol.pdf",
            f"raw/{study_id}/reference/protocol_v2.pdf",
            f"extractions/{study_id}/protocol/opendataloader/rendered/source.md",
            f"pipeline/{study_id}/ui_upload_manifest.json",
        ]
    )
    assert f"raw/{study_id}/" in result["blobPrefixes"]


def test_delete_study_rejects_unsafe_study_id(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    with pytest.raises(UiApiError) as exc:
        service.delete_study("../evil")
    assert exc.value.code == "VALIDATION_ERROR"


def test_step7_deviations_chat_and_refine(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    rule_path = tmp_path / study_id / "pipeline" / "rules" / "rules_parsed.json"
    review_path = tmp_path / study_id / "pipeline" / "review" / "deviations_review_state.json"
    pseudo_path = tmp_path / study_id / "pipeline" / "review" / "pseudo_logic_review_state.json"
    _touch(
        rule_path,
        '{"rules":[{"rule_id":"rule-001","title":"Visit window timing"}]}',
    )
    _touch(
        tmp_path / study_id / "pipeline" / "protocol_index" / "paragraph_index.json",
        '{"paragraphs":[{"paragraph_id":"p1","text":"Visit must be inside the allowed window."}]}',
    )
    _touch(
        review_path,
        (
            '{"schema_version":"1.0.0","study_id":"MY-STUDY","deviations":['
            '{"deviation_id":"dev-0001","rule_id":"rule-001","text":"Original","paragraph_refs":["p1"],'
            '"data_support_note":"Supported by SV date","status":"to_review","dm_comment":""}]}'
        ),
    )
    _touch(
        pseudo_path,
        (
            '{"schema_version":"1.0.0","study_id":"MY-STUDY","items":['
            '{"deviation_id":"dev-0001","rule_id":"rule-001","pseudo_logic":"SELECT 1",'
            '"programmable":true,"programmability_note":"ok"}]}'
        ),
    )

    from pdcheck_factory import pipeline_v2

    def fake_refine(
        *,
        study_id: str,
        output_dir: Path,
        row: dict,
        dm_comment: str,
        run_revision_cycle: bool,
        chat_history=None,
        also_generate_pseudo: bool = False,
    ):
        updated = dict(row)
        updated["text"] = f"{row.get('text')} :: refined"
        updated["dm_comment"] = dm_comment
        return updated, {
            "study_id": study_id,
            "review_type": "deviations",
            "deviation_id": row.get("deviation_id"),
            "updated_rows": 1,
            "revised_rows": 1,
            "run_revision_cycle": run_revision_cycle,
            "assistant_message": "Updated deviation based on your note.",
            "response_type": "revision",
            "missing_caveats": [],
        }

    monkeypatch.setattr(pipeline_v2, "refine_single_deviation_with_comment", fake_refine)

    list_payload = service.get_step7_deviations(study_id)
    assert list_payload["columns"] == ["rule_id", "deviation_id", "rule_title", "deviation_text", "paragraph_refs", "pseudo_logic"]
    assert list_payload["rows"][0]["deviation_id"] == "dev-0001"
    assert list_payload["rows"][0]["rule_title"] == "Visit window timing"
    assert list_payload["rows"][0]["data_support_note"] == "Supported by SV date"
    assert list_payload["rows"][0]["supporting_sentences"][0]["text"] == "Visit must be inside the allowed window."

    chat_payload = service.get_step7_deviation_chat(study_id, "dev-0001")
    assert chat_payload["messages"] == []

    refined = service.refine_step7_deviation(
        study_id=study_id,
        deviation_id="dev-0001",
        dm_comment="please refine",
        run_revision_cycle=True,
    )
    assert "refined" in refined["row"]["deviation_text"]
    assert len(refined["messages"]) == 2
    assert refined["messages"][0]["role"] == "dm"
    assert refined["messages"][1]["role"] == "assistant"
    assert "Updated deviation" in refined["messages"][1]["text"]
    assert refined.get("responseType") == "revision"

    updated = service.update_step7_deviation(
        study_id=study_id,
        deviation_id="dev-0001",
        status="accepted",
        dm_comment="approved",
    )
    assert updated["row"]["status"] == "accepted"
    assert updated["row"]["dm_comment"] == "approved"


def test_step7_manual_deviation_crud_and_xlsx_import(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="pending")

    added = service.create_step7_deviation(
        study_id,
        {
            "deviation_id": "dev-manual",
            "rule_id": "rule-001",
            "text": "Manual deviation",
            "paragraph_refs": ["p1"],
            "data_support_note": "Manual support",
        },
    )
    assert any(row["deviation_id"] == "dev-manual" for row in added["rows"])

    updated = service.patch_step7_deviation_fields(
        study_id,
        "dev-manual",
        {"text": "Manual deviation edited", "status": "accepted"},
    )
    assert updated["row"]["deviation_text"] == "Manual deviation edited"
    assert updated["row"]["status"] == "accepted"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["deviation_id", "rule_id", "deviation_text", "paragraph_refs", "data_support_note"])
    sheet.append(["dev-imported", "rule-001", "Imported deviation", "p1", "Imported support"])
    buffer = BytesIO()
    workbook.save(buffer)

    imported = service.import_step7_deviations_xlsx(study_id, buffer.getvalue())
    assert imported["imported"] == 1
    assert any(row["deviation_id"] == "dev-imported" for row in imported["rows"])

    with pytest.raises(UiApiError) as duplicate:
        service.import_step7_deviations_xlsx(study_id, buffer.getvalue())
    assert duplicate.value.code == "VALIDATION_ERROR"

    deleted = service.delete_step7_deviation(study_id, "dev-manual")
    assert all(row["deviation_id"] != "dev-manual" for row in deleted["rows"])

    state = read_json(paths.local_deviations_review_state(study_id, tmp_path))
    assert any(row.get("entry_source") == "imported" for row in state["deviations"])


def test_export_step7_deviations_xlsx_writes_workbook(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="accepted")

    pseudo_path = paths.local_pseudo_logic_review_state(study_id, tmp_path)
    pseudo_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        pseudo_path,
        {
            "items": [
                {
                    "deviation_id": "dev-0001",
                    "pseudo_logic": "SELECT 1",
                    "programmable": True,
                    "programmability_note": "ok",
                }
            ]
        },
    )

    exported = service.export_step7_deviations_xlsx(study_id)
    assert exported["rowCount"] == 1
    assert exported["fileName"].endswith(".xlsx")

    out_path = paths.local_deviations_review_export_xlsx(study_id, tmp_path)
    assert out_path.is_file()

    workbook = load_workbook(out_path, read_only=True, data_only=True)
    deviations = workbook["Deviations"]
    headers = [cell.value for cell in next(deviations.iter_rows(min_row=1, max_row=1))]
    assert "deviation_id" in headers
    assert "pseudo_logic" in headers
    assert "programmable" in headers
    assert "programmability_note" in headers
    assert "supporting_sentences" in headers

    data_row = next(deviations.iter_rows(min_row=2, max_row=2, values_only=True))
    row_map = dict(zip(headers, data_row))
    assert row_map["deviation_id"] == "dev-0001"
    assert row_map["pseudo_logic"] == "SELECT 1"
    assert row_map["programmable"] == "true"

    summary = workbook["Summary"]
    summary_rows = list(summary.iter_rows(min_row=2, values_only=True))
    summary_map = {row[0]: row[1] for row in summary_rows if row[0]}
    assert summary_map["total_deviations"] == 1
    assert summary_map["accepted"] == 1


def test_export_step7_deviations_coding_xlsx_writes_workbook(tmp_path: Path) -> None:
    from pdcheck_factory.pd_spec_export import PD_SPEC_HEADERS, PD_SPEC_SHEET_TITLE

    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="accepted")

    pseudo_path = paths.local_pseudo_logic_review_state(study_id, tmp_path)
    pseudo_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        pseudo_path,
        {
            "items": [
                {
                    "deviation_id": "dev-0001",
                    "pseudo_logic": "SELECT 1",
                    "programmable": True,
                    "programmability_note": "ok",
                }
            ]
        },
    )

    exported = service.export_step7_deviations_coding_xlsx(study_id)
    assert exported["rowCount"] == 1
    assert "company_pds" in exported["fileName"]

    out_path = paths.local_deviations_coding_export_xlsx(study_id, tmp_path)
    assert out_path.is_file()

    workbook = load_workbook(out_path, read_only=True, data_only=True)
    assert workbook.sheetnames[0] == PD_SPEC_SHEET_TITLE
    sheet = workbook[PD_SPEC_SHEET_TITLE]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == PD_SPEC_HEADERS

    data_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    row_map = dict(zip(headers, data_row))
    assert row_map["Protocol Deviation Category"] == "Visit window timing"
    assert row_map["Protocol Deviation Classification"] == "accepted"
    assert row_map["Manual or Programmable Deviation"] == "Programmable"
    assert "SELECT 1" in str(row_map["Programming Information"])
    assert "deviation_id: dev-0001" in str(row_map["Additional Information / Comments"])
    assert row_map["Programmer Comments"] == "ok"


def test_step7_manual_rule_crud(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="pending")

    created = service.create_step7_rule(
        study_id,
        {"rule_id": "rule-manual", "title": "Manual rule", "text": "Rule body"},
    )
    assert created["rule"]["rule_id"] == "rule-manual"

    updated = service.update_step7_rule(study_id, "rule-manual", {"title": "Manual rule edited"})
    assert updated["rule"]["title"] == "Manual rule edited"

    deleted = service.delete_step7_rule(study_id, "rule-manual")
    assert deleted["deletedRuleId"] == "rule-manual"


def _seed_step7_state(tmp_path: Path, study_id: str, status: str = "accepted") -> None:
    rule_path = tmp_path / study_id / "pipeline" / "rules" / "rules_parsed.json"
    review_path = tmp_path / study_id / "pipeline" / "review" / "deviations_review_state.json"
    validated_path = tmp_path / study_id / "pipeline" / "deviations" / "deviations_validated.json"
    _touch(
        rule_path,
        '{"rules":[{"rule_id":"rule-001","title":"Visit window timing"}]}',
    )
    state_json = (
        '{"schema_version":"1.0.0","study_id":"' + study_id + '","deviations":['
        '{"deviation_id":"dev-0001","rule_id":"rule-001","text":"Original","paragraph_refs":["p1"],'
        '"status":"' + status + '","dm_comment":""}]}'
    )
    _touch(review_path, state_json)
    _touch(validated_path, state_json)


def test_generate_step7_pseudo_logic_for_deviation_writes_state(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="accepted")

    from pdcheck_factory import paths, pipeline_v2

    def fake_single(*, study_id: str, output_dir: Path, deviation: dict, rule_by_id=None):
        return {
            "deviation_id": deviation["deviation_id"],
            "rule_id": deviation["rule_id"],
            "rule_title": "Visit window timing",
            "pseudo_logic": "SELECT * FROM dm",
            "programmable": True,
            "programmability_note": "ok",
            "status": "pending",
            "dm_comment": "",
        }

    monkeypatch.setattr(pipeline_v2, "generate_pseudo_logic_for_deviation", fake_single)

    payload = service.generate_step7_pseudo_logic_for_deviation(study_id, "dev-0001")
    assert payload["row"]["pseudo_logic"] == "SELECT * FROM dm"
    assert payload["row"]["programmable"] is True
    assert payload["row"]["programmability_note"] == "ok"

    review_state_path = paths.local_pseudo_logic_review_state(study_id, tmp_path)
    validated_path = paths.local_pseudo_logic_validated_json(study_id, tmp_path)
    assert review_state_path.is_file()
    assert validated_path.is_file()
    review_obj = read_json(review_state_path)
    assert any(item.get("deviation_id") == "dev-0001" for item in review_obj.get("items", []))


def test_generate_step7_pseudo_logic_for_deviation_rejects_non_accepted(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="pending")

    from pdcheck_factory import pipeline_v2

    def fake_single(**_kwargs):
        raise AssertionError("should not be called")

    monkeypatch.setattr(pipeline_v2, "generate_pseudo_logic_for_deviation", fake_single)

    with pytest.raises(UiApiError) as blocked:
        service.generate_step7_pseudo_logic_for_deviation(study_id, "dev-0001")
    assert blocked.value.code == "STEP_BLOCKED"
    assert blocked.value.status_code == 409


def test_accept_step7_deviations_bulk_accepts_pending_and_to_review(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="to_review")

    payload = service.accept_step7_deviations_bulk(study_id)
    assert payload["accepted"] == 1
    assert payload["rows"][0]["status"] == "accepted"

    state = read_json(paths.local_deviations_review_state(study_id, tmp_path))
    assert state["deviations"][0]["status"] == "accepted"


def test_accept_step7_deviations_bulk_skips_rejected(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="rejected")

    payload = service.accept_step7_deviations_bulk(study_id)
    assert payload["accepted"] == 0
    assert payload["rows"][0]["status"] == "rejected"


def test_generate_step7_pseudo_logic_bulk_returns_rows_and_count(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "MY-STUDY"
    _seed_step7_state(tmp_path, study_id, status="accepted")

    from pdcheck_factory import paths, pipeline_v2

    def fake_bulk(sid: str, output_dir: Path):
        out = {
            "schema_version": "1.0.0",
            "study_id": sid,
            "generated_at": "2024-01-01T00:00:00Z",
            "items": [
                {
                    "deviation_id": "dev-0001",
                    "rule_id": "rule-001",
                    "rule_title": "Visit window timing",
                    "pseudo_logic": "SELECT 1",
                    "programmable": True,
                    "programmability_note": "ok",
                    "status": "pending",
                    "dm_comment": "",
                }
            ],
        }
        review_state_path = paths.local_pseudo_logic_review_state(sid, output_dir)
        validated_path = paths.local_pseudo_logic_validated_json(sid, output_dir)
        write_json(review_state_path, out)
        write_json(validated_path, out)
        return out

    monkeypatch.setattr(pipeline_v2, "step8_generate_pseudo_logic", fake_bulk)

    payload = service.generate_step7_pseudo_logic_bulk(study_id)
    assert payload["generated"] == 1
    assert payload["rows"][0]["deviation_id"] == "dev-0001"
    assert payload["rows"][0]["pseudo_logic"] == "SELECT 1"
    assert payload["rows"][0]["rule_title"] == "Visit window timing"


def test_run_step1_extract_opendataloader_flags_and_choice(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "EX-S"
    captured: dict = {}

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    def fake_run_extract(**kwargs: object) -> None:
        captured.update(kwargs)

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)
    out = service.run_step1_extract(study_id, extractor="opendataloader")
    assert captured.get("opendataloader_only") is True
    assert captured.get("run_opendataloader_ocr") is True
    assert out["extractor"] == "opendataloader"
    choice_path = extraction_resolve.local_ui_extractor_choice_json(study_id, tmp_path)
    assert choice_path.is_file()
    assert read_json(choice_path)["extractor"] == "opendataloader"


def test_run_step1_extract_document_intelligence_flags(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "EX-S"
    captured: dict = {}

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    def fake_run_extract(**kwargs: object) -> None:
        captured.update(kwargs)

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)
    out = service.run_step1_extract(study_id, extractor="document_intelligence")
    assert captured.get("opendataloader_only") is False
    assert captured.get("run_opendataloader_ocr") is False
    assert out["extractor"] == "document_intelligence"
    choice_path = extraction_resolve.local_ui_extractor_choice_json(study_id, tmp_path)
    assert read_json(choice_path)["extractor"] == "document_intelligence"


def test_run_step1_extract_default_both(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "EX-S"
    captured: dict = {}

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    def fake_run_extract(**kwargs: object) -> None:
        captured.update(kwargs)

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)
    out = service.run_step1_extract(study_id, extractor=None)
    assert captured.get("run_opendataloader_ocr") is True
    assert captured.get("opendataloader_only") is False
    assert out["extractor"] == "both"
    assert read_json(extraction_resolve.local_ui_extractor_choice_json(study_id, tmp_path))["extractor"] == "both"


def test_upload_step1_files_persists_original_filenames(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "UP-S"

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    out = service.upload_step1_files(
        study_id,
        b"protocol-bytes",
        b"acrf-bytes",
        protocol_file_name="Protocol_v3_final.pdf",
        acrf_file_name="aCRF_annotated.pdf",
    )

    assert out["protocolFileName"] == "Protocol_v3_final.pdf"
    assert out["acrfFileName"] == "aCRF_annotated.pdf"
    manifest = read_json(paths.local_ui_upload_manifest(study_id, tmp_path))
    assert manifest["protocolFileName"] == "Protocol_v3_final.pdf"
    assert manifest["acrfFileName"] == "aCRF_annotated.pdf"

    preview = service.get_step1_preview(study_id)
    assert preview["protocolFileName"] == "Protocol_v3_final.pdf"
    assert preview["acrfFileName"] == "aCRF_annotated.pdf"


def test_get_step1_preview_filename_fallback_without_manifest(tmp_path: Path) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "FB-S"
    preview = service.get_step1_preview(study_id)
    assert preview["protocolFileName"] == "protocol.pdf"
    assert preview["acrfFileName"] == "acrf.pdf"


def test_get_specifications_preview_maps_workbook_and_review_state(tmp_path: Path) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    from pdcheck_factory.pd_spec_export import PD_SPEC_HEADERS, PD_SPEC_SHEET_TITLE

    service = UiStepService(output_dir=tmp_path)
    study_id = "SPEC-PREV"

    workbook_path = paths.local_pd_spec_workbook(study_id, tmp_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = PD_SPEC_SHEET_TITLE
    ws.append(PD_SPEC_HEADERS)
    ws.append(
        [
            "Eligibility Criteria",
            "Age",
            "Subject enrolled below minimum age",
            "",
            "Major",
            "Programmable",
            "",
            "",
            "RAVE",
            "",
            "",
            "",
            "",
        ]
    )
    buffer = BytesIO()
    wb.save(buffer)
    workbook_path.write_bytes(buffer.getvalue())

    review_path = paths.local_deviations_review_state(study_id, tmp_path)
    review_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        review_path,
        {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "deviations": [
                {
                    "deviation_id": "dev-0001",
                    "rule_id": "rule-001",
                    "text": "Generated deviation text",
                    "paragraph_refs": [],
                    "status": "pending",
                    "entry_source": "extracted",
                }
            ],
        },
    )

    preview = service.get_specifications_preview(study_id)
    keys = {source["key"] for source in preview["sources"]}
    assert "pd_spec_workbook" in keys
    assert "review_state" in keys

    workbook_source = next(s for s in preview["sources"] if s["key"] == "pd_spec_workbook")
    assert workbook_source["columns"] == PD_SPEC_HEADERS
    assert len(workbook_source["rows"]) == 1
    row = workbook_source["rows"][0]
    assert row[PD_SPEC_HEADERS[0]] == "Eligibility Criteria"
    assert row[PD_SPEC_HEADERS[2]] == "Subject enrolled below minimum age"
    assert row[PD_SPEC_HEADERS[9]] == "RAVE"

    review_source = next(s for s in preview["sources"] if s["key"] == "review_state")
    assert review_source["rows"][0]["deviation_text"] == "Generated deviation text"


def test_get_specifications_preview_loads_pd_spec_from_blob(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    from pdcheck_factory.pd_spec_export import PD_SPEC_HEADERS, PD_SPEC_SHEET_TITLE

    service = UiStepService(output_dir=tmp_path)
    study_id = "SPEC-BLOB"
    local_path = paths.local_pd_spec_workbook(study_id, tmp_path)
    assert not local_path.is_file()

    wb = Workbook()
    ws = wb.active
    ws.title = PD_SPEC_SHEET_TITLE
    ws.append(PD_SPEC_HEADERS)
    ws.append(
        [
            "Eligibility Criteria",
            "Age",
            "Blob-only row",
            "",
            "Major",
            "Programmable",
            "",
            "",
            "RAVE",
            "",
            "",
            "",
            "",
        ]
    )
    buffer = BytesIO()
    wb.save(buffer)
    blob_bytes = buffer.getvalue()

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "download_blob_bytes", lambda **_kwargs: blob_bytes)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    preview = service.get_specifications_preview(study_id)
    workbook_source = next(s for s in preview["sources"] if s["key"] == "pd_spec_workbook")
    assert workbook_source["rows"][0][PD_SPEC_HEADERS[2]] == "Blob-only row"
    assert local_path.is_file()


def test_get_step1_upload_status_includes_pd_spec_from_blob(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "PD-STATUS"
    manifest_path = paths.local_ui_upload_manifest(study_id, tmp_path)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        manifest_path,
        {
            "study_id": study_id,
            "pdSpecFileName": "company_specs.xlsx",
            "pdSpecSize": 2048,
        },
    )

    monkeypatch.setattr(service, "_blob_has_upload", lambda _sid, role: False)
    monkeypatch.setattr(service, "_blob_has_pd_spec_workbook", lambda _sid: True)
    monkeypatch.setattr(service, "_read_pd_spec_workbook_bytes", lambda _sid: b"cached")

    status = service.get_step1_upload_status(study_id)
    assert status["pdSpec"]["uploaded"] is True
    assert status["pdSpec"]["fileName"] == "company_specs.xlsx"
    assert status["pdSpec"]["size"] == 2048
    assert status["pdSpec"]["blob"] == paths.pd_spec_workbook_blob(study_id)


def test_get_step1_upload_status_reflects_blob_presence(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "US-S"

    def fake_exists(*, blob_path: str, **_kwargs: object) -> bool:
        return blob_path.endswith("protocol.pdf")

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", fake_exists)

    monkeypatch.setattr(service, "_blob_has_pd_spec_workbook", lambda _sid: False)

    status = service.get_step1_upload_status(study_id)
    assert status["protocol"]["uploaded"] is True
    assert status["acrf"]["uploaded"] is False
    assert status["pdSpec"]["uploaded"] is False
    assert status["bothUploaded"] is False


def test_upload_step1_single_file_partial(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "PART-S"
    uploaded: list[str] = []

    def fake_upload(*, blob_path: str, **_kwargs: object) -> None:
        uploaded.append(blob_path)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: False)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", fake_upload)

    out = service.upload_step1_files(
        study_id,
        protocol_bytes=b"proto",
        acrf_bytes=None,
        protocol_file_name="My Protocol.pdf",
    )
    assert out["bothUploaded"] is False
    assert out["protocolFileName"] == "My Protocol.pdf"
    assert any("protocol.pdf" in path for path in uploaded)


def test_run_step1_extract_requires_uploads(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: False)

    with pytest.raises(UiApiError) as exc:
        service.run_step1_extract("NO-UP", extractor="both")
    assert exc.value.code == "UPLOAD_REQUIRED"
    assert exc.value.status_code == 409


def test_run_step1_extract_invalid_extractor(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    def fake_run_extract(**_kwargs: object) -> None:
        raise AssertionError("run_extract should not be called")

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)
    with pytest.raises(UiApiError) as exc:
        service.run_step1_extract("EX-S", extractor="bogus")
    assert exc.value.code == "VALIDATION_ERROR"


def test_sync_study_returns_report_and_statuses(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from pdcheck_factory import study_artifact_sync

    monkeypatch.setattr(
        study_artifact_sync,
        "sync_study",
        lambda *_a, **_k: study_artifact_sync.SyncReport(uploaded=2, downloaded=1, skipped=3, errors=0),
    )

    service = UiStepService(output_dir=tmp_path)
    study_id = "SYNC-01"
    proto = extraction_resolve.resolve_protocol_rendered_source_md(study_id, tmp_path)
    acrf = extraction_resolve.resolve_acrf_rendered_source_md(study_id, tmp_path)
    proto.parent.mkdir(parents=True, exist_ok=True)
    acrf.parent.mkdir(parents=True, exist_ok=True)
    _touch(proto)
    _touch(acrf)

    out = service.sync_study(study_id)
    assert out["studyId"] == study_id
    assert out["sync"]["uploaded"] == 2
    assert out["sync"]["downloaded"] == 1
    assert out["sync"]["skipped"] == 3
    assert out["sync"]["errors"] == 0
    assert "extract-inputs" in out["stepStatuses"]


def _seed_processing_artifacts(service: UiStepService, study_id: str, tmp_path: Path) -> None:
    proto = extraction_resolve.resolve_protocol_rendered_source_md(study_id, tmp_path)
    acrf = extraction_resolve.resolve_acrf_rendered_source_md(study_id, tmp_path)
    _touch(proto)
    _touch(acrf)
    pindex = paths.local_protocol_paragraph_index_json(study_id, tmp_path)
    _touch(pindex, '{"paragraphs": []}')
    sections = extraction_resolve.resolve_acrf_sections_toc_dir(study_id, tmp_path)
    sections.mkdir(parents=True, exist_ok=True)
    _touch(sections / "section_01.md", "# Section")
    summary = paths.local_acrf_summary_text_merged(study_id, tmp_path)
    _touch(summary, '{"datasets": []}')
    rules = paths.local_rules_parsed_json(study_id, tmp_path)
    _touch(rules, '{"rules": []}')
    dev = paths.local_deviations_parsed_json(study_id, tmp_path)
    _touch(dev, '{"deviations": []}')
    review = paths.local_deviations_review_state(study_id, tmp_path)
    _touch(review, '{"deviations": []}')


def test_run_step1_extract_skips_when_artifacts_exist(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "SKIP-EX"
    _seed_processing_artifacts(service, study_id, tmp_path)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)

    called = {"run_extract": False}

    def fake_run_extract(**_kwargs: object) -> None:
        called["run_extract"] = True

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)

    out = service.run_step1_extract(study_id, extractor="both", force=False)
    assert out.get("skipped") is True
    assert called["run_extract"] is False
    assert out["stepStatuses"]["extract-inputs"] == "done"


def test_run_step1_extract_force_runs_when_artifacts_exist(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "FORCE-EX"
    _seed_processing_artifacts(service, study_id, tmp_path)

    monkeypatch.setattr(blob_io, "blob_service_from_env", lambda: object())
    monkeypatch.setattr(blob_io, "container_from_env", lambda: "container")
    monkeypatch.setattr(blob_io, "blob_exists", lambda **_kwargs: True)
    monkeypatch.setattr(blob_io, "upload_blob_bytes", lambda **_kwargs: None)

    called = {"run_extract": False}

    def fake_run_extract(**_kwargs: object) -> None:
        called["run_extract"] = True

    from pdcheck_factory import cli as cli_mod

    monkeypatch.setattr(cli_mod, "run_extract", fake_run_extract)

    out = service.run_step1_extract(study_id, extractor="both", force=True)
    assert out.get("skipped") is not True
    assert called["run_extract"] is True


def test_run_step_skips_when_artifact_exists(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from pdcheck_factory import pipeline_v2

    service = UiStepService(output_dir=tmp_path)
    study_id = "SKIP-RULES"
    _seed_processing_artifacts(service, study_id, tmp_path)

    called = {"rules": False}

    def fake_rules(sid: str, output_dir: Path, *args, **kwargs):
        called["rules"] = True
        return {"rules": []}

    monkeypatch.setattr(pipeline_v2, "step3_extract_rules", fake_rules)

    out = service.run_step(study_id, "extract-rules", force=False)
    assert out.get("skipped") is True
    assert called["rules"] is False
    assert "skipped" in out["summary"].lower()


def test_upload_status_includes_processing_complete_flags(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service = UiStepService(output_dir=tmp_path)
    study_id = "PROC-COMPLETE"
    _seed_processing_artifacts(service, study_id, tmp_path)

    monkeypatch.setattr(service, "_blob_has_upload", lambda _sid, role: True)
    monkeypatch.setattr(service, "_blob_has_pd_spec_workbook", lambda _sid: False)

    status = service.get_step1_upload_status(study_id)
    assert status["processingCoreComplete"] is True
    assert status["processingComplete"] is True


def test_get_step7_review_sources_and_isolated_state(tmp_path: Path) -> None:
    from pdcheck_factory import review_sources

    service = UiStepService(output_dir=tmp_path)
    study_id = "REVIEW-SRC"

    parsed = paths.local_deviations_parsed_json(study_id, tmp_path)
    parsed.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        parsed,
        {
            "schema_version": "1.0.0",
            "study_id": study_id,
            "deviations": [
                {
                    "deviation_id": "dev-0001",
                    "rule_id": "rule-001",
                    "text": "Generated deviation text",
                    "paragraph_refs": ["p1"],
                    "status": "pending",
                    "entry_source": "extracted",
                }
            ],
        },
    )

    workbook_path = paths.local_pd_spec_workbook(study_id, tmp_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    from io import BytesIO as _BytesIO

    from openpyxl import Workbook as _Workbook

    from pdcheck_factory.pd_spec_export import PD_SPEC_HEADERS, PD_SPEC_SHEET_TITLE

    wb = _Workbook()
    ws = wb.active
    ws.title = PD_SPEC_SHEET_TITLE
    ws.append(PD_SPEC_HEADERS)
    ws.append(
        [
            "Eligibility Criteria",
            "Age",
            "Imported PD row",
            "",
            "Major",
            "Manual",
            "",
            "",
            "RAVE",
            "",
            "",
            "",
            "",
        ]
    )
    buffer = _BytesIO()
    wb.save(buffer)
    workbook_path.write_bytes(buffer.getvalue())

    preview = service.get_step7_review_sources(study_id)
    keys = {source["key"] for source in preview["sources"]}
    assert review_sources.REVIEW_SOURCE_GENERATED in keys
    assert review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC in keys

    generated = service.get_step7_deviations(study_id, review_source=review_sources.REVIEW_SOURCE_GENERATED)
    assert generated["rows"][0]["deviation_text"] == "Generated deviation text"

    imported = service.get_step7_deviations(study_id, review_source=review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC)
    assert imported["rows"][0]["entry_source"] == "imported_pd_spec"
    assert "Imported PD row" in imported["rows"][0]["deviation_text"]

    imported["rows"][0]["status"] = "accepted"
    state_path = review_sources.review_state_path(
        study_id, tmp_path, review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC
    )
    state_obj = read_json(state_path)
    state_obj["deviations"][0]["status"] = "accepted"
    service._persist_state(  # noqa: SLF001
        study_id,
        state_obj,
        service._audit(study_id, action="test", target_id="x", updated_rows=1),  # noqa: SLF001
        review_source=review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC,
    )

    generated_again = service.get_step7_deviations(study_id, review_source=review_sources.REVIEW_SOURCE_GENERATED)
    assert generated_again["rows"][0]["status"] == "pending"
    imported_again = service.get_step7_deviations(study_id, review_source=review_sources.REVIEW_SOURCE_IMPORTED_PD_SPEC)
    assert imported_again["rows"][0]["status"] == "accepted"


def test_step7_enrichment_detail_endpoint(tmp_path: Path) -> None:
    from pdcheck_factory import review_sources

    service = UiStepService(output_dir=tmp_path)
    study_id = "ENRICH-API"
    deviation_id = "dev-import-enrich-1"
    artifact_path = paths.local_protocol_enrichment_json(study_id, tmp_path, deviation_id)
    artifact_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_path.write_text(
        json.dumps(
            {
                "schema_version": "1.0.0",
                "study_id": study_id,
                "deviation_id": deviation_id,
                "generated_at": "2026-05-21T00:00:00Z",
                "enrichment_status": "ok",
                "enrichment_summary": "Done",
                "enrichment_errors": {},
                "merged": {
                    "original_deviation_text": "Original import",
                    "improved_deviation_text": "Improved",
                    "assumptions": ["a1"],
                    "caveats": ["c1"],
                    "programmability_risk": "low",
                },
            }
        ),
        encoding="utf-8",
    )

    detail = service.get_step7_enrichment_detail(study_id, deviation_id)
    assert detail["original_deviation_text"] == "Original import"
    assert detail["suggested_deviation_text"] == "Improved"
    assert detail["improved_deviation_text"] == "Improved"
    assert detail["assumptions"] == ["a1"]
    assert detail["caveats"] == ["c1"]

    enriched_path = review_sources.review_state_path(
        study_id, tmp_path, review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )
    enriched_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        enriched_path,
        {
            "schema_version": "1.1.0",
            "study_id": study_id,
            "generated_at": "2026-05-21T00:00:00Z",
            "pd_spec_import_mode": "enrich",
            "deviations": [
                {
                    "deviation_id": deviation_id,
                    "rule_id": "pd-spec-rule-1",
                    "text": "Improved",
                    "original_deviation_text": "Original import",
                    "paragraph_refs": ["p1"],
                    "status": "pending",
                    "dm_comment": "",
                    "pd_spec_import": {
                        "entry_source": "imported_pd_spec",
                        "enrichment_status": "ok",
                        "enrichment_summary": "Done",
                    },
                }
            ],
        },
    )
    service._write_upload_manifest(  # noqa: SLF001
        study_id, review_display_source=review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )
    rows = service.get_step7_deviations(
        study_id, review_source=review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )["rows"]
    assert rows[0]["original_deviation_text"] == "Original import"
    assert "assumptions" not in rows[0] or rows[0].get("assumptions") in (None, [])


def test_step7_enrichment_detail_fallback_from_row(tmp_path: Path) -> None:
    from pdcheck_factory import review_sources

    service = UiStepService(output_dir=tmp_path)
    study_id = "ENRICH-FALLBACK"
    deviation_id = "dev-import-fallback-1"
    enriched_path = review_sources.review_state_path(
        study_id, tmp_path, review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )
    enriched_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        enriched_path,
        {
            "schema_version": "1.1.0",
            "study_id": study_id,
            "generated_at": "2026-05-21T00:00:00Z",
            "pd_spec_import_mode": "enrich",
            "deviations": [
                {
                    "deviation_id": deviation_id,
                    "rule_id": "pd-spec-rule-1",
                    "text": "Imported text",
                    "original_deviation_text": "Imported text",
                    "suggested_deviation_text": "Suggested from row",
                    "paragraph_refs": ["p1"],
                    "status": "pending",
                    "dm_comment": "",
                    "pd_spec_import": {
                        "entry_source": "imported_pd_spec",
                        "enrichment_status": "ok",
                        "enrichment_summary": "Done",
                    },
                }
            ],
        },
    )

    detail = service.get_step7_enrichment_detail(study_id, deviation_id)
    assert detail["suggested_deviation_text"] == "Suggested from row"
    assert detail["original_deviation_text"] == "Imported text"


def test_enriched_patch_accept_with_text_promotes_suggestion(tmp_path: Path) -> None:
    from pdcheck_factory import review_sources

    service = UiStepService(output_dir=tmp_path)
    study_id = "ENRICH-ACCEPT"
    deviation_id = "dev-enriched-accept-1"
    enriched_path = review_sources.review_state_path(
        study_id, tmp_path, review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )
    enriched_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        enriched_path,
        {
            "schema_version": "1.1.0",
            "study_id": study_id,
            "generated_at": "2026-05-21T00:00:00Z",
            "pd_spec_import_mode": "enrich",
            "deviations": [
                {
                    "deviation_id": deviation_id,
                    "rule_id": "pd-spec-rule-1",
                    "text": "Imported text",
                    "original_deviation_text": "Imported text",
                    "suggested_deviation_text": "Suggested enriched text",
                    "paragraph_refs": ["p1"],
                    "status": "to_review",
                    "dm_comment": "",
                    "pd_spec_import": {
                        "entry_source": "imported_pd_spec",
                        "enrichment_status": "ok",
                        "enrichment_summary": "Done",
                    },
                }
            ],
        },
    )
    service._write_upload_manifest(  # noqa: SLF001
        study_id, review_display_source=review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC
    )

    updated = service.patch_step7_deviation_fields(
        study_id,
        deviation_id,
        {"status": "accepted", "text": "Suggested enriched text"},
        review_source=review_sources.REVIEW_SOURCE_ENRICHED_PD_SPEC,
    )
    assert updated["row"]["deviation_text"] == "Suggested enriched text"
    assert updated["row"]["status"] == "accepted"
    assert updated["row"]["original_deviation_text"] == "Imported text"
    assert updated["row"]["suggested_deviation_text"] == "Suggested enriched text"

    persisted = read_json(enriched_path)
    row = persisted["deviations"][0]
    assert row["text"] == "Suggested enriched text"
    assert row["original_deviation_text"] == "Imported text"
    assert row["suggested_deviation_text"] == "Suggested enriched text"
