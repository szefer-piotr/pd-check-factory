from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from pdcheck_factory import paths
from pdcheck_factory.cli import (
    run_protocol_sections_extract,
    run_step2_apply_review,
)
from pdcheck_factory.json_util import write_json
from pdcheck_factory.protocol_markdown import write_manifest
from pdcheck_factory.step2_review import export_step2_review_workbook


class CliAcrfSummaryContextTests(unittest.TestCase):
    def test_step1_respects_no_use_acrf_summary_flag(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            output_dir = Path(td) / "output"
            study_id = "study-x"
            manifest_path = paths.local_protocol_sections_manifest(study_id, output_dir)
            manifest = {
                "manifest_schema_version": "1.1.0",
                "study_id": study_id,
                "di_page_markers_stripped": True,
                "rollup_max_section_level": 1,
                "sections": [
                    {
                        "section_id": "sec:abc",
                        "section_path": ["Inclusion"],
                        "heading_level": 1,
                        "body_markdown": "Patient must be >= 18 years.",
                        "sentences": [{"id": "sec:abc#s1", "text": "Patient must be >= 18 years."}],
                    }
                ],
            }
            write_manifest(manifest_path, manifest)
            acrf_merged = paths.local_acrf_summary_merged(study_id, output_dir)
            acrf_merged.parent.mkdir(parents=True, exist_ok=True)
            acrf_merged.write_text(json.dumps({"dataset_index": []}), encoding="utf-8")

            with patch(
                "pdcheck_factory.cli._optional_acrf_summary_context",
                side_effect=AssertionError("summary loader should not be called"),
            ), patch(
                "pdcheck_factory.llm.extract_protocol_section_step1",
                return_value={
                    "schema_version": "2.0.0",
                    "study_id": study_id,
                    "generated_at": "2026-01-01T00:00:00+00:00",
                    "section_id": "sec:abc",
                    "section_path": ["Inclusion"],
                    "rules": [],
                },
            ):
                run_protocol_sections_extract(
                    study_id=study_id,
                    output_dir=output_dir,
                    upload=False,
                    all_sections=True,
                    section_id=[],
                    match_regex=None,
                    skip_section_id=[],
                    skip_regex=None,
                    include_acrf=True,
                    use_acrf_summary=False,
                )

    def test_step2_revalidation_receives_summary_context_by_default(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            output_dir = Path(td) / "output"
            study_id = "study-x"
            proto_md = paths.local_extraction_layout(study_id, "protocol", output_dir) / "rendered" / "source.md"
            proto_md.parent.mkdir(parents=True, exist_ok=True)
            proto_md.write_text("# Protocol\nText", encoding="utf-8")

            merged = {
                "schema_version": "2.1.0",
                "study_id": study_id,
                "generated_at": "2026-01-01T00:00:00+00:00",
                "rules": [
                    {
                        "rule_id": "rule-001",
                        "title": "Rule",
                        "atomic_requirement": "Req",
                        "sentence_refs": ["sec:abc#s1"],
                        "source_section_ids": ["sec:abc"],
                        "source_section_paths": [["Inclusion"]],
                        "candidate_deviations": [
                            {
                                "deviation_id": "dev-001a",
                                "scenario_description": "Scenario",
                                "example_violation_narrative": "Example",
                                "sentence_refs": ["sec:abc#s1"],
                                "programmable": True,
                                "source_section_ids": ["sec:abc"],
                                "source_section_paths": [["Inclusion"]],
                            }
                        ],
                    }
                ],
            }
            step2_path = paths.local_protocol_sections_step2_merged(study_id, output_dir)
            write_json(step2_path, merged)

            workbook_path = output_dir / "review.xlsx"
            export_step2_review_workbook(step2_json_path=step2_path, workbook_path=workbook_path)
            wb = load_workbook(workbook_path)
            ws = wb.active
            ws["L2"] = "to_review"
            ws["M2"] = "Please refine."
            wb.save(workbook_path)

            acrf_merged = paths.local_acrf_summary_merged(study_id, output_dir)
            acrf_merged.parent.mkdir(parents=True, exist_ok=True)
            acrf_merged.write_text(json.dumps({"dataset_index": [{"dataset_name": "DM"}]}), encoding="utf-8")

            with patch(
                "pdcheck_factory.llm.revalidate_deviation_with_dm_feedback",
                return_value=[
                    {
                        "deviation_id": "dev-001a",
                        "scenario_description": "Scenario updated",
                        "example_violation_narrative": "Example updated",
                        "sentence_refs": ["sec:abc#s1"],
                        "programmable": True,
                        "source_section_ids": ["sec:abc"],
                        "source_section_paths": [["Inclusion"]],
                    }
                ],
            ) as mock_revalidate:
                run_step2_apply_review(
                    study_id=study_id,
                    output_dir=output_dir,
                    workbook=workbook_path,
                    context_mode="full_protocol",
                    strict=True,
                    upload=False,
                )

            self.assertEqual(mock_revalidate.call_count, 1)
            self.assertTrue(mock_revalidate.call_args.kwargs["acrf_summary_context"])


if __name__ == "__main__":
    unittest.main()
